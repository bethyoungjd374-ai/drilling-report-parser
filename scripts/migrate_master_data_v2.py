from __future__ import annotations

import argparse
import hashlib
import json
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from drilling_report_parser.master_data_service import normalize_alias  # noqa: E402
from drilling_report_parser.mysql_database import _connect, initialize_database  # noqa: E402
from drilling_report_parser.report_normalization_service import synchronize_saved_report  # noqa: E402


DEFAULT_PROJECT_CONFIG = ROOT / "outputs" / "project_team_config.json"
ALIAS_REVIEW_PAIRS = (
    ("well", "PCN-041", "PCNA-041"),
    ("well", "SHSG-160", "SHSH-160"),
    ("block", "WAYRA", "YURALPA"),
    ("block", "SHUSHUFINDI", "SSFD"),
)

# These aliases are backed by the cited monthly source workbooks, not by fuzzy
# string similarity.  JVNC-024 appears in the daily PDF while the same
# SINOPEC220 / PUCUNA job is JVN-024 in both appendix 4 and appendix 6.
CONFIRMED_SOURCE_ALIASES = (
    ("well", "JVNC-024", "JVN-024", "2026-06-monthly-table4-table6"),
)

WELL_PROFILE_CODES = {
    "VERTICAL": "VERTICAL",
    "DIRECTIONAL": "DIRECTIONAL",
    "HORIZONTAL": "HORIZONTAL",
    "SIDETRACK": "SIDETRACK",
    "直井": "VERTICAL",
    "定向井": "DIRECTIONAL",
    "水平井": "HORIZONTAL",
    "侧钻井": "SIDETRACK",
}


def _normalize_well_profile_code(value: object) -> str:
    raw = str(value or "").strip()
    return WELL_PROFILE_CODES.get(raw, WELL_PROFILE_CODES.get(raw.upper(), ""))


def main() -> int:
    parser = argparse.ArgumentParser(description="Seed and backfill the V2 master-data model.")
    parser.add_argument("--project-config", default=str(DEFAULT_PROJECT_CONFIG))
    parser.add_argument("--table4", default="")
    parser.add_argument("--table5", default="")
    parser.add_argument("--table6", default="")
    parser.add_argument("--batch-code", default="master-data-v2-2026")
    parser.add_argument("--actor", default="migration")
    parser.add_argument("--rollback", default="", metavar="BATCH_CODE")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    initialize_database()
    with _connect() as connection:
        try:
            with connection.cursor() as cursor:
                if args.rollback:
                    summary = rollback_batch(cursor, args.rollback, actor=args.actor)
                else:
                    summary = migrate(
                        cursor,
                        project_config=Path(args.project_config),
                        workbook_paths={
                            "table4": Path(args.table4) if args.table4 else None,
                            "table5": Path(args.table5) if args.table5 else None,
                            "table6": Path(args.table6) if args.table6 else None,
                        },
                        batch_code=args.batch_code,
                        actor=args.actor,
                    )
            if args.dry_run:
                connection.rollback()
                summary["dry_run"] = True
            else:
                connection.commit()
        except Exception:
            connection.rollback()
            raise
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


def migrate(
    cursor: Any,
    *,
    project_config: Path,
    workbook_paths: dict[str, Path | None],
    batch_code: str,
    actor: str,
) -> dict[str, Any]:
    batch_id = _start_batch(cursor, batch_code, project_config, actor)
    counters: dict[str, int] = {
        "organizations": 0,
        "blocks": 0,
        "rig_models": 0,
        "rigs": 0,
        "wells": 0,
        "contracts": 0,
        "projects": 0,
        "assignments": 0,
        "aliases": 0,
        "reports": 0,
        "issues": 0,
    }
    config = _load_json(project_config)

    for team in config.get("teams", []) if isinstance(config, dict) else []:
        raw_name = str(team.get("name", "") or "")
        if raw_name:
            _, created = _ensure_rig(cursor, raw_name, batch_id, f"config:team:{raw_name}", actor)
            counters["rigs"] += int(created)
            counters["aliases"] += 1

    for project in config.get("projects", []) if isinstance(config, dict) else []:
        project_id, created_contract, created_project = _ensure_project(
            cursor, project, batch_id, actor
        )
        counters["contracts"] += int(created_contract)
        counters["projects"] += int(created_project)
        project_start = str(project.get("start_date", "") or "") or "1900-01-01"
        project_end = _exclusive_end(str(project.get("end_date", "") or ""))
        for rig_item in project.get("rigs", []) if isinstance(project.get("rigs"), list) else []:
            raw_rig = str(rig_item.get("rig", "") or "")
            if not raw_rig:
                continue
            rig_id, created = _ensure_rig(cursor, raw_rig, batch_id, f"config:project:{project_id}:rig:{raw_rig}", actor)
            counters["rigs"] += int(created)
            cursor.execute("SELECT team_id FROM md_rig WHERE id=%s", (rig_id,))
            team_id = int((cursor.fetchone() or {}).get("team_id") or 0)
            if not team_id:
                raise RuntimeError(f"设备 {raw_rig} 未关联正式队伍主数据。")
            valid_from = str(rig_item.get("start_date", "") or "") or project_start
            valid_to = _exclusive_end(str(rig_item.get("end_date", "") or "")) or project_end
            counters["assignments"] += int(_ensure_assignment(
                cursor,
                "rel_project_team_assignment",
                {
                    "project_id": project_id,
                    "team_id": team_id,
                    "valid_from": f"{valid_from} 00:00:00",
                    "valid_to": f"{valid_to} 00:00:00" if valid_to else None,
                    "service_discipline": "",
                    "priority": 100,
                },
                ("project_id", "team_id", "valid_from"),
                batch_id,
                f"config:project:{project_id}:team:{team_id}:{valid_from}",
                actor,
            ))
            for raw_well in rig_item.get("wells", []) if isinstance(rig_item.get("wells"), list) else []:
                well_id, created_well = _ensure_well(
                    cursor,
                    raw_well,
                    batch_id,
                    f"config:project:{project_id}:well:{raw_well}",
                    actor,
                )
                counters["wells"] += int(created_well)
                counters["assignments"] += int(_ensure_assignment(
                    cursor,
                    "rel_project_well_scope",
                    {
                        "project_id": project_id,
                        "well_id": well_id,
                        "job_type": "",
                        "valid_from": f"{valid_from} 00:00:00",
                        "valid_to": f"{valid_to} 00:00:00" if valid_to else None,
                    },
                    ("project_id", "well_id", "job_type", "valid_from"),
                    batch_id,
                    f"config:project:{project_id}:well:{well_id}:{valid_from}",
                    actor,
                ))

    for table_kind, path in workbook_paths.items():
        if not path:
            continue
        if not path.exists():
            raise FileNotFoundError(path)
        for item in _workbook_master_rows(path, table_kind):
            org_id, created = _ensure_simple_master(
                cursor,
                "md_organization",
                "organization_code",
                item["organization"],
                {"organization_name": item["organization"], "organization_type": "regional_company"},
                batch_id,
                f"{path}:organization:{item['organization']}",
                "organization",
                actor,
            )
            counters["organizations"] += int(created)
            block_id, created = _ensure_simple_master(
                cursor,
                "md_block",
                "block_code",
                item["block"],
                {"block_name": item["block"], "country": item["country"]},
                batch_id,
                f"{path}:block:{item['block']}",
                "block",
                actor,
            )
            counters["blocks"] += int(created)
            model_id, created = _ensure_simple_master(
                cursor,
                "md_rig_model",
                "model_code",
                item["rig_model"],
                {"model_name": item["rig_model"], "equipment_type": item["discipline"]},
                batch_id,
                f"{path}:rig-model:{item['rig_model']}",
                "rig_model",
                actor,
            )
            counters["rig_models"] += int(created)
            _, created = _ensure_rig(
                cursor,
                item["rig"],
                batch_id,
                f"{path}:rig:{item['rig']}",
                actor,
                rig_model_id=model_id,
                owner_organization_id=org_id,
                rig_type=item["discipline"],
            )
            counters["rigs"] += int(created)
            _, created_well = _ensure_well(
                cursor,
                item["wellbore"],
                batch_id,
                f"{path}:{item['sheet']}:{item['row_no']}",
                actor,
                block_id=block_id,
                well_profile=item.get("well_profile", ""),
            )
            counters["wells"] += int(created_well)

    counters["aliases"] += _ensure_confirmed_source_aliases(
        cursor,
        batch_id=batch_id,
        actor=actor,
    )

    cursor.execute("SELECT DISTINCT rig, wellbore FROM dpr_report_record")
    for row in cursor.fetchall():
        raw_rig = str(row.get("rig", "") or "")
        raw_well = str(row.get("wellbore", "") or "")
        if raw_rig:
            _, created = _ensure_rig(cursor, raw_rig, batch_id, f"mysql:rig:{raw_rig}", actor)
            counters["rigs"] += int(created)
        if raw_well:
            well_id = _existing_well_id(cursor, raw_well)
            if well_id:
                _ensure_alias(cursor, "well", raw_well, well_id, batch_id, f"mysql:well:{raw_well}", actor)

    for entity_type, left, right in ALIAS_REVIEW_PAIRS:
        issue_key = f"migration:alias-review:{entity_type}:{left}:{right}"
        cursor.execute(
            """
            INSERT INTO dq_issue
              (issue_key, issue_type, severity, entity_type, entity_id, details_json,
               status, created_by, updated_by)
            VALUES (%s,'ALIAS_REVIEW','warning',%s,'',%s,'OPEN',%s,%s)
            ON DUPLICATE KEY UPDATE details_json=VALUES(details_json), status='OPEN',
              updated_by=VALUES(updated_by), version=version+1
            """,
            (issue_key, entity_type, json.dumps({"left": left, "right": right}, ensure_ascii=False), actor, actor),
        )
        counters["issues"] += 1

    cursor.execute(
        """
        SELECT r.record_id, r.report_type, f.fields_json
        FROM dpr_report_record r
        JOIN dpr_report_field f ON f.record_id=r.record_id
        ORDER BY r.report_date, r.record_id
        """
    )
    dpr_report_row = cursor.fetchall()
    for report in dpr_report_row:
        record_id = str(report["record_id"])
        fields = _json_value(report.get("fields_json"), {})
        cursor.execute(
            "SELECT module_name,row_json FROM dpr_report_row WHERE record_id=%s ORDER BY module_name,row_no",
            (record_id,),
        )
        payload: dict[str, Any] = {"report_fields": fields}
        for row in cursor.fetchall():
            payload.setdefault(str(row.get("module_name", "") or ""), []).append(_json_value(row.get("row_json"), {}))
        operations = payload.get("operations", [])
        old_refs = _report_refs(cursor, record_id)
        result = synchronize_saved_report(
            cursor,
            record_id=record_id,
            report_type=str(report["report_type"]),
            fields=fields,
            operations=operations,
            payload=payload,
            actor=actor,
        )
        _record_entry(
            cursor,
            batch_id,
            f"mysql:report:{record_id}",
            "report_backfill",
            record_id,
            old_refs,
            result,
        )
        counters["reports"] += 1

    cursor.execute(
        "UPDATE migration_batch SET batch_status='COMPLETED', summary_json=%s, completed_at=NOW() WHERE id=%s",
        (json.dumps(counters, ensure_ascii=False), batch_id),
    )
    return {"batch_id": batch_id, "batch_code": batch_code, **counters}


def rollback_batch(cursor: Any, batch_code: str, *, actor: str) -> dict[str, Any]:
    cursor.execute("SELECT id, batch_status FROM migration_batch WHERE batch_code=%s FOR UPDATE", (batch_code,))
    batch = cursor.fetchone()
    if not batch:
        raise KeyError(batch_code)
    batch_id = int(batch["id"])
    cursor.execute(
        "SELECT * FROM migration_entry WHERE batch_id=%s ORDER BY id DESC",
        (batch_id,),
    )
    restored = 0
    skipped = 0
    for entry in cursor.fetchall():
        entity_type = str(entry.get("entity_type", "") or "")
        old_value = _json_value(entry.get("old_value_json"), {})
        new_value = _json_value(entry.get("new_value_json"), {})
        try:
            if entity_type == "report_backfill" and old_value:
                cursor.execute(
                    """
                    UPDATE dpr_report_record
                    SET rig_id=%s, well_id=%s, project_id=%s, job_id=%s,
                        master_match_status=%s, master_match_message=%s
                    WHERE record_id=%s
                    """,
                    (
                        old_value.get("rig_id"),
                        old_value.get("well_id"),
                        old_value.get("project_id"),
                        old_value.get("job_id"),
                        old_value.get("master_match_status", ""),
                        old_value.get("master_match_message", ""),
                        entry.get("entity_id"),
                    ),
                )
                cursor.execute("DELETE FROM dpr_report WHERE record_id=%s", (entry.get("entity_id"),))
                restored += 1
            elif new_value.get("created") and new_value.get("table") and new_value.get("id"):
                table = str(new_value["table"])
                if table not in _rollback_tables():
                    skipped += 1
                    continue
                cursor.execute(f"DELETE FROM {table} WHERE id=%s", (int(new_value["id"]),))
                restored += int(cursor.rowcount == 1)
            else:
                skipped += 1
        except Exception:
            skipped += 1
    cursor.execute(
        "UPDATE migration_batch SET batch_status='ROLLED_BACK', rolled_back_at=NOW(), "
        "summary_json=%s WHERE id=%s",
        (json.dumps({"restored": restored, "skipped": skipped, "actor": actor}, ensure_ascii=False), batch_id),
    )
    return {"batch_id": batch_id, "batch_code": batch_code, "restored": restored, "skipped": skipped}


def _ensure_project(cursor: Any, project: dict[str, Any], batch_id: int, actor: str) -> tuple[int, bool, bool]:
    contract_no = str(project.get("contract_no", "") or project.get("project_name", "") or "").strip()
    project_code = str(project.get("project_name", "") or contract_no).strip()
    contract_id, created_contract = _ensure_simple_master(
        cursor,
        "md_contract",
        "contract_no",
        contract_no,
        {
            "contract_name": str(project.get("project_name", "") or contract_no),
            "valid_from": str(project.get("start_date", "") or "") or None,
            "valid_to": str(project.get("end_date", "") or "") or None,
        },
        batch_id,
        f"config:contract:{contract_no}",
        "contract",
        actor,
    )
    project_id, created_project = _ensure_simple_master(
        cursor,
        "md_project",
        "project_code",
        project_code,
        {
            "project_name": str(project.get("project_name", "") or contract_no),
            "contract_id": contract_id,
            "valid_from": str(project.get("start_date", "") or "") or None,
            "valid_to": str(project.get("end_date", "") or "") or None,
        },
        batch_id,
        f"config:project:{project_code}",
        "project",
        actor,
    )
    return project_id, created_contract, created_project


def _ensure_rig(
    cursor: Any,
    raw_name: object,
    batch_id: int,
    locator: str,
    actor: str,
    *,
    rig_model_id: int | None = None,
    owner_organization_id: int | None = None,
    rig_type: str = "",
) -> tuple[int, bool]:
    canonical = normalize_alias("rig", raw_name)
    rig_id, created = _ensure_simple_master(
        cursor,
        "md_rig",
        "rig_code",
        canonical,
        {
            "rig_name": canonical,
            "rig_model_id": rig_model_id,
            "owner_organization_id": owner_organization_id,
            "rig_type": rig_type,
        },
        batch_id,
        locator,
        "rig",
        actor,
    )
    team_id, _ = _ensure_simple_master(
        cursor,
        "md_team",
        "team_code",
        canonical,
        {
            "team_name": canonical,
            "team_type_code": "WORKOVER" if rig_type == "workover" else "DRILLING",
            "company_id": owner_organization_id,
        },
        batch_id,
        f"{locator}:team",
        "team",
        actor,
    )
    cursor.execute("UPDATE md_rig SET team_id=%s,updated_by=%s WHERE id=%s", (team_id, actor, rig_id))
    _ensure_alias(cursor, "rig", raw_name, rig_id, batch_id, locator, actor)
    return rig_id, created


def _ensure_well(
    cursor: Any,
    raw_code: object,
    batch_id: int,
    locator: str,
    actor: str,
    *,
    block_id: int | None = None,
    well_type: str = "",
    well_profile: str = "",
) -> tuple[int, bool]:
    code = normalize_alias("well", raw_code)
    profile_code = _normalize_well_profile_code(well_profile)
    existing_well_id = _existing_well_id(cursor, raw_code)
    if existing_well_id:
        cursor.execute(
            "UPDATE md_well SET block_id=COALESCE(block_id,%s), "
            "well_type_code=IF(well_type_code='',%s,well_type_code), "
            "version=version+IF(well_profile_code='' AND %s<>'',1,0), "
            "change_reason=IF(well_profile_code='' AND %s<>'','月报表4井型明确字段迁移',change_reason), "
            "well_profile_code=IF(well_profile_code='',%s,well_profile_code), updated_by=%s "
            "WHERE id=%s",
            (block_id, well_type or "DEVELOPMENT", profile_code, profile_code, profile_code, actor, existing_well_id),
        )
        _ensure_alias(cursor, "well", raw_code, existing_well_id, batch_id, locator, actor)
        return existing_well_id, False
    well_id, created_well = _ensure_simple_master(
        cursor,
        "md_well",
        "well_code",
        code,
        {
            "well_name": code,
            "block_id": block_id,
            "well_type_code": well_type or "DEVELOPMENT",
            "well_profile_code": profile_code,
        },
        batch_id,
        f"{locator}:well",
        "well",
        actor,
    )
    if created_well and profile_code:
        cursor.execute(
            "UPDATE md_well SET change_reason='月报表4井型明确字段迁移' WHERE id=%s",
            (well_id,),
        )
    _ensure_alias(cursor, "well", raw_code, well_id, batch_id, locator, actor)
    return well_id, created_well


def _existing_well_id(cursor: Any, raw_code: object) -> int | None:
    code = normalize_alias("well", raw_code)
    if not code:
        return None
    cursor.execute(
        "SELECT entity_id FROM md_alias "
        "WHERE entity_type='well' AND normalized_alias=%s AND status='active' "
        "AND confirmation_status='confirmed' "
        "ORDER BY source_system='manual' DESC, id LIMIT 1",
        (code,),
    )
    alias_row = cursor.fetchone()
    if alias_row:
        return int(alias_row["entity_id"])
    cursor.execute("SELECT id FROM md_well WHERE well_code=%s AND status='active' LIMIT 1", (code,))
    well_row = cursor.fetchone()
    return int(well_row["id"]) if well_row else None


def _ensure_confirmed_source_aliases(cursor: Any, *, batch_id: int, actor: str) -> int:
    applied = 0
    for entity_type, alias_value, canonical_code, source_system in CONFIRMED_SOURCE_ALIASES:
        if entity_type != "well":
            continue
        cursor.execute("SELECT id FROM md_well WHERE well_code=%s AND status='active' LIMIT 1", (canonical_code,))
        row = cursor.fetchone()
        if not row:
            continue
        entity_id = int(row["id"])
        normalized_alias = normalize_alias(entity_type, alias_value)
        cursor.execute(
            "UPDATE md_alias SET entity_id=%s, confirmation_status='confirmed', status='active', "
            "change_reason='月报表4/表6与日报交叉确认', updated_by=%s, version=version+1 "
            "WHERE entity_type=%s AND normalized_alias=%s AND source_system<>'manual' AND entity_id<>%s",
            (entity_id, actor, entity_type, normalized_alias, entity_id),
        )
        cursor.execute(
            "INSERT INTO md_alias "
            "(entity_type,source_system,alias_value,normalized_alias,entity_id,confirmation_status,status,change_reason,created_by,updated_by) "
            "VALUES (%s,%s,%s,%s,%s,'confirmed','active','月报表4/表6与日报交叉确认',%s,%s) "
            "ON DUPLICATE KEY UPDATE entity_id=VALUES(entity_id),alias_value=VALUES(alias_value),"
            "confirmation_status='confirmed',status='active',change_reason=VALUES(change_reason),updated_by=VALUES(updated_by)",
            (entity_type, source_system, alias_value, normalized_alias, entity_id, actor, actor),
        )
        _record_entry(
            cursor,
            batch_id,
            f"source-alias:{source_system}:{entity_type}:{alias_value}",
            "alias",
            f"{entity_type}:{alias_value}",
            {},
            {"entity_id": entity_id, "canonical_code": canonical_code, "source_system": source_system},
        )
        applied += 1
    return applied


def _ensure_simple_master(
    cursor: Any,
    table: str,
    code_field: str,
    code: object,
    values: dict[str, Any],
    batch_id: int,
    locator: str,
    entity_type: str,
    actor: str,
) -> tuple[int, bool]:
    clean_code = str(code or "").strip()
    if not clean_code:
        raise ValueError(f"Missing {code_field} at {locator}")
    cursor.execute(f"SELECT * FROM {table} WHERE {code_field}=%s", (clean_code,))
    existing = cursor.fetchone()
    if existing:
        updates = {
            field: value
            for field, value in values.items()
            if value not in (None, "") and existing.get(field) in (None, "")
        }
        if updates:
            cursor.execute(
                f"UPDATE {table} SET {', '.join(f'{field}=%s' for field in updates)}, "
                "updated_by=%s, version=version+1 WHERE id=%s",
                [*updates.values(), actor, existing["id"]],
            )
        return int(existing["id"]), False
    columns = [code_field, *values.keys(), "status", "change_reason", "created_by", "updated_by"]
    data = [clean_code, *values.values(), "active", "2026首期主数据迁移", actor, actor]
    cursor.execute(
        f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({', '.join(['%s'] * len(columns))})",
        data,
    )
    entity_id = int(cursor.lastrowid)
    _record_entry(
        cursor,
        batch_id,
        locator,
        entity_type,
        str(entity_id),
        {},
        {"created": True, "table": table, "id": entity_id},
    )
    return entity_id, True


def _ensure_alias(
    cursor: Any,
    entity_type: str,
    raw_value: object,
    entity_id: int,
    batch_id: int,
    locator: str,
    actor: str,
) -> None:
    raw = str(raw_value or "").strip()
    normalized = normalize_alias(entity_type, raw)
    cursor.execute(
        """
        INSERT INTO md_alias
          (entity_type, source_system, alias_value, normalized_alias, entity_id,
           confirmation_status, status, change_reason, created_by, updated_by)
        VALUES (%s,'migration',%s,%s,%s,'confirmed','active','2026首期迁移',%s,%s)
        ON DUPLICATE KEY UPDATE
          entity_id=VALUES(entity_id), alias_value=VALUES(alias_value),
          confirmation_status='confirmed', status='active', updated_by=VALUES(updated_by),
          version=version+1
        """,
        (entity_type, raw, normalized, entity_id, actor, actor),
    )


def _ensure_assignment(
    cursor: Any,
    table: str,
    values: dict[str, Any],
    key_fields: tuple[str, ...],
    batch_id: int,
    locator: str,
    actor: str,
) -> bool:
    cursor.execute(
        f"SELECT id FROM {table} WHERE {' AND '.join(f'{field}<=>%s' for field in key_fields)}",
        [values.get(field) for field in key_fields],
    )
    row = cursor.fetchone()
    if row:
        return False
    columns = [*values.keys(), "status", "change_reason", "created_by", "updated_by"]
    cursor.execute(
        f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({', '.join(['%s'] * len(columns))})",
        [*values.values(), "active", "2026首期项目关系迁移", actor, actor],
    )
    entity_id = int(cursor.lastrowid)
    _record_entry(
        cursor,
        batch_id,
        locator,
        "assignment",
        str(entity_id),
        {},
        {"created": True, "table": table, "id": entity_id},
    )
    return True


def _workbook_master_rows(path: Path, kind: str) -> Iterable[dict[str, Any]]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    sheet = next((item for item in workbook.worksheets if item.max_row > 1), workbook.active)
    if kind in {"table4", "table5"}:
        start_row = 4
        columns = {
            "rig": 2,
            "country": 3,
            "organization": 4,
            "block": 5,
            "rig_model": 6,
            "wellbore": 7,
            ("well_profile" if kind == "table4" else "source_well_kind"): 8,
        }
        discipline = "drilling" if kind == "table4" else "workover"
    else:
        start_row = 6
        columns = {"rig": 2, "wellbore": 3, "discipline": 4, "country": 5, "organization": 6, "block": 7, "rig_model": 8}
        discipline = ""
    for row_no in range(start_row, sheet.max_row + 1):
        sequence = sheet.cell(row_no, 1).value
        if not isinstance(sequence, (int, float)):
            continue
        values = {key: str(sheet.cell(row_no, column).value or "").strip() for key, column in columns.items()}
        if not values.get("rig") or not values.get("wellbore"):
            continue
        values["discipline"] = discipline or values.get("discipline", "")
        if "well_profile" in values:
            values["well_profile"] = _normalize_well_profile_code(values["well_profile"])
        values["sheet"] = sheet.title
        values["row_no"] = row_no
        yield values


def _start_batch(cursor: Any, batch_code: str, source_path: Path, actor: str) -> int:
    cursor.execute("SELECT id, batch_status FROM migration_batch WHERE batch_code=%s FOR UPDATE", (batch_code,))
    existing = cursor.fetchone()
    if existing:
        cursor.execute(
            "UPDATE migration_batch SET batch_status='RUNNING', started_at=NOW(), "
            "completed_at=NULL, rolled_back_at=NULL, created_by=%s WHERE id=%s",
            (actor, existing["id"]),
        )
        return int(existing["id"])
    cursor.execute(
        "INSERT INTO migration_batch "
        "(batch_code, source_type, source_path, batch_status, started_at, created_by) "
        "VALUES (%s,'master_data_v2',%s,'RUNNING',NOW(),%s)",
        (batch_code, str(source_path), actor),
    )
    return int(cursor.lastrowid)


def _record_entry(
    cursor: Any,
    batch_id: int,
    locator: str,
    entity_type: str,
    entity_id: str,
    old_value: dict[str, Any],
    new_value: dict[str, Any],
) -> None:
    locator_hash = hashlib.sha256(locator.encode("utf-8")).hexdigest()
    cursor.execute(
        """
        INSERT INTO migration_entry
          (batch_id, source_locator, source_locator_hash, entity_type, entity_id,
           old_value_json, new_value_json, entry_status)
        VALUES (%s,%s,%s,%s,%s,%s,%s,'APPLIED')
        ON DUPLICATE KEY UPDATE
          entity_id=VALUES(entity_id), old_value_json=VALUES(old_value_json),
          new_value_json=VALUES(new_value_json), entry_status='APPLIED'
        """,
        (
            batch_id,
            locator,
            locator_hash,
            entity_type,
            entity_id,
            json.dumps(old_value, ensure_ascii=False, default=str),
            json.dumps(new_value, ensure_ascii=False, default=str),
        ),
    )


def _report_refs(cursor: Any, record_id: str) -> dict[str, Any]:
    cursor.execute(
        "SELECT rig_id, well_id, project_id, job_id, master_match_status, master_match_message "
        "FROM dpr_report_record WHERE record_id=%s",
        (record_id,),
    )
    return dict(cursor.fetchone() or {})


def _exclusive_end(value: str) -> str:
    if not value:
        return ""
    parsed = date.fromisoformat(value)
    return (parsed + timedelta(days=1)).isoformat()


def _json_value(value: object, default: Any) -> Any:
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(str(value or ""))
    except json.JSONDecodeError:
        return default


def _load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    return data if isinstance(data, dict) else {}


def _rollback_tables() -> set[str]:
    return {
        "rel_job_rig_assignment",
        "rel_project_well_scope",
        "rel_project_team_assignment",
        "biz_job",
        "md_alias",
        "md_project",
        "md_contract",
        "md_well",
        "md_rig",
        "md_rig_model",
        "md_block",
        "md_organization",
    }


if __name__ == "__main__":
    raise SystemExit(main())
