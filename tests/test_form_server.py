from __future__ import annotations

import hashlib
import http.client
import json
import os
import tempfile
import threading
import unittest
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any
from unittest.mock import Mock, patch

from pypdf import PdfReader, PdfWriter
from openpyxl import load_workbook

from drilling_report_parser import excel_database, form_server
from drilling_report_parser.excel_database import load_report_payload
from tests.test_pdf_report_parser import sample_pdf


def _yellow_filled_cells(worksheet) -> list[str]:
    matches: list[str] = []
    for row in worksheet.iter_rows():
        for cell in row:
            color = str(getattr(cell.fill.fgColor, "rgb", "") or "").upper()
            indexed = getattr(cell.fill.fgColor, "indexed", None)
            if color.endswith("FFFF00") or indexed == 6:
                matches.append(cell.coordinate)
    return matches


def test_operational_parameter_values_are_normalized_without_units() -> None:
    payload = {"report_fields": {
        "lastCasingSize": "13.375in", "lastCasingDepth": "7,610ft",
        "nextCasingSize": "9.625in", "nextCasingDepth": "11,720ft",
        "formTestEmw": "FIT / 13.70 ppg", "lastBopPressTest": "06/06/2026",
        "stringWeightUpDown": "325.0 230.0 kip/", "torqueOnBottom": "27,000.0 ft-lbf",
    }}
    form_server._normalize_payload_values(payload)
    fields = payload["report_fields"]
    assert fields["lastCasingSize"] == "13.375"
    assert fields["lastCasingDepth"] == "7610"
    assert fields["nextCasingSize"] == "9.625"
    assert fields["nextCasingDepth"] == "11720"
    assert fields["formTestType"] == "FIT"
    assert fields["formTestEmw"] == "13.7"
    assert fields["lastBopPressTest"] == "2026-06-06"
    assert fields["stringWeightUp"] == "325"
    assert fields["stringWeightDown"] == "230"
    assert fields["torqueOnBottom"] == "27000"


def test_admin_project_form_exposes_type_and_editable_npt_allowance() -> None:
    source = Path(__file__).parents[1].joinpath("web_form", "admin.js").read_text(encoding="utf-8")

    assert '["project_type", "项目类型", "project-type"]' in source
    assert '["npt_allowance_hours", "允许 NPT（h）", "number"]' in source
    assert "{ drilling: 5, completion: 5, workover: 12 }" in source
    assert "data-admin-new-project" in source


def test_daily_report_forms_expose_all_pdf_basic_information_fields() -> None:
    html = Path(__file__).parents[1].joinpath("web_form", "index.html").read_text(encoding="utf-8")
    drilling = html.split('<form id="reportForm"', 1)[1].split("</form>", 1)[0]
    completion = html.split('<form id="completionReportForm"', 1)[1].split("</form>", 1)[0]
    workover = html.split('<form id="workoverReportForm"', 1)[1].split("</form>", 1)[0]

    for name in (
        "wellboreNo", "dfs", "groundElev", "afeMdDays", "dailyCost", "cumulativeCost",
        "afeCost", "avgRopSlide", "avgRopRot", "supervisor1", "supervisor2", "engineer",
        "pamEngineer", "geologist", "totalPersonnel",
    ):
        assert f'name="{name}"' in drilling
    for name in (
        "completionNo", "wellboreNo", "rigContractName", "groundElev", "dol", "dfs",
        "dailyCost", "cumulativeCost", "afeCost", "supervisor1", "supervisor2", "engineer",
        "pamEngineer", "geologist", "totalPersonnel",
    ):
        assert f'name="{name}"' in completion
    for name in (
        "workoverNo", "wellboreNo", "rigContractName", "groundElev", "dol", "dfs",
        "dailyCost", "cumulativeCost", "afeCost", "supervisor1", "supervisor2", "engineer",
        "pamEngineer", "geologist", "totalPersonnel",
    ):
        assert f'name="{name}"' in workover


def test_blank_bop_test_date_does_not_render_native_date_placeholder() -> None:
    root = Path(__file__).parents[1]
    html = root.joinpath("web_form", "index.html").read_text(encoding="utf-8")
    script = root.joinpath("web_form", "app.js").read_text(encoding="utf-8")
    styles = root.joinpath("web_form", "styles.css").read_text(encoding="utf-8")

    assert 'name="lastBopPressTest" type="date" data-blank-date' in html
    assert 'function syncBlankDateDisplay(input)' in script
    assert 'input.classList.toggle("blank-date-value", !input.value)' in script
    assert '.blank-date-value::-webkit-datetime-edit' in styles


def test_report_content_language_is_global_and_persistent_across_modules() -> None:
    script = Path(__file__).parents[1].joinpath("web_form", "app.js").read_text(encoding="utf-8")

    assert 'const REPORT_CONTENT_LANGUAGE_STORAGE_KEY = "drillingReportContentLanguage"' in script
    assert 'localStorage.setItem(REPORT_CONTENT_LANGUAGE_STORAGE_KEY, language)' in script
    assert 'const selectedLanguage = globalReportContentLanguage' in script
    assert 'selectedLanguage: globalReportContentLanguage' in script
    assert 'void applyGlobalLanguageToReportDetail("drilling")' in script
    assert 'void applyGlobalLanguageToReportDetail("completion")' in script
    assert 'void applyGlobalLanguageToReportDetail("workover")' in script
    assert 'return globalReportContentLanguage === "zh"' in script


def test_monthly_report_navigation_uses_short_names_in_all_languages() -> None:
    root = Path(__file__).parents[1]
    html = root.joinpath("web_form", "index.html").read_text(encoding="utf-8")
    script = root.joinpath("web_form", "app.js").read_text(encoding="utf-8")

    for label in ("钻井月报", "修井月报", "钻修井月报", "月度工作量统计"):
        assert label in html
        assert label in script
    for label in (
        "Drilling Monthly Report",
        "Workover Monthly Report",
        "Drilling & Workover Monthly Report",
        "Monthly Workload Statistics",
        "Reporte Mensual de Perforación",
        "Reporte Mensual de Workover",
        "Reporte Mensual de Perforación y Workover",
        "Estadística Mensual de Carga de Trabajo",
    ):
        assert label in script


def test_four_monthly_report_headers_include_shared_global_actions() -> None:
    root = Path(__file__).parents[1]
    html = root.joinpath("web_form", "index.html").read_text(encoding="utf-8")

    for page_id in (
        "drillingBasicMonthlyReportPage",
        "workoverBasicMonthlyReportPage",
        "drillingWorkoverEfficiencyMonthlyReportPage",
        "monthlyTeamWorkloadReportPage",
    ):
        page_start = html.index(f'id="{page_id}"')
        next_page_start = html.find('<section id="', page_start + 1)
        page_html = html[page_start:next_page_start if next_page_start >= 0 else None]
        assert 'class="top-actions"' in page_html, f"missing top actions in {page_id}"
        assert 'class="language-switch"' in page_html, f"missing language switch in {page_id}"
        assert 'data-lang="original"' in page_html
        assert 'data-lang="zh"' in page_html
        assert 'data-lang="es"' in page_html


def test_drilling_basic_monthly_report_page_matches_appendix_4_structure() -> None:
    root = Path(__file__).parents[1]
    html = root.joinpath("web_form", "index.html").read_text(encoding="utf-8")
    script = root.joinpath("web_form", "app.js").read_text(encoding="utf-8")
    styles = root.joinpath("web_form", "styles.css").read_text(encoding="utf-8")

    monthly_group = html.index('data-i18n="menuMonthlyReport"')
    basic_item = html.index('data-menu-target="drilling-basic-monthly-report"', monthly_group)
    efficiency_item = html.index('data-menu-target="drilling-workover-efficiency-monthly-report"', monthly_group)
    assert basic_item < efficiency_item
    assert 'id="drillingBasicMonthlyReportPage"' in html
    assert 'data-drilling-basic-monthly-table' in html
    assert 'data-drilling-basic-export' in html
    assert '<span>填报日期</span><input type="date" data-drilling-basic-date' not in html
    assert '<span>项目</span><select data-drilling-basic-project' not in html
    assert '<span>队伍编号</span><select data-drilling-basic-team' not in html
    assert 'class="monthly-month-picker" data-monthly-month-picker' in html
    assert '<input type="hidden" data-drilling-basic-date' in html
    assert 'class="monthly-month-trigger" type="button" aria-label="填报月份"' in html
    assert 'new URLSearchParams({ report_month:' in script
    assert 'function monthlyReportFillDate(reportMonth)' in script
    assert 'function renderMonthlyMonthPicker(select)' in script
    assert 'data-monthly-picker-year-step="-1"' in script
    assert 'data-monthly-picker-value=' in script
    assert '.monthly-month-grid {' in styles
    assert 'name="report_date" data-drilling-basic-date' not in html
    assert 'data-drilling-basic-status-note' not in html
    assert 'minimumFractionDigits: 1, maximumFractionDigits: 1' in script
    assert 'escapeHtml(row.well_control_incident)' in script
    assert 'escapeHtml(row.accident_waiting)' in script
    assert 'escapeHtml(row.remarks)' in script
    assert script.count('escapeHtml(row.well_control_incident)') >= 2
    assert 'escapeHtml(row.nonproductive_description)' in script
    assert ".template-highlight {\n  background: #fff;" in styles
    assert ".template-fill-time {\n  margin-left: 64px;" in styles
    assert ".drilling-basic-monthly-workspace {\n  display: grid;\n  gap: 14px;\n  padding: 14px 0 28px;" in styles
    for header in (
        "序号", "队伍编号", "施工国家和地区", "隶属地区公司", "施工区块", "钻机规格型号", "井号", "井型",
        "开钻日期", "完钻日期", "完井日期", "设计井深（英尺）", "当前井深（英尺）", "月进尺（英尺）",
        "本年累计进尺（英尺）", "计划钻井周期（天-时）", "计划完井周期（天-时）", "实际钻井周期（天-时）",
        "实际完井周期（天-时）", "是否有溢流、井涌或井喷", "事故及停待情况", "备注",
    ):
        assert f'"{header}"' in script


def test_drilling_basic_monthly_template_export_populates_exact_static_workbook() -> None:
    template = form_server.MONTHLY_DRILLING_BASIC_TEMPLATE
    assert template.exists()
    assert template.read_bytes().startswith(b"PK")
    handler = object.__new__(form_server.FormHandler)
    handler.send_response = Mock()
    handler.send_header = Mock()
    handler.end_headers = Mock()
    handler.wfile = BytesIO()

    payload = {
        "report_date": "2026-07-20",
        "rows": [{
            "sequence": 1, "team_code": "SINOPEC 248钻井队", "country_region": "厄瓜多尔", "team_company": "西南工程",
            "block_name": "PUCUNA", "rig_model": "ZJ70D", "well_name": "PCNA-001", "well_profile": "定向井",
            "drilling_start_date": "2026-07-01", "drilling_end_date": "2026-07-10", "completion_date": "2026-07-12",
            "design_depth_ft": 10000, "current_depth_ft": 10000, "month_progress_ft": 10000, "year_progress_ft": 10000,
            "actual_drilling_cycle_days": 9.5, "actual_completion_cycle_days": 2,
            "well_control_incident": "无", "accident_waiting": "", "remarks": "",
        }],
    }
    with patch.object(form_server, "_monthly_drilling_basic_report_payload", return_value=payload):
        handler._monthly_drilling_basic_template_export("report_date=2026-07-20")

    handler.send_response.assert_called_once_with(200)
    workbook = load_workbook(BytesIO(handler.wfile.getvalue()), data_only=False)
    assert workbook.sheetnames == ["表4钻井基础指标数据月报"]
    worksheet = workbook["表4钻井基础指标数据月报"]
    assert "2026年7月20日" in worksheet["A2"].value
    assert worksheet["B4"].value == "SINOPEC 248钻井队"
    assert worksheet["B4"].alignment.wrap_text is not True
    assert worksheet["A1"].font.name == worksheet["A3"].font.name == worksheet["B4"].font.name == "宋体"
    assert worksheet["A1"].font.sz == 20
    assert worksheet["A3"].font.sz == 11
    assert worksheet["B4"].font.sz == 10
    assert worksheet.column_dimensions["B"].width == 16
    assert _yellow_filled_cells(worksheet) == []
    assert worksheet["G4"].value == "PCNA-001"
    assert worksheet["I4"].value.strftime("%Y-%m-%d") == "2026-07-01"
    assert worksheet["T4"].value == "无"
    assert worksheet["U4"].value is None
    assert worksheet["V4"].value is None
    assert worksheet["H5"].value == "开钻井数"
    assert worksheet["I5"].value == "=COUNT(I4:I4)"
    assert worksheet["J5"].value == "交井数"
    assert worksheet["K5"].value == "=COUNT(K4:K4)"
    assert worksheet["N5"].value == "=SUM(N4:N4)"
    assert worksheet["O5"].value == "=SUM(O4:O4)"
    assert worksheet["M6"].value == "进尺（米）"
    assert worksheet["N6"].value == "=N5*0.3048"
    assert worksheet["A7"].value == "填报说明："
    assert worksheet["A10"].value.startswith("3.开钻日期")
    assert worksheet["H53"].value is None
    assert worksheet["P4"].number_format == "0.0"
    assert worksheet["R4"].number_format == "0.0"
    assert worksheet["S4"].number_format == "0.0"


def test_drilling_basic_monthly_export_places_zero_totals_after_header() -> None:
    workbook_bytes = form_server._monthly_drilling_basic_workbook_bytes({
        "report_date": "2026-07-20",
        "rows": [],
    })
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    worksheet = workbook["表4钻井基础指标数据月报"]

    assert worksheet["H4"].value == "开钻井数"
    assert worksheet["I4"].value == "=0"
    assert worksheet["J4"].value == "交井数"
    assert worksheet["K4"].value == "=0"
    assert worksheet["N4"].value == "=0"
    assert worksheet["O4"].value == "=0"
    assert worksheet["M5"].value == "进尺（米）"
    assert worksheet["N5"].value == "=N4*0.3048"
    assert worksheet["A6"].value == "填报说明："


def test_drilling_basic_monthly_payload_filters_rows_and_keeps_ai_fields_blank() -> None:
    source = {
        "month_start": "2026-07-01", "month_end": "2026-07-31", "year_start": "2026-01-01",
        "rows": [
            {"job_id": 1, "project_id": 10, "project_name": "A", "team_code": "248", "drilling_start_date": "2026-07-01", "completion_date": "", "month_progress_ft": 100, "year_progress_ft": 100, "well_control_incident": "", "accident_waiting": "", "remarks": ""},
            {"job_id": 2, "project_id": 20, "project_name": "B", "team_code": "168", "drilling_start_date": "2026-07-02", "completion_date": "2026-07-10", "month_progress_ft": 200, "year_progress_ft": 300, "well_control_incident": "", "accident_waiting": "", "remarks": ""},
        ],
    }
    with patch.object(form_server, "load_drilling_basic_monthly_report_rows", return_value=source):
        payload = form_server._monthly_drilling_basic_report_payload(
            Path("mysql"), {"report_date": ["2026-07-20"], "project": ["20"]}
        )

    assert payload["report_month"] == "2026-07"
    assert payload["report_date"] == form_server._monthly_report_fill_date("2026-07")
    assert [row["job_id"] for row in payload["rows"]] == [2]
    assert payload["rows"][0]["sequence"] == 1
    assert payload["summary"] == {
        "drilling_start_count": 1,
        "completion_count": 1,
        "month_progress_ft": 200.0,
        "year_progress_ft": 300.0,
    }


def test_drilling_basic_monthly_payload_uses_team_month_well_control_result() -> None:
    source = {
        "month_start": "2026-07-01", "month_end": "2026-07-31", "year_start": "2026-01-01",
        "rows": [{
            "job_id": 1, "job_sequence_no": 1, "project_id": 10, "project_name": "A", "well_id": 11,
            "team_id": 9, "team_code": "248",
            "well_name": "W-1", "drilling_start_date": "2026-07-01", "completion_date": "",
            "month_progress_ft": 100, "year_progress_ft": 100, "well_control_incident": "",
            "accident_waiting": "", "remarks": "",
        }],
    }
    with (
        patch.object(form_server, "load_drilling_basic_monthly_report_rows", return_value=source),
        patch.object(form_server, "_aggregate_extraction_values", side_effect=[
            {"WELL_JOB:10:11:1": {"well_control_incident": "旧单井结果"}},
            {"TEAM_MONTH:9:drilling:2026-07": {"team_month_well_control_incident": "无"}},
        ]) as load_well_control,
    ):
        payload = form_server._monthly_drilling_basic_report_payload(
            Path("mysql"), {"report_date": ["2026-07-20"]},
        )

    assert payload["rows"][0]["well_control_incident"] == "无"
    assert load_well_control.call_count == 2
    well_job_scope = load_well_control.call_args_list[0].args[1][0]
    team_month_scope = load_well_control.call_args_list[1].args[1][0]
    assert well_job_scope["grain"] == "well_job"
    assert team_month_scope["grain"] == "team_month"
    assert team_month_scope["team_id"] == 9
    assert team_month_scope["profession"] == "drilling"
    assert team_month_scope["period_start"] == "2026-07-01"
    assert team_month_scope["period_end"] == "2026-07-31"


def test_monthly_report_fill_date_uses_today_for_current_month_and_month_end_for_history() -> None:
    today = date(2026, 7, 21)

    assert form_server._monthly_report_fill_date("2026-07", today=today) == "2026-07-21"
    assert form_server._monthly_report_fill_date("2026-06", today=today) == "2026-06-30"
    assert form_server._monthly_report_fill_date("2024-02", today=today) == "2024-02-29"


def test_monthly_report_falls_back_to_latest_available_month() -> None:
    class FixedDate(date):
        @classmethod
        def today(cls) -> "FixedDate":
            return cls(2026, 7, 21)

    sources = [
        {"available_months": ["2026-06", "2026-05"], "rows": []},
        {"available_months": ["2026-06", "2026-05"], "rows": [{"project_id": 1}]},
    ]
    loader = Mock(side_effect=sources)

    with patch.object(form_server, "date", FixedDate):
        selected_month, selected_date, source, options = form_server._monthly_selected_source(
            Path("mysql"), {"report_month": ["2099-12"]}, loader,
        )

    assert selected_month == "2026-06"
    assert selected_date == "2026-06-30"
    assert source["rows"] == [{"project_id": 1}]
    assert options == [
        {"value": "2026-06", "label": "2026年6月"},
        {"value": "2026-05", "label": "2026年5月"},
    ]


def test_workover_basic_monthly_report_page_matches_appendix_5_structure() -> None:
    root = Path(__file__).parents[1]
    html = root.joinpath("web_form", "index.html").read_text(encoding="utf-8")
    script = root.joinpath("web_form", "app.js").read_text(encoding="utf-8")

    monthly_group = html.index('data-i18n="menuMonthlyReport"')
    drilling_item = html.index('data-menu-target="drilling-basic-monthly-report"', monthly_group)
    workover_item = html.index('data-menu-target="workover-basic-monthly-report"', monthly_group)
    efficiency_item = html.index('data-menu-target="drilling-workover-efficiency-monthly-report"', monthly_group)
    assert drilling_item < workover_item < efficiency_item
    assert 'id="workoverBasicMonthlyReportPage"' in html
    assert 'data-workover-basic-monthly-table' in html
    assert 'data-workover-basic-export' in html
    assert '<span>填报日期</span><input type="date" data-workover-basic-date' not in html
    assert '<span>项目</span><select data-workover-basic-project' not in html
    assert '<span>队伍编号</span><select data-workover-basic-team' not in html
    assert '<input type="hidden" data-workover-basic-date' in html
    for header in (
        "序号", "队伍编号", "施工国家和地区", "隶属地区公司", "施工区块", "钻机规格型号", "井号", "井型",
        "开工日期", "完工日期", "作业主要内容", "是否有溢流、井涌或井喷", "事故及停待情况", "备注",
    ):
        assert f'"{header}"' in script


def test_workover_basic_monthly_template_export_populates_exact_static_workbook() -> None:
    template = form_server.MONTHLY_WORKOVER_BASIC_TEMPLATE
    assert template.exists()
    assert template.read_bytes().startswith(b"PK")
    handler = object.__new__(form_server.FormHandler)
    handler.send_response = Mock()
    handler.send_header = Mock()
    handler.end_headers = Mock()
    handler.wfile = BytesIO()

    payload = {
        "report_date": "2026-07-20",
        "rows": [{
            "sequence": 1, "team_code": "SINOPEC 932修井队", "country_region": "厄瓜多尔", "team_company": "华东工程",
            "block_name": "AUCA", "rig_model": "XJ650", "well_name": "CNOA-047", "well_profile": "油井",
            "workover_start_date": "2026-07-01", "workover_end_date": "2026-07-10", "primary_operation": "检泵",
            "well_control_incident": "", "accident_waiting": "", "remarks": "",
        }],
    }
    with patch.object(form_server, "_monthly_workover_basic_report_payload", return_value=payload):
        handler._monthly_workover_basic_template_export("report_date=2026-07-20")

    handler.send_response.assert_called_once_with(200)
    workbook = load_workbook(BytesIO(handler.wfile.getvalue()), data_only=False)
    assert workbook.sheetnames == ["表5 修井基础指标数据月报"]
    worksheet = workbook["表5 修井基础指标数据月报"]
    assert worksheet["A2"].value == "填报单位：厄瓜多尔子公司"
    assert worksheet["K2"].value == "填报时间：2026年7月20日"
    assert worksheet["B4"].value == "SINOPEC 932修井队"
    assert worksheet["A1"].font.name == worksheet["A3"].font.name == worksheet["B4"].font.name == "宋体"
    assert worksheet["A1"].font.sz == 20
    assert worksheet["A3"].font.sz == 11
    assert worksheet["B4"].font.sz == 10
    assert worksheet.column_dimensions["B"].width == 16
    assert _yellow_filled_cells(worksheet) == []
    assert worksheet["I4"].value.strftime("%Y-%m-%d") == "2026-07-01"
    assert worksheet["K4"].value == "检泵"
    assert worksheet["L4"].value is None
    assert worksheet["M4"].value is None
    assert worksheet["N4"].value is None
    assert worksheet["I5"].value == "修井完工口数"
    assert worksheet["J5"].value == "=COUNT(J4:J4)"
    assert worksheet["A6"].value == "填报说明："
    assert worksheet["A7"].value.startswith("1.设备目前状态")
    assert worksheet["A8"].value.startswith("2.开工日期")
    assert worksheet["I75"].value is None


def test_workover_basic_monthly_export_places_zero_total_after_header() -> None:
    workbook_bytes = form_server._monthly_workover_basic_workbook_bytes({
        "report_date": "2026-07-20",
        "rows": [],
    })
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    worksheet = workbook["表5 修井基础指标数据月报"]

    assert worksheet["I4"].value == "修井完工口数"
    assert worksheet["J4"].value == "=0"
    assert worksheet["A5"].value == "填报说明："
    assert worksheet["A6"].value.startswith("1.设备目前状态")
    assert worksheet["A7"].value.startswith("2.开工日期")


def test_workover_basic_monthly_payload_filters_rows_and_keeps_ai_fields_blank() -> None:
    source = {
        "month_start": "2026-07-01", "month_end": "2026-07-31",
        "rows": [
            {"job_id": 1, "project_id": 10, "project_name": "A", "team_code": "SINOPEC 932", "workover_end_date": "", "primary_operation": "", "well_control_incident": "", "accident_waiting": "", "remarks": ""},
            {"job_id": 2, "project_id": 20, "project_name": "B", "team_code": "SINOPEC 976", "workover_end_date": "2026-07-10", "primary_operation": "检泵", "well_control_incident": "", "accident_waiting": "", "remarks": ""},
        ],
    }
    with patch.object(form_server, "load_workover_basic_monthly_report_rows", return_value=source):
        payload = form_server._monthly_workover_basic_report_payload(
            Path("mysql"), {"report_date": ["2026-07-20"], "project": ["20"]}
        )

    assert payload["report_month"] == "2026-07"
    assert payload["report_date"] == form_server._monthly_report_fill_date("2026-07")
    assert [row["job_id"] for row in payload["rows"]] == [2]
    assert payload["rows"][0]["sequence"] == 1
    assert payload["summary"] == {"completion_count": 1}


def test_drilling_workover_efficiency_monthly_page_matches_appendix_6_structure() -> None:
    root = Path(__file__).parents[1]
    html = root.joinpath("web_form", "index.html").read_text(encoding="utf-8")
    script = root.joinpath("web_form", "app.js").read_text(encoding="utf-8")

    monthly_group = html.index('data-i18n="menuMonthlyReport"')
    drilling_item = html.index('data-menu-target="drilling-basic-monthly-report"', monthly_group)
    workover_item = html.index('data-menu-target="workover-basic-monthly-report"', monthly_group)
    efficiency_item = html.index('data-menu-target="drilling-workover-efficiency-monthly-report"', monthly_group)
    assert drilling_item < workover_item < efficiency_item
    assert 'id="drillingWorkoverEfficiencyMonthlyReportPage"' in html
    assert 'data-drilling-workover-efficiency-monthly-table' in html
    assert 'data-drilling-workover-efficiency-export' in html
    assert '<span>填报日期</span><input type="date" data-drilling-workover-efficiency-date' not in html
    assert '<span>项目</span><select data-drilling-workover-efficiency-project' not in html
    assert '<span>队伍编号</span><select data-drilling-workover-efficiency-team' not in html
    assert '<input type="hidden" data-drilling-workover-efficiency-date' in html
    for header in (
        "序号", "队伍编号", "施工井号", "专业", "施工国家和地区", "隶属地区公司", "施工区块", "钻机规格型号",
        "搬安时间", "生产时间", "非生产时间", "修理时间", "事故、复杂情况时间", "单井生产时效",
        "非生产时间原因描述", "平均生产时效", "备注",
    ):
        assert header in script
    assert "reservedRowCount" not in script
    assert "reservedRowCount - dataValues.length" not in script


def test_drilling_workover_efficiency_monthly_template_export_uses_exact_appendix_6_workbook() -> None:
    template = form_server.MONTHLY_DRILLING_WORKOVER_EFFICIENCY_TEMPLATE
    assert template.exists()
    assert template.read_bytes().startswith(b"PK")
    handler = object.__new__(form_server.FormHandler)
    handler.send_response = Mock()
    handler.send_header = Mock()
    handler.end_headers = Mock()
    handler.wfile = BytesIO()
    payload = {
        "report_date": "2026-07-21",
        "rows": [{
            "sequence": 1, "team_code": "SINOPEC 933修井队", "well_name": "YLBD-043HS1", "profession_label": "修井",
            "country_region": "厄瓜多尔", "team_company": "华东工程", "block_name": "SHUSHUFINDI", "rig_model": "ZJ30",
            "move_hours": 12.25, "production_hours": 100.04, "paid_repair_hours": 12, "zero_rate_repair_hours": 3.5,
            "accident_complex_hours": 0, "other_hours": 0, "well_efficiency": 0.8654,
            "nonproductive_description": "", "average_efficiency": None, "remarks": "",
        }],
    }
    with patch.object(form_server, "_monthly_drilling_workover_efficiency_report_payload", return_value=payload):
        handler._monthly_drilling_workover_efficiency_template_export("report_date=2026-07-21")

    handler.send_response.assert_called_once_with(200)
    workbook = load_workbook(BytesIO(handler.wfile.getvalue()), data_only=False)
    assert workbook.sheetnames == ["表6钻修井基础时效数据月报 "]
    worksheet = workbook["表6钻修井基础时效数据月报 "]
    assert "2026年7月21日" in worksheet["A2"].value
    assert worksheet["B6"].value == "SINOPEC 933修井队"
    assert worksheet["B6"].alignment.wrap_text is not True
    assert worksheet["A1"].font.name == worksheet["A3"].font.name == worksheet["B6"].font.name == "宋体"
    assert worksheet["A1"].font.sz == 20
    assert worksheet["A3"].font.sz == 11
    assert worksheet["B6"].font.sz == 10
    assert worksheet.column_dimensions["B"].width == 16
    assert _yellow_filled_cells(worksheet) == []
    assert worksheet["C6"].value == "YLBD-043HS1"
    assert worksheet["D6"].value == "修井"
    assert worksheet["I6"].value == 12.25
    assert worksheet["I6"].number_format == "0.0"
    assert worksheet["K6"].value == 12
    assert worksheet["L6"].value == 3.5
    assert worksheet["M6"].value == 0
    assert worksheet["N6"].value == 0
    assert worksheet["O6"].value == '=IF(J6+SUM(K6:N6)=0,"",J6/(J6+SUM(K6:N6)))'
    assert worksheet["P6"].value is None
    assert worksheet["Q6"].value is None
    assert worksheet["R6"].value is None
    assert worksheet["A7"].value == "填报说明："
    assert worksheet["A8"].value == "1. 单井生产时效=单井生产时间/（单井生产时间+单井非生产时间）"
    assert worksheet.max_row == 9


def test_drilling_workover_efficiency_monthly_payload_filters_rows() -> None:
    source = {
        "month_start": "2026-07-01", "month_end": "2026-07-31",
        "rows": [
            {"project_id": 10, "project_name": "A", "team_code": "SINOPEC 168", "well_name": "W-1"},
            {"project_id": 20, "project_name": "B", "team_code": "SINOPEC 933", "well_name": "W-2"},
        ],
    }
    with patch.object(form_server, "load_drilling_workover_efficiency_monthly_report_rows", return_value=source):
        payload = form_server._monthly_drilling_workover_efficiency_report_payload(
            Path("mysql"), {"report_date": ["2026-07-21"], "project": ["20"]}
        )

    assert payload["report_month"] == "2026-07"
    assert payload["report_date"] == form_server._monthly_report_fill_date("2026-07")
    assert [row["well_name"] for row in payload["rows"]] == ["W-2"]
    assert payload["rows"][0]["sequence"] == 1


def test_drilling_workover_efficiency_monthly_hides_ai_text_when_nonproductive_hours_are_zero() -> None:
    source = {
        "month_start": "2026-07-01", "month_end": "2026-07-31",
        "rows": [{
            "project_id": 20, "project_name": "B", "well_id": 30,
            "team_code": "SINOPEC 933", "well_name": "W-2", "profession": "drilling",
            "paid_repair_hours": 0, "zero_rate_repair_hours": 0,
            "accident_complex_hours": 0, "other_hours": 0,
            "nonproductive_description": "",
        }],
    }

    def inject_stale_text(_database_path, rows, **_kwargs):
        rows[0]["nonproductive_description"] = "不应显示的旧提炼结果"

    with (
        patch.object(form_server, "load_drilling_workover_efficiency_monthly_report_rows", return_value=source),
        patch.object(form_server, "_apply_monthly_aggregate_extractions", side_effect=inject_stale_text),
    ):
        payload = form_server._monthly_drilling_workover_efficiency_report_payload(
            Path("mysql"), {"report_date": ["2026-07-20"]},
        )

    assert payload["rows"][0]["nonproductive_description"] == "无"


def test_monthly_team_workload_page_is_fourth_monthly_report_and_has_shared_controls() -> None:
    root = Path(__file__).parents[1]
    html = root.joinpath("web_form", "index.html").read_text(encoding="utf-8")
    script = root.joinpath("web_form", "app.js").read_text(encoding="utf-8")

    monthly_group = html.index('data-i18n="menuMonthlyReport"')
    efficiency_item = html.index('data-menu-target="drilling-workover-efficiency-monthly-report"', monthly_group)
    workload_item = html.index('data-menu-target="monthly-team-workload-report"', monthly_group)
    assert efficiency_item < workload_item
    assert 'id="monthlyTeamWorkloadReportPage"' in html
    assert '<input type="hidden" data-monthly-team-workload-date' in html
    assert 'data-monthly-team-workload-project' in html
    assert 'data-monthly-team-workload-team' in html
    assert 'data-monthly-team-workload-export' in html
    assert 'src="./assets/sinopec-logo.png"' in script
    assert "填报单位：厄瓜多尔子公司" not in script.split('function renderMonthlyTeamWorkloadTable', 1)[1].split('async function loadMonthlyTeamWorkloadReport', 1)[0]
    for header in ("作业队伍", "工作时间（单位：小时）", "有人待工", "无人待工", "不可抗力待工", "维修/零日费"):
        assert header in script


def test_hsse_collection_page_uses_structured_daily_four_category_entry() -> None:
    root = Path(__file__).parents[1]
    html = root.joinpath("web_form", "index.html").read_text(encoding="utf-8")
    script = root.joinpath("web_form", "app.js").read_text(encoding="utf-8")
    schema = root.joinpath("db", "init.sql").read_text(encoding="utf-8")

    assert 'id="hsseCollectionPage"' in html
    assert '"hsse-collection": "hsseCollectionPage"' in script
    for code in ("UNSAFE_BEHAVIOR", "SAFETY_HAZARD", "CONCERN_EMPLOYEE", "PRODUCTION_ANOMALY"):
        assert f'data-hsse-category="{code}"' in html
        assert code in schema
    assert "CREATE TABLE IF NOT EXISTS hsse_daily_record" in schema
    assert "UNIQUE KEY uq_hsse_daily_team_date (team_id, record_date)" in schema
    assert "data_source_type VARCHAR(32)" in schema
    assert "source_context_json JSON" in schema
    assert "CREATE TABLE IF NOT EXISTS hsse_daily_record_well" in schema
    assert "CREATE TABLE IF NOT EXISTS hsse_daily_item" in schema
    assert 'type="hidden" data-hsse-project' in html
    assert "data-hsse-well-trigger" in html
    assert "data-hsse-well-id" in script
    assert "井队24小时工况摘要" in html
    assert "preferredWellIds" in script
    assert 'adminRequest("/api/hsse/daily-records"' in script


def test_hsse_dashboard_uses_full_roster_and_excludes_rectification_workflow() -> None:
    root = Path(__file__).parents[1]
    html = root.joinpath("web_form", "index.html").read_text(encoding="utf-8")
    script = root.joinpath("web_form", "app.js").read_text(encoding="utf-8")
    service = root.joinpath("drilling_report_parser", "hsse_service.py").read_text(encoding="utf-8")
    server = root.joinpath("drilling_report_parser", "form_server.py").read_text(encoding="utf-8")

    assert 'id="hsseDashboardPage"' in html
    assert '"hsse-dashboard": "hsseDashboardPage"' in script
    assert 'adminRequest(`/api/hsse/dashboard?' in script
    assert 'parsed.path == "/api/hsse/dashboard"' in server
    assert 'parsed.path == "/api/hsse/dashboard-export"' in server
    assert "rel_project_team_assignment" in service
    assert "expected_team_days" in service
    assert "EXCEL_IMPORT" in service
    assert "SIMULATED" in service
    assert "Excel真实记录" in service
    assert "模拟数据" in service
    dashboard_markup = html.split('id="hsseDashboardPage"', 1)[1].split('id="modulePlaceholder"', 1)[0]
    assert "data-hsse-dashboard-source" in dashboard_markup
    assert "数据来源" in dashboard_markup
    assert "漏报（天）" not in dashboard_markup
    assert "事项次数 / 本月应填报天数" in dashboard_markup
    assert "value / expectedDays * 100" in script
    assert "source_summary" in service
    assert "source_reference" in service
    for forbidden in ("整改措施", "整改状态", "闭环状态", "计划完成", "审核状态"):
        assert forbidden not in dashboard_markup


def test_monthly_team_workload_payload_groups_projects_by_standard_team_and_profession() -> None:
    source = {
        "month_start": "2026-06-01",
        "month_end": "2026-06-30",
        "available_months": ["2026-06"],
        "rows": [
            {"project_id": 10, "project_name": "A", "team_code": "SINOPEC 127", "team_name": "SINOPEC 127", "profession": "drilling", "operation_hours": 30.25, "move_hours": 6, "manned_standby_hours": 0, "unmanned_standby_hours": 0, "force_majeure_hours": 0, "zero_rate_repair_hours": 1.25},
            {"project_id": 20, "project_name": "B", "team_code": "SINOPEC 127", "team_name": "SINOPEC 127", "profession": "drilling", "operation_hours": 20.25, "move_hours": 18, "manned_standby_hours": 0, "unmanned_standby_hours": 0, "force_majeure_hours": 0, "zero_rate_repair_hours": 2.25},
            {"project_id": 20, "project_name": "B", "team_code": "SINOPEC 127", "team_name": "SINOPEC 127", "profession": "workover", "operation_hours": 24, "move_hours": 0, "manned_standby_hours": 0, "unmanned_standby_hours": 0, "force_majeure_hours": 0, "zero_rate_repair_hours": 0},
        ],
    }
    with patch.object(form_server, "load_monthly_team_workload_report_rows", return_value=source):
        payload = form_server._monthly_team_workload_report_payload(Path("mysql"), {"report_month": ["2026-06"]})

    assert len(payload["rows"]) == 2
    drilling, workover = payload["rows"]
    assert drilling["category_label"] == "钻机"
    assert drilling["team_name"] == "SINOPEC 127"
    assert drilling["operation_hours"] == 50.5
    assert drilling["move_hours"] == 24.0
    assert drilling["zero_rate_repair_hours"] == 3.5
    assert drilling["total_hours"] == 78.0
    assert workover["category_label"] == "修井"
    assert workover["total_hours"] == 24.0


def test_monthly_team_workload_export_has_dynamic_rows_and_clean_formulas() -> None:
    data = form_server._monthly_team_workload_workbook_bytes({
        "report_month": "2026-06",
        "report_date": "2026-06-30",
        "rows": [
            {"category_label": "钻机", "team_name": "SINOPEC 127", "operation_hours": 100.2, "move_hours": 24, "manned_standby_hours": 0, "unmanned_standby_hours": 0, "force_majeure_hours": 0, "zero_rate_repair_hours": 5.3, "remarks": ""},
            {"category_label": "修井", "team_name": "SINOPEC 903", "operation_hours": 72, "move_hours": 0, "manned_standby_hours": 0, "unmanned_standby_hours": 0, "force_majeure_hours": 0, "zero_rate_repair_hours": 1, "remarks": ""},
        ],
    })
    workbook = load_workbook(BytesIO(data), data_only=False)
    assert workbook.sheetnames == ["6月份"]
    worksheet = workbook["6月份"]

    assert worksheet["A1"].value == "2026年6月厄子公司石油工程项目工作量统计表"
    assert worksheet["F2"].value is None
    assert "A1:J2" in {str(item) for item in worksheet.merged_cells.ranges}
    assert worksheet["B5"].value == "SINOPEC 127"
    assert worksheet["I5"].value == "=SUM(C5:H5)"
    assert worksheet["I5"].number_format == "0.0"
    assert worksheet["B6"].value == "SINOPEC 903"
    assert worksheet["A7"].value == "负责人："
    assert worksheet["A8"].value == "备注：28日至月末期间的工作量为预估值。"
    assert worksheet.max_row == 8
    assert worksheet.max_column == 10
    assert len(worksheet._images) == 1
    logo_anchor = worksheet._images[0].anchor
    assert round(logo_anchor.ext.cx / 9525) == 71
    assert round(logo_anchor.ext.cy / 9525) == 74
    assert round(logo_anchor._from.rowOff / 9525) == 7
    assert _yellow_filled_cells(worksheet) == []
    assert not any(
        isinstance(cell.value, str) and "#REF!" in cell.value
        for row in worksheet.iter_rows()
        for cell in row
    )


def test_completeness_is_not_inferred_without_an_explicit_period() -> None:
    records = [
        {"reportDate": "2026-06-01", "validation_status": "ok"},
        {"reportDate": "2026-06-30", "validation_status": "warning"},
    ]

    result = form_server._completeness(records)

    assert result == {
        "assessed": False,
        "percent": None,
        "missing_days": None,
        "warning_days": 1,
        "observed_days": 2,
        "coverage_basis": "NOT_ASSESSED_WITHOUT_EXPLICIT_PERIOD",
    }


def test_completeness_uses_only_the_explicit_calendar_period() -> None:
    records = [
        {"reportDate": "2026-06-01", "validation_status": "ok"},
        {"reportDate": "2026-06-03", "validation_status": "warning"},
        {"reportDate": "2026-07-01", "validation_status": "warning"},
    ]

    result = form_server._completeness(records, date_from="2026-06-01", date_to="2026-06-03")

    assert result == {
        "assessed": True,
        "percent": 66.7,
        "missing_days": 1,
        "warning_days": 1,
        "observed_days": 2,
        "expected_days": 3,
        "coverage_basis": "CALENDAR_DAY_WITH_ANY_MATCHED_REPORT",
    }


def test_analytics_quality_exposes_pending_hours_and_readiness() -> None:
    operations = [
        {
            "record_id": "ready",
            "source_row_no": 1,
            "hours": 8,
            "classification_status": "AUTO_CONFIRMED",
            "statistics_ready": True,
        },
        {
            "record_id": "pending",
            "source_row_no": 1,
            "hours": 4,
            "classification_status": "PENDING",
            "statistics_ready": False,
        },
    ]

    result = form_server._analytics_quality_fields([], operations)

    assert result["unconfirmed_classification_count"] == 1
    assert result["unconfirmed_classification_hours"] == 4.0
    assert result["statistics_ready_hours"] == 8.0
    assert result["statistics_ready_percent"] == 66.7
    assert result["statistics_ready"] is False


class FormServerImportTest(unittest.TestCase):
    def setUp(self) -> None:
        self.repository_tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.repository_tmp.cleanup)
        self.repository_path = Path(self.repository_tmp.name) / "report_database.xlsx"
        original_database_path = form_server.DATABASE_PATH
        form_server.DATABASE_PATH = self.repository_path
        self.addCleanup(setattr, form_server, "DATABASE_PATH", original_database_path)

        def repository(path: Path | str | None) -> Path:
            value = Path(path or self.repository_path)
            return self.repository_path if value == Path("mysql") else value

        def save_payload(path: Path, payload: dict[str, Any], report_type: str, **kwargs: Any) -> dict[str, Any]:
            kwargs.pop("invalidate_translations", None)
            return excel_database.save_report_payload(repository(path), payload, report_type, **kwargs)

        def load_payload(path: Path, record_id: str) -> dict[str, Any]:
            return excel_database.load_report_payload(repository(path), record_id)

        def operation_translations(path: Path, record_ids: list[str]) -> list[dict[str, str]]:
            return [
                row
                for record_id in record_ids
                for row in excel_database.load_translation_content(repository(path), record_id)
                if str(row.get("entity_type", "")) == "operations"
            ]

        def records(path: Path | None = None, **filters: str) -> list[dict[str, str]]:
            values = excel_database.list_records(repository(path))
            return [
                row for row in values
                if (not filters.get("report_type") or row.get("report_type") == filters["report_type"])
                and (not filters.get("wellbore") or row.get("wellbore") == filters["wellbore"])
                and (not filters.get("date") or row.get("reportDate") == filters["date"])
                and (not filters.get("date_from") or str(row.get("reportDate", "")) >= filters["date_from"])
                and (not filters.get("date_to") or str(row.get("reportDate", "")) <= filters["date_to"])
            ]

        patchers = [
            patch.object(form_server, "save_report_payload", side_effect=save_payload),
            patch.object(form_server, "load_report_payload", side_effect=load_payload),
            patch.object(form_server, "list_records", side_effect=records),
            patch.object(form_server, "load_operation_translations", side_effect=operation_translations),
            patch.object(form_server, "load_extraction_results", return_value=[]),
            patch.object(form_server, "clear_extraction_results"),
            patch.object(form_server, "_extraction_jobs_enabled", return_value=False),
        ]
        for patcher in patchers:
            patcher.start()
            self.addCleanup(patcher.stop)

    def test_report_identity_requires_date_well_and_rig(self) -> None:
        payload = {"report_fields": {"reportDate": "", "wellbore": "PCNC-039", "rig": ""}}

        self.assertEqual(form_server._report_identity_errors(payload), ["日报日期", "井队"])
        self.assertEqual(form_server._report_identity_errors({"report_fields": {
            "reportDate": "2026-05-22", "wellbore": "PCNC-039", "rig": "SINOPEC 248",
        }}), [])

    def test_all_pdf_categories_use_shared_multi_report_import(self) -> None:
        handler = object.__new__(form_server.FormHandler)
        handler._import_report_pdf = Mock()

        handler._import_pdf()
        handler._import_completion_pdf()
        handler._import_workover_pdf()
        handler._import_move_pdf()

        self.assertEqual(
            [call.args[0] for call in handler._import_report_pdf.call_args_list],
            [
                "drilling",
                "completion",
                "workover",
                "drilling",
            ],
        )

    def test_drilling_import_passes_selected_template_profile(self) -> None:
        handler = object.__new__(form_server.FormHandler)
        handler.path = "/api/import-pdf?template_profile=compatible"
        handler._import_report_pdf = Mock()

        handler._import_pdf()

        handler._import_report_pdf.assert_called_once_with("drilling", template_profile="compatible")

    def test_unknown_drilling_template_returns_validation_error(self) -> None:
        handler = object.__new__(form_server.FormHandler)
        handler._read_pdf_upload = Mock(return_value=form_server.UploadedFile(
            filename="daily.pdf",
            data=b"%PDF-unknown-template",
        ))
        handler._send_json = Mock()

        handler._import_report_pdf("drilling", template_profile="unknown")

        response = handler._send_json.call_args
        self.assertIn("不支持的钻井 PDF 模板", response.args[0]["error"])
        self.assertEqual(response.kwargs["status"], 400)

    def test_multi_report_pdf_import_stores_every_segment_and_returns_reports(self) -> None:
        writer = PdfWriter()
        for width in (101, 150, 201, 250):
            writer.add_blank_page(width=width, height=300)
        output = BytesIO()
        writer.write(output)

        def parser(source: bytes) -> dict[str, Any]:
            pages = PdfReader(BytesIO(source)).pages
            width = round(float(pages[0].mediabox.width))
            identities = {
                101: ("2026-06-01", "WELL-A", "1"),
                201: ("2026-06-02", "WELL-A", "2"),
            }
            identity = identities.get(width)
            fields = {}
            if identity:
                fields = {
                    "event": "DEV COMPLETION",
                    "reportDate": identity[0],
                    "wellbore": identity[1],
                    "reportNo": identity[2],
                    "rig": "SINOPEC 191",
                }
            return {"metadata": {}, "report_fields": fields, "operations": []}

        handler = object.__new__(form_server.FormHandler)
        handler._read_pdf_upload = Mock(return_value=form_server.UploadedFile(
            filename="combined-completion.pdf",
            data=output.getvalue(),
        ))
        handler._store_payload = Mock()
        handler._store_source_pdf = Mock()
        handler._send_json = Mock()

        strategy = Mock(storage_report_type="completion", parser=parser)
        with patch.object(form_server, "pdf_import_strategy", return_value=strategy):
            handler._import_report_pdf("completion")

        self.assertEqual(handler._store_payload.call_count, 2)
        self.assertEqual(handler._store_source_pdf.call_count, 2)
        first_payload = handler._store_payload.call_args_list[0].args[0]
        second_payload = handler._store_payload.call_args_list[1].args[0]
        self.assertEqual(first_payload["metadata"]["source_page_start"], 1)
        self.assertEqual(first_payload["metadata"]["source_page_end"], 2)
        self.assertEqual(second_payload["metadata"]["source_page_start"], 3)
        self.assertEqual(second_payload["metadata"]["source_page_end"], 4)
        response = handler._send_json.call_args.args[0]
        self.assertTrue(response["metadata"]["multi_report"])
        self.assertEqual(response["metadata"]["imported_count"], 2)
        self.assertEqual(len(response["reports"]), 2)

    def test_pdf_import_rejects_event_type_mismatch_before_storing(self) -> None:
        writer = PdfWriter()
        writer.add_blank_page(width=101, height=300)
        output = BytesIO()
        writer.write(output)

        payload = {
            "metadata": {},
            "report_fields": {
                "event": "WORKOVER",
                "reportDate": "2026-07-15",
                "wellbore": "ACAM-148",
                "reportNo": "22",
                "rig": "SINOPEC 976",
            },
            "operations": [],
        }
        handler = object.__new__(form_server.FormHandler)
        handler._read_pdf_upload = Mock(return_value=form_server.UploadedFile(
            filename="workover-in-completion.pdf",
            data=output.getvalue(),
        ))
        handler._store_payload = Mock()
        handler._store_source_pdf = Mock()
        handler._send_json = Mock()

        strategy = Mock(storage_report_type="completion", parser=Mock(return_value=payload))
        with patch.object(form_server, "pdf_import_strategy", return_value=strategy):
            handler._import_report_pdf("completion")

        handler._store_payload.assert_not_called()
        handler._store_source_pdf.assert_not_called()
        error_payload = handler._send_json.call_args.args[0]
        self.assertIn("识别为修井日报", error_payload["error"])
        self.assertIn("当前上传入口是完井日报", error_payload["error"])
        self.assertEqual(handler._send_json.call_args.kwargs["status"], 400)

    def test_rig_move_event_is_accepted_from_drilling_and_stored_as_drilling(self) -> None:
        writer = PdfWriter()
        writer.add_blank_page(width=101, height=300)
        output = BytesIO()
        writer.write(output)

        payload = {
            "metadata": {},
            "report_fields": {
                "event": "MAJOR RIG MOVE",
                "reportDate": "2026-06-10",
                "wellbore": "TCHA-006I",
                "reportNo": "5",
                "rig": "SINOPEC 168",
            },
            "operations": [],
        }
        handler = object.__new__(form_server.FormHandler)
        handler._read_pdf_upload = Mock(return_value=form_server.UploadedFile(
            filename="rig-move.pdf",
            data=output.getvalue(),
        ))
        handler._store_payload = Mock()
        handler._store_source_pdf = Mock()
        handler._send_json = Mock()

        strategy = Mock(storage_report_type="drilling", parser=Mock(return_value=payload))
        with patch.object(form_server, "pdf_import_strategy", return_value=strategy):
            handler._import_report_pdf("drilling")

        handler._store_payload.assert_called_once()
        self.assertEqual(handler._store_payload.call_args.args[1], "drilling")
        stored_payload = handler._store_payload.call_args.args[0]
        self.assertEqual(stored_payload["metadata"]["detected_event_type"], "move")
        self.assertEqual(stored_payload["metadata"]["detected_report_type"], "drilling")

    def test_multi_report_type_mismatch_rejects_the_whole_batch(self) -> None:
        writer = PdfWriter()
        writer.add_blank_page(width=101, height=300)
        writer.add_blank_page(width=201, height=300)
        output = BytesIO()
        writer.write(output)

        def parser(source: bytes) -> dict[str, Any]:
            width = round(float(PdfReader(BytesIO(source)).pages[0].mediabox.width))
            is_first = width == 101
            return {
                "metadata": {},
                "report_fields": {
                    "event": "DEV COMPLETION" if is_first else "WORKOVER",
                    "reportDate": "2026-07-14" if is_first else "2026-07-15",
                    "wellbore": "WELL-A",
                    "reportNo": "21" if is_first else "22",
                    "rig": "SINOPEC 976",
                },
                "operations": [],
            }

        handler = object.__new__(form_server.FormHandler)
        handler._read_pdf_upload = Mock(return_value=form_server.UploadedFile(
            filename="mixed-types.pdf",
            data=output.getvalue(),
        ))
        handler._store_payload = Mock()
        handler._store_source_pdf = Mock()
        handler._send_json = Mock()

        strategy = Mock(storage_report_type="completion", parser=parser)
        with patch.object(form_server, "pdf_import_strategy", return_value=strategy):
            handler._import_report_pdf("completion")

        handler._store_payload.assert_not_called()
        handler._store_source_pdf.assert_not_called()
        error_payload = handler._send_json.call_args.args[0]
        self.assertIn("合并 PDF 第2份日报", error_payload["error"])
        self.assertIn("识别为修井日报", error_payload["error"])

    def test_batch_rig_is_inherited_only_when_same_well_is_unanimous(self) -> None:
        payloads = [
            {"metadata": {}, "report_fields": {"wellbore": "WELL-A", "rig": ""}},
            {"metadata": {}, "report_fields": {"wellbore": "WELL-A", "rig": "SINOPEC 191"}},
            {"metadata": {}, "report_fields": {"wellbore": "WELL-A", "rig": "SINOPEC-191"}},
            {"metadata": {}, "report_fields": {"wellbore": "WELL-B", "rig": ""}},
            {"metadata": {}, "report_fields": {"wellbore": "WELL-B", "rig": "SINOPEC 127"}},
            {"metadata": {}, "report_fields": {"wellbore": "WELL-B", "rig": "SINOPEC 129"}},
        ]

        form_server._inherit_consistent_batch_rigs(payloads)

        self.assertEqual(payloads[0]["report_fields"]["rig"], "SINOPEC-191")
        self.assertEqual(payloads[0]["metadata"]["batch_inherited_fields"], ["rig"])
        self.assertEqual(payloads[3]["report_fields"]["rig"], "")
        self.assertNotIn("batch_inherited_fields", payloads[3]["metadata"])

    def test_batch_source_metadata_is_persisted(self) -> None:
        payload = {
            "metadata": {
                "source_file": "combined.pdf",
                "source_page_start": 3,
                "source_page_end": 4,
                "source_report_index": 2,
                "source_report_count": 5,
                "batch_inherited_fields": ["rig"],
            },
            "report_fields": {
                "event": "DEV COMPLETION",
                "reportDate": "2026-06-02",
                "reportNo": "2",
                "wellbore": "WELL-A",
                "rig": "SINOPEC 191",
            },
            "operations": [],
        }
        handler = object.__new__(form_server.FormHandler)

        handler._store_payload(payload, "completion", from_upload=True)

        stored = form_server.load_report_payload(
            form_server.DATABASE_PATH,
            "completion:WELL-A:2026-06-02:2",
        )
        self.assertEqual(stored["metadata"]["source_page_start"], "3")
        self.assertEqual(stored["metadata"]["source_page_end"], "4")
        self.assertEqual(stored["metadata"]["source_report_index"], "2")
        self.assertEqual(stored["metadata"]["source_report_count"], "5")
        self.assertEqual(stored["metadata"]["batch_inherited_fields"], ["rig"])

    def test_uploaded_report_is_auto_translated_and_reupload_is_queued_again(self) -> None:
        payload = {
            "metadata": {},
            "report_fields": {
                "event": "DEV DRILLING",
                "reportDate": "2026-05-17",
                "reportNo": "5",
                "wellbore": "PCNC-039",
                "rig": "SINOPEC 248",
                "currentOps": "DRILL AHEAD.",
            },
            "operations": [],
        }
        handler = object.__new__(form_server.FormHandler)

        with (
            patch.object(form_server, "_load_translation_tuning_config", return_value={
                **form_server._default_translation_tuning_config(),
                "auto_translate_on_upload": True,
            }),
            patch.object(form_server, "_translation_jobs_enabled", return_value=True),
            patch.object(form_server, "_invalidate_translation_jobs") as invalidate_jobs,
            patch.object(form_server, "_schedule_translation_job") as schedule_job,
        ):
            handler._store_payload(payload, "drilling", from_upload=True)
            handler._store_payload(payload, "drilling", from_upload=True)

        record_id = "drilling:PCNC-039:2026-05-17:5"
        self.assertEqual(payload["metadata"]["translation_status"], "QUEUED")
        self.assertEqual(schedule_job.call_args_list, [unittest.mock.call(record_id), unittest.mock.call(record_id)])
        self.assertEqual(invalidate_jobs.call_count, 2)
        self.assertTrue(form_server.save_report_payload.call_args.kwargs["invalidate_translations"])

    def test_manual_report_save_does_not_use_upload_auto_translation(self) -> None:
        payload = {
            "metadata": {},
            "report_fields": {
                "reportDate": "2026-05-17",
                "reportNo": "5",
                "wellbore": "PCNC-039",
                "rig": "SINOPEC 248",
                "currentOps": "DRILL AHEAD.",
            },
            "operations": [],
        }
        handler = object.__new__(form_server.FormHandler)

        with (
            patch.object(form_server, "_load_translation_tuning_config", return_value={
                **form_server._default_translation_tuning_config(),
                "auto_translate_on_upload": True,
            }),
            patch.object(form_server, "_translation_jobs_enabled", return_value=True),
            patch.object(form_server, "_schedule_translation_job") as schedule_job,
        ):
            handler._store_payload(payload, "drilling")

        schedule_job.assert_not_called()
        self.assertEqual(payload["metadata"]["translation_status"], "PENDING")

    def test_extraction_requeues_only_after_relevant_input_changes(self) -> None:
        payload = {
            "metadata": {},
            "report_fields": {
                "reportDate": "2026-05-17", "reportNo": "5", "wellbore": "PCNC-039",
                "rig": "SINOPEC 248", "currentOps": "DRILL AHEAD.",
            },
            "operations": [{
                "from": "03:00", "to": "06:00", "hours": "3", "op_code": "BHA",
                "op_type": "NPT", "operation_details": "NPT A CARGO DE SINOPEC",
            }],
        }
        config = form_server._normalize_ai_extraction_config({
            "auto_execute": True,
            "rules": [{
                "id": "npt-owner", "name": "NPT责任方", "report_type": "drilling",
                "source_section": "operations", "source_field": "operation_details",
                "instruction": "提取责任公司", "target_field": "service_line",
                "output_format": "company", "enabled": True,
            }],
        })
        handler = object.__new__(form_server.FormHandler)

        with (
            patch.object(form_server, "_load_ai_extraction_config", return_value=config),
            patch.object(form_server, "_extraction_jobs_enabled", return_value=True),
            patch.object(form_server, "_schedule_extraction_job") as schedule_job,
            patch.object(form_server, "_invalidate_extraction_jobs") as invalidate_jobs,
            patch.object(form_server, "clear_extraction_results") as clear_results,
        ):
            handler._store_payload(payload, "drilling", from_upload=True)
            handler._store_payload(payload, "drilling", from_upload=True)
            payload["report_fields"]["currentOps"] = "UNRELATED FIELD CHANGED"
            handler._store_payload(payload, "drilling")
            payload["operations"][0]["op_code"] = "CIRCULATING"
            handler._store_payload(payload, "drilling")

        record_id = "drilling:PCNC-039:2026-05-17:5"
        self.assertEqual(schedule_job.call_args_list, [unittest.mock.call(record_id), unittest.mock.call(record_id)])
        invalidate_jobs.assert_called_once_with([record_id])
        clear_results.assert_called_once_with(form_server.DATABASE_PATH, [record_id])

    def test_removing_all_extraction_units_clears_stale_results(self) -> None:
        payload = {
            "metadata": {},
            "report_fields": {
                "reportDate": "2026-05-17", "reportNo": "5", "wellbore": "PCNC-039", "rig": "SINOPEC 248",
            },
            "operations": [{"op_type": "NPT", "operation_details": "WAIT ON SERVICE"}],
        }
        config = form_server._normalize_ai_extraction_config({
            "auto_execute": True,
            "rules": [{
                "id": "npt-owner", "name": "NPT责任方", "report_type": "drilling",
                "source_section": "operations", "source_field": "operation_details",
                "instruction": "提取责任公司", "target_field": "service_line",
                "output_format": "company", "enabled": True,
            }],
        })
        handler = object.__new__(form_server.FormHandler)

        with (
            patch.object(form_server, "_load_ai_extraction_config", return_value=config),
            patch.object(form_server, "_extraction_jobs_enabled", return_value=True),
            patch.object(form_server, "_schedule_extraction_job"),
            patch.object(form_server, "_invalidate_extraction_jobs") as invalidate_jobs,
            patch.object(form_server, "clear_extraction_results") as clear_results,
        ):
            handler._store_payload(payload, "drilling", from_upload=True)
            payload["operations"] = []
            handler._store_payload(payload, "drilling")

        record_id = "drilling:PCNC-039:2026-05-17:5"
        self.assertEqual(payload["metadata"]["extraction_status"], "NOT_REQUIRED")
        invalidate_jobs.assert_called_once_with([record_id])
        clear_results.assert_called_once_with(form_server.DATABASE_PATH, [record_id])

    def test_lm_studio_qwen_disabled_thinking_payload_uses_compatibility_prefill(self) -> None:
        model = {
            "base_url": "http://127.0.0.1:1234/v1",
            "model": "qwen3.5-9b",
            "thinking_mode": "disabled",
        }
        payload = form_server._openai_payload_for_model(model, {
            "model": "qwen3.5-9b",
            "messages": [{"role": "user", "content": "translate"}],
        })

        self.assertEqual(payload["chat_template_kwargs"], {"enable_thinking": False})
        self.assertEqual(payload["messages"][-1], {"role": "assistant", "content": "<think>\n\n</think>\n\n"})

    def test_system_model_temperature_is_fixed_for_deterministic_workflows(self) -> None:
        model = form_server._normalize_ai_model({
            "id": "deterministic-model",
            "name": "Deterministic",
            "api_type": "openai-compatible",
            "base_url": "https://example.test/v1",
            "model": "example-model",
            "temperature": 1.7,
        })

        self.assertNotIn("temperature", model)
        config = form_server._translation_config_for_model(model)
        self.assertEqual(config.openai_temperature, 0.0)

    def test_deepseek_thinking_and_manual_request_options_use_provider_wire_format(self) -> None:
        model = {
            "base_url": "https://api.deepseek.com",
            "model": "deepseek-v4-pro",
            "thinking_mode": "disabled",
            "request_options": {"reasoning_effort": "high"},
        }

        payload = form_server._openai_payload_for_model(model, {
            "model": "deepseek-v4-pro",
            "messages": [{"role": "user", "content": "translate"}],
        })

        self.assertEqual(payload["thinking"], {"type": "disabled"})
        self.assertEqual(payload["reasoning_effort"], "high")
        self.assertNotIn("chat_template_kwargs", payload)

    def test_manual_request_options_cannot_override_core_request_fields(self) -> None:
        options = form_server._normalize_model_request_options({
            "model": "wrong-model",
            "messages": [{"role": "user", "content": "wrong"}],
            "Authorization": "secret",
            "thinking": {"type": "disabled"},
        })

        self.assertEqual(options, {"thinking": {"type": "disabled"}})

    def test_operation_translation_accepts_whitespace_normalization_and_rejects_stale_source(self) -> None:
        source = "WAIT ON SERVICE\nCOMPANY"
        stored_source = "WAIT ON SERVICE COMPANY"
        translation = {
            "source_text": stored_source,
            "source_hash": form_server.hashlib.sha256(stored_source.encode("utf-8")).hexdigest(),
            "translated_text": "等待服务公司",
            "translation_status": "COMPLETED",
        }

        self.assertEqual(form_server._current_operation_translation(source, translation), ("等待服务公司", "COMPLETED"))
        self.assertEqual(form_server._current_operation_translation("DIFFERENT SOURCE", translation), ("", "COMPLETED"))

    def test_production_npt_ranking_includes_rigs_with_zero_npt(self) -> None:
        records = [
            {"rig": "RIG A", "wellbore": "WELL A", "reportDate": "2026-05-22", "report_type": "drilling", "validation_status": "ok"},
            {"rig": "RIG B", "wellbore": "WELL B", "reportDate": "2026-05-22", "report_type": "drilling", "validation_status": "ok"},
        ]
        operations = [
            {**records[0], "hours": 2.0, "op_type": "NPT"},
            {**records[1], "hours": 24.0, "op_type": "P"},
        ]

        with patch.object(form_server, "_filtered_fact_rows", return_value={"records": records, "operations": operations}):
            payload = form_server._production_summary_payload(Path("mysql"), {})

        self.assertEqual(payload["npt_by_rig"], [
            {"label": "RIG A", "hours": 2.0},
            {"label": "RIG B", "hours": 0.0},
        ])

    def test_mysql_production_analytics_uses_only_ready_view_hours(self) -> None:
        record = {
            "record_id": "drilling:well-a:2026-06-01:1", "report_type": "drilling",
            "reportDate": "2026-06-01", "rig": "RIG A", "wellbore": "WELL A",
            "project_id": "5", "project_name": "Project A", "project_contract": "C-1",
            "validation_status": "ok", "event": "DRILLING", "master_match_status": "MATCHED",
        }
        rows = {
            "records": [record],
            "operations": [
                {**record, "hours": 8.0, "op_type": "P", "source_op_type": "P", "statistics_ready": True,
                 "classification_status": "AUTO_CONFIRMED", "reason": "DRILLING"},
                {**record, "hours": 16.0, "op_type": "PENDING", "source_op_type": "NPT", "statistics_ready": False,
                 "classification_status": "PENDING", "reason": "WAITING"},
            ],
            "quality": {"unassigned_count": 3, "ambiguous_count": 1},
        }
        with patch.object(form_server, "load_analytics_view_rows", return_value=rows) as loader, patch.object(
            form_server, "_enrich_operation_translation_rows"
        ), patch.object(form_server, "_enrich_operation_extraction_rows"):
            payload = form_server._production_summary_payload(Path("mysql"), {})

        loader.assert_called_once()
        self.assertEqual(payload["kpis"]["total_hours"], 8.0)
        self.assertEqual(payload["details"][0]["drilling_hours"], 8.0)
        self.assertEqual(payload["unassigned_count"], 3)
        self.assertEqual(payload["unconfirmed_classification_count"], 1)

    def test_monthly_efficiency_report_keeps_pending_fields_out_of_weighted_efficiency(self) -> None:
        source_rows = [
            {
                "job_id": 11, "job_code": "JOB-11", "job_type": "drilling",
                "project_id": 5, "project_name": "Project A", "well_id": 7, "well_name": "WELL-A",
                "well_profile_code": "",
                "rig_name": "RIG-A", "report_start_date": "2026-06-01", "report_end_date": "2026-06-03",
                "report_count": 3, "operation_count": 6, "pending_operation_count": 0,
                "production_hours": 18, "npt_hours": 2, "sc_hours": 4, "pending_hours": 0,
                "events": {"DRILLING_START": "2026-06-01", "DRILLING_END": "2026-06-03"},
                "month_progress_ft": 600, "year_progress_ft": 900, "month_progress_count": 3,
                "year_progress_count": 5, "current_depth_ft": 1500,
            },
            {
                "job_id": 12, "job_code": "JOB-12", "job_type": "workover",
                "project_id": 5, "project_name": "Project A", "well_id": 8, "well_name": "WELL-B",
                "rig_name": "RIG-A", "report_start_date": "2026-06-10", "report_end_date": "2026-06-10",
                "report_count": 1, "operation_count": 2, "pending_operation_count": 1,
                "production_hours": 8, "npt_hours": 0, "sc_hours": 0, "pending_hours": 8,
                "events": {"WORKOVER_START": "2026-06-10"},
            },
        ]
        with patch.object(
            form_server,
            "load_monthly_efficiency_report_rows",
            return_value={"latest_month": "2026-06", "rows": source_rows},
        ) as loader:
            payload = form_server._monthly_efficiency_report_payload(
                Path("mysql"), {"date_from": ["2026-06-01"], "date_to": ["2026-06-30"]}
            )

        loader.assert_called_once_with(
            Path("mysql"), date_from="2026-06-01", date_to="2026-06-30", year_start="2026-01-01"
        )
        self.assertEqual(payload["grain"], "date_range + project + well + job_type + job_id")
        self.assertEqual(payload["kpis"]["weighted_efficiency"], 0.9)
        self.assertEqual(payload["kpis"]["efficiency_job_count"], 1)
        self.assertEqual(payload["details"][0]["actual_cycle_days"], 3)
        self.assertEqual(payload["details"][0]["well_profile"], "")
        self.assertNotIn("sequence_no", payload["details"][0])
        self.assertEqual(payload["details"][0]["source_status"], "AVAILABLE")
        self.assertIsNone(payload["details"][0]["move_setup_hours"])
        self.assertIn("搬安时长", payload["details"][0]["pending_fields"])
        self.assertIsNone(payload["details"][1]["efficiency"])
        self.assertEqual(payload["details"][1]["source_status"], "PARTIAL")

    def test_monthly_efficiency_export_writes_pending_instead_of_zero(self) -> None:
        data = form_server._monthly_efficiency_workbook_bytes([{
            "month": "2026-06", "project_name": "Project A", "wellbore": "WELL-A",
            "job_type_label": "钻井", "production_hours": 12,
            "well_profile": "",
            "npt_hours": 0, "sc_hours": 0, "efficiency": 1.0,
            "move_setup_hours": None, "source_status": "AVAILABLE", "pending_fields": ["搬安时长"],
        }], "2026-06-01", "2026-06-30", "zh")

        workbook = form_server.load_workbook(BytesIO(data), data_only=True)
        sheet = workbook["时效报表"]
        headers = {cell.value: cell.column for cell in sheet[1]}
        self.assertEqual(sheet.cell(2, headers["搬安(h)"]).value, "待定")
        self.assertIsNone(sheet.cell(2, headers["井型"]).value)
        self.assertEqual(sheet.cell(2, headers["作业时效"]).value, 1.0)
        self.assertEqual(workbook.sheetnames, ["时效报表"])

    def test_monthly_efficiency_uses_current_operation_translation_for_chinese_reason(self) -> None:
        source_rows = [{
            "job_id": 11, "job_type": "drilling", "project_id": 5, "project_name": "Project A",
            "well_id": 7, "well_name": "WELL-A", "operation_count": 1, "pending_operation_count": 0,
            "production_hours": 0, "npt_hours": 1, "sc_hours": 0, "pending_hours": 0,
            "nonproductive_description": "ESPERA DE SERVICIO",
            "nonproductive_operations": [{
                "record_id": "record-1", "source_row_no": 3,
                "operation_details": "ESPERA DE SERVICIO",
                "operation_details_normalized": "ESPERA DE SERVICIO",
            }],
        }]
        translation = {
            "record_id": "record-1", "entity_id": "record-1:operations:3",
            "source_text": "ESPERA DE SERVICIO", "translated_text": "等待服务公司",
            "source_hash": hashlib.sha256("ESPERA DE SERVICIO".encode("utf-8")).hexdigest(),
            "translation_status": "COMPLETED",
        }
        with patch.object(
            form_server, "load_monthly_efficiency_report_rows",
            return_value={"latest_month": "2026-06", "rows": source_rows},
        ), patch.object(form_server, "load_operation_translations", return_value=[translation]):
            payload = form_server._monthly_efficiency_report_payload(
                Path("mysql"), {"date_from": ["2026-06-01"], "date_to": ["2026-06-30"]}
            )

        self.assertEqual(payload["details"][0]["nonproductive_description"], "ESPERA DE SERVICIO")
        self.assertEqual(payload["details"][0]["nonproductive_description_zh"], "等待服务公司")

    def test_monthly_efficiency_export_omits_removed_source_columns_and_uses_chinese_reason(self) -> None:
        data = form_server._monthly_efficiency_workbook_bytes([{
            "month": "2026-06", "project_name": "Project A", "wellbore": "WELL-A",
            "job_type_label": "钻井",
            "nonproductive_description": "ESPERA DE SERVICIO",
            "nonproductive_description_zh": "等待服务公司",
            "other_remarks": "不应导出", "report_count": 3, "source_status": "AVAILABLE",
            "pending_fields": ["搬安时长"],
        }], "2026-06-01", "2026-06-30", "zh")

        workbook = form_server.load_workbook(BytesIO(data), data_only=True)
        sheet = workbook["时效报表"]
        headers = {cell.value: cell.column for cell in sheet[1]}
        self.assertEqual(sheet.cell(2, headers["非生产原因"]).value, "等待服务公司")
        self.assertNotIn("日报其他备注", headers)
        self.assertNotIn("来源日报数", headers)
        self.assertNotIn("时效数据状态", headers)
        self.assertNotIn("待定字段", headers)
        self.assertNotIn("批次", headers)

    def test_monthly_efficiency_default_date_range_is_unlimited(self) -> None:
        with patch.object(
            form_server,
            "load_monthly_efficiency_report_rows",
            return_value={"available_date_from": "2026-04-01", "available_date_to": "2026-07-20", "rows": []},
        ) as loader:
            payload = form_server._monthly_efficiency_report_payload(Path("mysql"), {})

        loader.assert_called_once_with(Path("mysql"), date_from="", date_to="", year_start="")
        self.assertEqual(payload["date_from"], "")
        self.assertEqual(payload["date_to"], "")
        self.assertEqual(payload["filters"]["available_date_from"], "2026-04-01")

    def test_monthly_efficiency_export_follows_spanish_language_and_visible_columns(self) -> None:
        data = form_server._monthly_efficiency_workbook_bytes([{
            "project_name": "Proyecto A", "wellbore": "POZO-A", "job_type": "drilling",
            "job_type_label": "钻井",
            "nonproductive_description": "ESPERA DE SERVICIO",
            "nonproductive_description_zh": "等待服务公司",
        }], "", "", "es")

        workbook = form_server.load_workbook(BytesIO(data), data_only=True)
        sheet = workbook["Reporte eficiencia"]
        headers = {cell.value: cell.column for cell in sheet[1]}
        self.assertEqual(len(headers), 30)
        self.assertEqual(sheet.cell(2, headers["Especialidad"]).value, "Perforación")
        self.assertEqual(sheet.cell(2, headers["Motivo no productivo"]).value, "ESPERA DE SERVICIO")
        self.assertNotIn("统计月份", headers)
        self.assertNotIn("合同号", headers)
        self.assertNotIn("Lote", headers)

    def test_ai_extraction_rules_validate_source_and_target_fields(self) -> None:
        config = form_server._normalize_ai_extraction_config({
            "rules": [
                {
                    "id": "npt-owner",
                    "name": "NPT责任方",
                    "report_type": "drilling",
                    "source_section": "operations",
                    "source_field": "operation_details",
                    "condition": "仅NPT",
                    "instruction": "提取责任公司",
                    "target_field": "service_line",
                    "output_format": "company",
                    "enabled": True,
                },
                {
                    "id": "invalid-target",
                    "name": "覆盖井号",
                    "report_type": "drilling",
                    "source_section": "report_fields",
                    "source_field": "currentOps",
                    "instruction": "错误规则",
                    "target_field": "wellbore",
                },
            ],
        })

        self.assertEqual(len(config["rules"]), 1)
        self.assertEqual(config["rules"][0]["target_field"], "service_line")
        self.assertEqual(config["rules"][0]["output_format"], "text")
        self.assertEqual(
            [item["value"] for item in config["catalog"]["output_formats"]],
            ["text", "number", "date"],
        )
        self.assertIn("target_fields", config["catalog"])

    def test_ai_extraction_supports_grouped_multi_source_fields_and_all_modules(self) -> None:
        config = form_server._normalize_ai_extraction_config({
            "rules": [{
                "id": "monthly-incident-summary",
                "name": "月度事故摘要",
                "report_type": "all",
                "source_fields": [
                    {"section": "report_fields", "field": "incidentComments"},
                    {"section": "report_fields", "field": "otherRemarks"},
                    {"section": "operations", "field": "operation_details"},
                    {"section": "perforation_intervals", "field": "comments"},
                ],
                "instruction": "总结事故和停待情况",
                "target_field": "monthly_incident_summary",
                "target_field_label": "月度事故摘要",
                "output_format": "text",
                "enabled": True,
            }],
        })

        rule = config["rules"][0]
        # The all-report rule keeps fields that apply to any selected report
        # type; unavailable sections are simply empty for that report.
        self.assertEqual(
            rule["source_fields"],
            [
                {"section": "report_fields", "field": "incidentComments", "report_types": ["drilling"]},
                {"section": "report_fields", "field": "otherRemarks", "report_types": ["drilling", "completion", "workover"]},
                {"section": "operations", "field": "operation_details", "report_types": ["drilling", "completion", "workover"]},
                {"section": "perforation_intervals", "field": "comments", "report_types": ["completion", "workover"]},
            ],
        )
        sections = {item["value"] for item in config["catalog"]["report_types"][0]["sections"]}
        self.assertTrue({
            "report_fields", "operations", "survey_data", "bha_components", "fluid_losses",
            "bulks", "mud_products", "perforation_intervals",
        }.issubset(sections))
        custom = next(item for item in config["catalog"]["target_fields"] if item["value"] == "monthly_incident_summary")
        self.assertEqual(custom["label"], "月度事故摘要")

    def test_ai_extraction_catalog_uses_chinese_labels_without_changing_field_codes(self) -> None:
        catalog = form_server._ai_extraction_catalog()
        actual_reports = [item for item in catalog["report_types"] if item["value"] != "all"]
        fields = [field for report in actual_reports for section in report["sections"] for field in section["fields"]]

        self.assertTrue(fields)
        self.assertTrue(all(field["value"] in form_server.FIELD_LABELS_ZH for field in fields))
        self.assertTrue(all(field["label"] == form_server.FIELD_LABELS_ZH[field["value"]] for field in fields))
        self.assertTrue(all(any("\u4e00" <= char <= "\u9fff" for char in field["label"]) for field in fields))
        self.assertIn({"value": "reportDate", "label": "报告日期"}, fields)
        self.assertIn({"value": "afeNumber", "label": "AFE编号"}, fields)

    def test_ai_extraction_record_target_combines_selected_fields_into_one_unit(self) -> None:
        payload = {
            "report_fields": {"currentOps": "DRILL AHEAD", "otherRemarks": "NO INCIDENT"},
            "operations": [
                {"from": "00:00", "to": "06:00", "operation_details": "DRILLING"},
                {"from": "06:00", "to": "08:00", "operation_details": "WAIT ON SERVICE"},
            ],
        }
        rule = {
            "source_fields": [
                {"section": "report_fields", "field": "currentOps"},
                {"section": "report_fields", "field": "otherRemarks"},
                {"section": "operations", "field": "operation_details"},
            ],
            "target_field": "monthly_incident_summary",
        }

        units = form_server._ai_extraction_units(payload, rule)

        self.assertEqual(len(units), 1)
        self.assertEqual(units[0]["source_section"], "multiple")
        self.assertIn("DRILL AHEAD", units[0]["prompt_text"])
        self.assertIn("WAIT ON SERVICE", units[0]["prompt_text"])

    def test_ai_extraction_admin_uses_grouped_checkbox_sources_and_custom_targets(self) -> None:
        script = Path(__file__).parents[1].joinpath("web_form", "admin.js").read_text(encoding="utf-8")

        self.assertIn('name="aiExtractionSourceField"', script)
        self.assertIn("aiExtractionSourceFieldPicker", script)
        self.assertIn('class="ai-extraction-source-dropdown"', script)
        self.assertIn("data-ai-extraction-source-preview", script)
        self.assertIn("ai-extraction-source-report", script)
        self.assertIn("ai-extraction-source-module", script)
        self.assertIn('name="aiExtractionReportType"', script)
        self.assertIn("data-ai-extraction-report-dropdown", script)
        self.assertIn("data-ai-extraction-section-toggle", script)
        self.assertIn("data-ai-extraction-new-target", script)
        self.assertIn('name="aiExtractionCustomTargetActive"', script)
        self.assertNotIn('<option value="__new__"', script)
        self.assertIn('name="aiExtractionCustomTargetCode"', script)
        self.assertIn('name="aiExtractionGrain"', script)
        self.assertIn("ai-extraction-scope-preview", script)

    def test_monthly_well_control_rule_uses_well_job_storage_grain(self) -> None:
        rule = form_server._normalize_ai_extraction_rule({
            **form_server._monthly_well_control_rule(),
            "grain": "team_month",
        }, 0)

        self.assertEqual(rule["grain"], "well_job")
        catalog = form_server._ai_extraction_catalog()
        target = next(item for item in catalog["target_fields"] if item["value"] == "well_control_incident")
        self.assertEqual(target["grain"], "well_job")
        self.assertEqual(target["allowed_grains"], ["well_job"])
        self.assertEqual(
            [item["value"] for item in catalog["grains"]],
            ["detail_row", "report", "well", "well_job", "well_month", "team_month"],
        )

    def test_team_month_well_control_rule_has_its_own_target_field(self) -> None:
        rule = form_server._normalize_ai_extraction_rule(form_server._team_month_well_control_rule(), 0)
        catalog = form_server._ai_extraction_catalog()
        target = next(
            item for item in catalog["target_fields"]
            if item["value"] == "team_month_well_control_incident"
        )

        self.assertEqual(rule["grain"], "team_month")
        self.assertEqual(rule["report_types"], ["drilling", "completion"])
        self.assertEqual(rule["target_field"], "team_month_well_control_incident")
        self.assertEqual(target["grain"], "team_month")
        self.assertEqual(target["allowed_grains"], ["team_month"])

    def test_canonical_monthly_ai_rules_cover_every_narrative_report_field(self) -> None:
        rules = form_server._canonical_monthly_extraction_rules()
        rule_targets = {str(rule["target_field"]): str(rule["grain"]) for rule in rules}

        self.assertEqual(rule_targets, {
            "well_control_incident": "well_job",
            "team_month_well_control_incident": "team_month",
            "accident_waiting": "well_job",
            "basic_monthly_remarks": "well_job",
            "nonproductive_description": "well_month",
            "efficiency_monthly_remarks": "well_month",
            "workload_monthly_remarks": "team_month",
        })
        self.assertTrue(all(rule["enabled"] for rule in rules))
        self.assertTrue(all(rule["source_fields"] for rule in rules))

    def test_monthly_nonproductive_rule_only_reads_confirmed_npt_operations(self) -> None:
        rule = form_server._normalize_ai_extraction_rule(form_server._monthly_npt_description_rule(), 0)
        self.assertEqual(rule["source_fields"], [{
            "section": "operations", "field": "operation_details",
            "report_types": ["drilling", "completion", "workover"],
        }])
        payload = {
            "operations": [
                {"hours": 5, "op_type": "P", "operation_details": "WAITING BUT CLASSIFIED P"},
                {"hours": 2, "op_type": "SC", "operation_details": "SPECIAL CONDITION"},
                {"hours": 1.5, "op_type": "SC", "operation_details": "CONFIRMED NPT EVENT"},
            ],
        }
        classifications = {
            ("r-1", 1): {"source_op_type": "P", "confirmed_op_type": "P", "confirmation_status": "AUTO_CONFIRMED"},
            ("r-1", 2): {"source_op_type": "SC", "confirmed_op_type": "SC", "confirmation_status": "CONFIRMED"},
            ("r-1", 3): {"source_op_type": "SC", "confirmed_op_type": "NPT", "confirmation_status": "CONFIRMED"},
        }

        unit, sources = form_server._aggregate_rule_record_unit(
            payload, rule, "drilling", "r-1", classifications,
        )

        self.assertEqual(sources, [{"section": "operations", "field": "operation_details"}])
        self.assertIsNotNone(unit)
        self.assertNotIn("WAITING BUT CLASSIFIED P", unit["source_text"])
        self.assertNotIn("SPECIAL CONDITION", unit["source_text"])
        self.assertIn("CONFIRMED NPT EVENT", unit["source_text"])
        self.assertEqual(unit["source_entries"][0]["row_no"], 3)

    def test_monthly_nonproductive_rule_has_no_unit_without_confirmed_npt(self) -> None:
        rule = form_server._normalize_ai_extraction_rule(form_server._monthly_npt_description_rule(), 0)
        payload = {"operations": [{"hours": 4, "op_type": "P", "operation_details": "WAIT ON INSTRUCTIONS"}]}
        classifications = {
            ("r-1", 1): {"source_op_type": "P", "confirmed_op_type": "P", "confirmation_status": "AUTO_CONFIRMED"},
        }

        unit, _ = form_server._aggregate_rule_record_unit(
            payload, rule, "drilling", "r-1", classifications,
        )

        self.assertIsNone(unit)

    def test_monthly_narrative_empty_variants_normalize_to_no(self) -> None:
        self.assertEqual(form_server._normalize_monthly_narrative_result("NPT: 无"), "无")
        self.assertEqual(form_server._normalize_monthly_narrative_result("无特别说明。"), "无")
        self.assertEqual(form_server._normalize_monthly_narrative_result("等待指令5小时"), "等待指令5小时")

    def test_extraction_task_scope_maps_completion_into_drilling_monthly_profession(self) -> None:
        scope = form_server._extraction_task_scope("team_month", {
            "record_id": "r-1", "report_type": "completion", "report_date": "2026-07-15",
            "project_id": 3, "well_id": 7, "team_id": 9,
        })

        self.assertIsNotNone(scope)
        self.assertEqual(scope["profession"], "drilling")
        self.assertEqual(scope["period_start"], "2026-07-01")
        self.assertEqual(scope["period_end"], "2026-07-31")
        self.assertEqual(scope["scope_key"], "TEAM_MONTH:9:drilling:2026-07")
        self.assertNotIn("project_id", scope)
        self.assertNotIn("well_id", scope)
        self.assertNotIn("job_sequence_no", scope)

    def test_extraction_task_scopes_discard_dimensions_from_other_grains(self) -> None:
        record = {
            "record_id": "r-1", "report_type": "drilling", "report_date": "2026-07-15",
            "project_id": 3, "job_id": 5, "job_sequence_no": 2, "well_id": 7, "team_id": 9,
        }

        job_scope = form_server._extraction_task_scope("well_job", record)
        month_scope = form_server._extraction_task_scope("well_month", record)

        self.assertEqual(job_scope["scope_key"], "WELL_JOB:3:7:2")
        self.assertNotIn("team_id", job_scope)
        self.assertNotIn("period_start", job_scope)
        self.assertNotIn("period_end", job_scope)
        self.assertEqual(month_scope["scope_key"], "WELL_MONTH:3:7:drilling:2026-07")
        self.assertNotIn("team_id", month_scope)
        self.assertNotIn("job_id", month_scope)
        self.assertNotIn("job_sequence_no", month_scope)

    def test_extraction_task_scope_requires_aggregate_master_dimensions(self) -> None:
        scope = form_server._extraction_task_scope("well_job", {
            "record_id": "r-1", "report_type": "drilling", "report_date": "2026-07-15",
            "project_id": 0, "well_id": 0, "job_sequence_no": 0,
        })

        self.assertIsNone(scope)

    def test_admin_extraction_queue_uses_unified_task_language(self) -> None:
        script = Path(__file__).parents[1].joinpath("web_form", "admin.js").read_text(encoding="utf-8")

        self.assertIn("源日报覆盖", script)
        self.assertIn("单井作业周期", script)
        self.assertIn("data-ai-extraction-task", script)
        self.assertIn("生成全部历史任务", script)
        self.assertIn("系统执行门控", script)

    def test_ai_extraction_target_fields_belong_to_exactly_one_grain(self) -> None:
        catalog = form_server._ai_extraction_catalog([{
            "field_code": "monthly_custom_summary",
            "field_label": "月度自定义摘要",
            "allowed_grains": ["well_month", "team_month"],
        }])

        for target in catalog["target_fields"]:
            self.assertIn(target["grain"], {item["value"] for item in catalog["grains"]})
            self.assertEqual(target["allowed_grains"], [target["grain"]])
        custom = next(item for item in catalog["target_fields"] if item["value"] == "monthly_custom_summary")
        self.assertEqual(custom["grain"], "well_month")

    def test_ai_extraction_sources_can_differ_by_report_type(self) -> None:
        rule = form_server._normalize_ai_extraction_rule({
            "id": "typed-sources",
            "name": "分类来源字段",
            "report_types": ["drilling", "workover"],
            "source_fields": [
                {"section": "report_fields", "field": "incidentComments", "report_types": ["drilling"]},
                {"section": "report_fields", "field": "description", "report_types": ["workover"]},
            ],
            "instruction": "分别读取来源字段",
            "target_field": "remarks",
        }, 0)

        self.assertIsNotNone(rule)
        self.assertEqual(
            form_server._ai_extraction_rule_sources(rule, "drilling"),
            [{"section": "report_fields", "field": "incidentComments"}],
        )
        self.assertEqual(
            form_server._ai_extraction_rule_sources(rule, "workover"),
            [{"section": "report_fields", "field": "description"}],
        )

    def test_well_job_well_control_rule_returns_none_without_signal_and_persists_lineage(self) -> None:
        rule = form_server._normalize_ai_extraction_rule(form_server._monthly_well_control_rule(), 0)
        records = [{
            "record_id": "drilling:W-1:2026-07-01:1", "report_type": "drilling",
            "report_date": "2026-07-01", "project_id": 10, "well_id": 7, "well_name": "W-1",
            "job_sequence_no": 1, "team_id": 9, "team_name": "队伍9",
        }]
        payloads = {
            records[0]["record_id"]: {
                "metadata": {"report_type": "drilling"},
                "report_fields": {"summary24h": "正常钻进，无异常。"},
                "operations": [],
            },
        }
        scope = {"grain": "well_job", "project_id": 10, "well_id": 7, "job_sequence_no": 1}
        with (
            patch.object(form_server, "_enabled_aggregate_extraction_rules", return_value=[rule]),
            patch.object(form_server, "list_aggregate_scope_report_records", return_value=records),
            patch.object(form_server, "load_report_payloads", return_value=payloads),
            patch.object(form_server, "load_aggregate_extraction_results", return_value=[]),
            patch.object(form_server, "save_aggregate_extraction_results") as save_results,
            patch.object(form_server, "save_extraction_result_sources") as save_sources,
            patch.object(form_server, "_load_ai_extraction_config", return_value={"version": "rules-v1"}),
            patch.object(form_server, "_run_ai_extraction_test") as run_model,
        ):
            values = form_server._aggregate_extraction_values(Path("mysql"), [scope])

        self.assertEqual(values, {"WELL_JOB:10:7:1": {"well_control_incident": "无"}})
        run_model.assert_not_called()
        saved = save_results.call_args.args[1][0]
        self.assertEqual(saved["grain"], "well_job")
        self.assertEqual(saved["source_record_ids"], [records[0]["record_id"]])
        self.assertEqual(saved["result_text"], "无")
        save_sources.assert_called_once()

    def test_well_job_well_control_rule_uses_model_for_actual_signal(self) -> None:
        rule = form_server._normalize_ai_extraction_rule(form_server._monthly_well_control_rule(), 0)
        records = [{
            "record_id": "completion:W-2:2026-07-02:2", "report_type": "completion",
            "report_date": "2026-07-02", "project_id": 12, "well_id": 6, "well_name": "W-2",
            "job_sequence_no": 2, "team_id": 8, "team_name": "队伍8",
        }]
        payloads = {
            records[0]["record_id"]: {
                "metadata": {"report_type": "completion"},
                "report_fields": {"summary24h": "SE REGISTRA INFLUJO DURANTE LA OPERACIÓN."},
                "operations": [],
            },
        }
        scope = {"grain": "well_job", "project_id": 12, "well_id": 6, "job_sequence_no": 2}
        with (
            patch.object(form_server, "_enabled_aggregate_extraction_rules", return_value=[rule]),
            patch.object(form_server, "list_aggregate_scope_report_records", return_value=records),
            patch.object(form_server, "load_report_payloads", return_value=payloads),
            patch.object(form_server, "load_aggregate_extraction_results", return_value=[]),
            patch.object(form_server, "save_aggregate_extraction_results"),
            patch.object(form_server, "save_extraction_result_sources"),
            patch.object(form_server, "_load_ai_extraction_config", return_value={"version": "rules-v1"}),
            patch.object(form_server, "_extraction_model", return_value={"id": "model-1"}),
            patch.object(form_server, "_run_ai_extraction_test", return_value={"result": "W-2井发生井涌1次"}) as run_model,
        ):
            values = form_server._aggregate_extraction_values(Path("mysql"), [scope])

        self.assertEqual(values, {"WELL_JOB:12:6:2": {"well_control_incident": "W-2井发生井涌1次。"}})
        self.assertIn("INFLUJO", run_model.call_args.args[2])

    def test_team_month_well_control_rule_persists_by_team_profession_and_month(self) -> None:
        rule = form_server._normalize_ai_extraction_rule(form_server._team_month_well_control_rule(), 0)
        records = [{
            "record_id": "drilling:W-1:2026-07-01:1", "report_type": "drilling",
            "report_date": "2026-07-01", "well_id": 7, "well_name": "W-1",
            "team_id": 9, "team_name": "队伍9",
        }]
        payloads = {records[0]["record_id"]: {
            "metadata": {"report_type": "drilling"},
            "report_fields": {"summary24h": "正常钻进，作业平稳。"},
            "operations": [],
        }}
        scope = {
            "grain": "team_month", "team_id": 9, "profession": "drilling",
            "period_start": "2026-07-01", "period_end": "2026-07-31",
        }
        with (
            patch.object(form_server, "_enabled_aggregate_extraction_rules", return_value=[rule]),
            patch.object(form_server, "list_aggregate_scope_report_records", return_value=records),
            patch.object(form_server, "load_report_payloads", return_value=payloads),
            patch.object(form_server, "load_aggregate_extraction_results", return_value=[]),
            patch.object(form_server, "save_aggregate_extraction_results") as save_results,
            patch.object(form_server, "save_extraction_result_sources"),
            patch.object(form_server, "_load_ai_extraction_config", return_value={"version": "rules-v1"}),
            patch.object(form_server, "_run_ai_extraction_test") as run_model,
        ):
            values = form_server._aggregate_extraction_values(Path("mysql"), [scope])

        self.assertEqual(values, {
            "TEAM_MONTH:9:drilling:2026-07": {"team_month_well_control_incident": "无"},
        })
        run_model.assert_not_called()
        saved = save_results.call_args.args[1][0]
        self.assertEqual(saved["grain"], "team_month")
        self.assertEqual(saved["target_field"], "team_month_well_control_incident")

    def test_monthly_report_does_not_display_stale_aggregate_result(self) -> None:
        rule = form_server._normalize_ai_extraction_rule(form_server._monthly_well_control_rule(), 0)
        records = [{
            "record_id": "drilling:W-1:2026-07-01:1", "report_type": "drilling",
            "report_date": "2026-07-01", "project_id": 10, "well_id": 7,
            "well_name": "W-1", "job_sequence_no": 1,
        }]
        payloads = {records[0]["record_id"]: {
            "metadata": {"report_type": "drilling"},
            "report_fields": {"summary24h": "正常钻进，无异常。"}, "operations": [],
        }}
        stale = {
            "scope_key": "WELL_JOB:10:7:1", "rule_id": rule["id"],
            "target_field": rule["target_field"], "extraction_status": "COMPLETED",
            "result_text": "不应显示的旧结果", "source_hash": "old-source",
            "rule_version": "old-rules",
        }
        with (
            patch.object(form_server, "_enabled_aggregate_extraction_rules", return_value=[rule]),
            patch.object(form_server, "list_aggregate_scope_report_records", return_value=records),
            patch.object(form_server, "load_report_payloads", return_value=payloads),
            patch.object(form_server, "load_aggregate_extraction_results", return_value=[stale]),
            patch.object(form_server, "_load_ai_extraction_config", return_value={"version": "rules-v2"}),
        ):
            values = form_server._aggregate_extraction_values(
                Path("mysql"),
                [{"grain": "well_job", "project_id": 10, "well_id": 7, "job_sequence_no": 1}],
                execute_missing=False,
            )

        self.assertEqual(values, {})

    def test_stopped_extraction_is_available_to_continue(self) -> None:
        self.assertTrue(form_server._extraction_record_needs_processing(
            {"extraction_status": "STOPPED", "extraction_version": "rules-v1"},
            "rules-v1",
        ))
        self.assertFalse(form_server._extraction_record_needs_processing(
            {"extraction_status": "IN_PROGRESS", "extraction_version": "rules-v1"},
            "rules-v1",
        ))

    def test_stop_translation_invalidates_jobs_and_rotates_executor(self) -> None:
        old_executor = Mock()
        new_executor = Mock()
        original_executor = form_server.TRANSLATION_EXECUTOR
        form_server.TRANSLATION_EXECUTOR = old_executor
        self.addCleanup(setattr, form_server, "TRANSLATION_EXECUTOR", original_executor)
        form_server.TRANSLATION_JOB_GENERATIONS["report-1"] = 4
        self.addCleanup(form_server.TRANSLATION_JOB_GENERATIONS.pop, "report-1", None)

        with (
            patch.object(form_server, "list_ai_job_status", return_value=[{
                "record_id": "report-1", "status": "IN_PROGRESS", "progress": "42",
            }]),
            patch.object(form_server, "update_record_translation_status") as update_status,
            patch.object(form_server, "_write_translation_metric"),
            patch.object(form_server, "ThreadPoolExecutor", return_value=new_executor),
        ):
            stopped = form_server._stop_active_translation_jobs()

        self.assertEqual(stopped, 1)
        self.assertIs(form_server.TRANSLATION_EXECUTOR, new_executor)
        self.assertEqual(form_server.TRANSLATION_JOB_GENERATIONS["report-1"], 5)
        old_executor.shutdown.assert_called_once_with(wait=False, cancel_futures=True)
        update_status.assert_called_once_with(
            form_server.DATABASE_PATH, "report-1", status="STOPPED", progress="42", error="",
        )

    def test_ai_job_status_snapshot_is_lightweight_and_reports_processing_count(self) -> None:
        records = [
            {"record_id": "report-1", "translation_status": "IN_PROGRESS", "translation_progress": "38", "translation_updated_at": "now"},
            {"record_id": "report-2", "translation_status": "COMPLETED", "translation_progress": "100", "translation_updated_at": "before"},
        ]
        lightweight_records = [
            {"record_id": row["record_id"], "status": row["translation_status"], "progress": row["translation_progress"], "updated_at": row["translation_updated_at"]}
            for row in records
        ]
        with patch.object(form_server, "list_ai_job_status", return_value=lightweight_records):
            snapshot = form_server._ai_job_status_snapshot("translation")

        self.assertEqual(snapshot["processing_count"], 1)
        self.assertEqual(snapshot["records"][0], {
            "record_id": "report-1", "status": "IN_PROGRESS", "progress": "38", "error": "", "updated_at": "now",
        })

    def test_translation_telemetry_promotes_provider_cache_usage_to_metrics(self) -> None:
        emit = form_server._translation_telemetry("report-1", 3, "zh-CN")
        with (
            patch.object(form_server, "_write_translation_debug_log") as write_debug,
            patch.object(form_server, "_write_translation_metric") as write_metric,
        ):
            emit({
                "event": "model_wire_response",
                "provider": "openai-compatible",
                "model": "translation-model",
                "prompt_prefix_hash": "abc123",
                "elapsed_ms": 240,
                "usage_metrics": {
                    "input_tokens": 100,
                    "cached_input_tokens": 80,
                    "prompt_cache_hit": True,
                    "prompt_cache_hit_ratio": 0.8,
                },
            })

        write_debug.assert_called_once()
        write_metric.assert_called_once_with(
            "model_usage",
            record_id="report-1",
            generation=3,
            language="zh-CN",
            provider="openai-compatible",
            model="translation-model",
            prompt_prefix_hash="abc123",
            elapsed_ms=240,
            input_tokens=100,
            cached_input_tokens=80,
            prompt_cache_hit=True,
            prompt_cache_hit_ratio=0.8,
        )

    def test_ai_job_monitor_keeps_separate_translation_and_extraction_streams(self) -> None:
        original_path = form_server.AI_JOB_MONITOR_PATH
        form_server.AI_JOB_MONITOR_PATH = Path(self.repository_tmp.name) / "ai_job_monitor.jsonl"
        self.addCleanup(setattr, form_server, "AI_JOB_MONITOR_PATH", original_path)

        form_server._write_ai_job_monitor("translation", "request", record_id="report-1", source_preview="DRILL AHEAD")
        form_server._write_ai_job_monitor("translation", "response", record_id="report-1", response_preview="继续钻进")
        form_server._write_ai_job_monitor("extraction", "response", record_id="report-2", response_preview="SLB")

        translation = form_server._ai_job_monitor_snapshot("translation", 10)
        extraction = form_server._ai_job_monitor_snapshot("extraction", 10)
        self.assertEqual([row["event"] for row in translation["events"]], ["request", "response"])
        self.assertEqual(extraction["events"][0]["record_id"], "report-2")

    def test_translation_debug_log_retention_limits_age_count_and_rotated_file(self) -> None:
        debug_path = Path(self.repository_tmp.name) / "translation_debug_logs.jsonl"
        rotated_path = debug_path.with_suffix(debug_path.suffix + ".1")
        now = datetime(2026, 7, 13, 12, 0, tzinfo=timezone.utc)
        rows = [
            {"time": "2026-07-01T12:00:00Z", "event": "old", "index": 0},
            *[
                {"time": f"2026-07-{day:02d}T12:00:00Z", "event": "recent", "index": day}
                for day in range(9, 14)
            ],
        ]
        debug_path.write_text("".join(json.dumps(row) + "\n" for row in rows), encoding="utf-8")
        rotated_path.write_text("old rotated data", encoding="utf-8")

        with (
            patch.object(form_server, "TRANSLATION_DEBUG_LOG_PATH", debug_path),
            patch.object(form_server, "TRANSLATION_DEBUG_RETENTION_DAYS", 7),
            patch.object(form_server, "TRANSLATION_DEBUG_MAX_ENTRIES", 3),
            patch.object(form_server, "TRANSLATION_DEBUG_MAX_BYTES", 1024 * 1024),
        ):
            result = form_server._prune_translation_debug_logs(now=now)

        retained = [json.loads(line) for line in debug_path.read_text(encoding="utf-8").splitlines()]
        self.assertEqual([row["index"] for row in retained], [11, 12, 13])
        self.assertEqual(result["after"], 3)
        self.assertFalse(rotated_path.exists())

    def test_default_ai_extraction_rule_is_disabled_until_reviewed(self) -> None:
        config = form_server._default_ai_extraction_config()

        self.assertEqual(config["rules"][0]["id"], "npt-service-line")
        self.assertFalse(config["rules"][0]["enabled"])

    def test_ai_extraction_source_includes_row_context(self) -> None:
        source, count = form_server._ai_extraction_source_from_payload({
            "operations": [
                {"from": "08:00", "to": "09:30", "hours": "1.5", "op_type": "NPT", "operation_details": "WAIT ON SERVICE COMPANY"},
                {"from": "09:30", "to": "10:00", "hours": "0.5", "op_type": "P", "operation_details": "RESUME DRILLING"},
            ],
        }, {"source_section": "operations", "source_field": "operation_details"})

        self.assertEqual(count, 2)
        self.assertIn('"op_type": "NPT"', source)
        self.assertIn("WAIT ON SERVICE COMPANY", source)

    def test_ai_extraction_version_is_stable_and_changes_with_rules(self) -> None:
        raw = {
            "auto_execute": True,
            "rules": [{
                "id": "npt-owner", "name": "NPT责任方", "report_type": "drilling",
                "source_section": "operations", "source_field": "operation_details",
                "instruction": "提取责任公司", "target_field": "service_line",
                "output_format": "company", "enabled": True,
            }],
        }
        first = form_server._normalize_ai_extraction_config(raw)
        second = form_server._normalize_ai_extraction_config(raw)
        switch_only = form_server._normalize_ai_extraction_config({**raw, "auto_execute": False})
        changed = form_server._normalize_ai_extraction_config({
            **raw,
            "rules": [{**raw["rules"][0], "instruction": "提取责任公司；没有证据则返回空值"}],
        })

        self.assertEqual(first["version"], second["version"])
        self.assertEqual(first["version"], switch_only["version"])
        self.assertNotEqual(first["version"], changed["version"])

    def test_ai_extraction_units_only_include_npt_rows_for_service_line(self) -> None:
        units = form_server._ai_extraction_units({"operations": [
            {"op_type": "P", "operation_details": "DRILLING"},
            {"op_type": "NPT", "from": "08:00", "to": "09:00", "operation_details": "WAIT ON SINOPEC"},
        ]}, {
            "source_section": "operations", "source_field": "operation_details",
            "target_field": "service_line", "condition": "仅处理 NPT",
        })

        self.assertEqual(len(units), 1)
        self.assertEqual(units[0]["source_row_no"], 2)
        self.assertIn("WAIT ON SINOPEC", units[0]["prompt_text"])

    def test_ai_extraction_unit_uses_only_matching_row_translation(self) -> None:
        record_id = "drilling:PCNC-039:2026-05-17:5"
        source = "NPT A CARGO DE SINOPEC"
        payload = {
            "metadata": {"record_id": record_id},
            "report_fields": {"rig": "SINOPEC 248"},
            "operations": [{"op_type": "NPT", "operation_details": source}],
            "translation_content": [{
                "entity_id": f"{record_id}:operations:1",
                "field_code": "operations.operation_details",
                "target_language": "zh-CN",
                "source_hash": form_server.source_hash(source),
                "translated_text": "NPT由SINOPEC负责",
                "translation_status": "COMPLETED",
            }],
        }
        rule = {
            "source_section": "operations", "source_field": "operation_details",
            "target_field": "service_line", "condition": "仅处理 NPT",
        }

        unit = form_server._ai_extraction_units(payload, rule)[0]

        self.assertEqual(unit["translated_text"], "NPT由SINOPEC负责")
        self.assertIn("作业原文", unit["prompt_text"])
        self.assertIn("中文参考译文（仅辅助理解）", unit["prompt_text"])
        self.assertIn("以作业原文为准", unit["prompt_text"])

    def test_ai_extraction_unit_rejects_stale_translation(self) -> None:
        record_id = "drilling:PCNC-039:2026-05-17:5"
        payload = {
            "metadata": {"record_id": record_id},
            "operations": [{"op_type": "NPT", "operation_details": "CURRENT SOURCE"}],
            "translation_content": [{
                "entity_id": f"{record_id}:operations:1",
                "field_code": "operations.operation_details",
                "target_language": "zh-CN",
                "source_hash": form_server.source_hash("OLD SOURCE"),
                "translated_text": "错误旧译文",
                "translation_status": "COMPLETED",
            }],
        }
        rule = {
            "source_section": "operations", "source_field": "operation_details",
            "target_field": "service_line", "condition": "仅处理 NPT",
        }

        unit = form_server._ai_extraction_units(payload, rule)[0]

        self.assertEqual(unit["translated_text"], "")
        self.assertNotIn("错误旧译文", unit["prompt_text"])

    def test_extraction_input_signature_includes_prompt_context_and_row_mapping(self) -> None:
        rule = {
            "id": "npt-owner", "source_section": "operations", "source_field": "operation_details",
            "target_field": "service_line", "condition": "仅处理 NPT",
        }
        original = {"report_fields": {"rig": "SINOPEC 248"}, "operations": [
            {"op_type": "NPT", "op_code": "BHA", "operation_details": "WAIT ON SERVICE"},
        ]}
        context_changed = {"report_fields": {"rig": "SINOPEC 129"}, "operations": original["operations"]}
        row_inserted = {"report_fields": original["report_fields"], "operations": [
            {"op_type": "P", "operation_details": "DRILL"},
            original["operations"][0],
        ]}

        original_signature, count = form_server._extraction_input_signature(original, [rule])

        self.assertEqual(count, 1)
        self.assertNotEqual(original_signature, form_server._extraction_input_signature(context_changed, [rule])[0])
        self.assertNotEqual(original_signature, form_server._extraction_input_signature(row_inserted, [rule])[0])

    def test_extraction_waits_while_translation_is_running(self) -> None:
        with (
            patch.object(form_server, "_translation_is_running_for_extraction", return_value=True),
            patch.object(form_server, "_retry_extraction_job_after_translation") as retry,
            patch.object(form_server, "background_job_lock") as job_lock,
        ):
            form_server._run_extraction_job("report-1", 7)

        retry.assert_called_once_with("report-1", 7, False)
        job_lock.assert_not_called()

    def test_translation_revision_requeues_related_extraction(self) -> None:
        payload = {
            "metadata": {"record_id": "report-1", "report_type": "drilling", "extraction_status": "COMPLETED"},
            "operations": [{"op_type": "NPT", "operation_details": "WAIT ON SERVICE"}],
        }
        rule = {
            "id": "npt-owner", "report_type": "drilling", "source_section": "operations",
            "source_field": "operation_details", "target_field": "service_line", "condition": "仅处理 NPT",
        }
        with (
            patch.object(form_server, "load_report_payload", return_value=payload),
            patch.object(form_server, "_enabled_extraction_rules", return_value=[rule]),
            patch.object(form_server, "_load_ai_extraction_config", return_value={"auto_execute": True, "version": "rules-v2"}),
            patch.object(form_server, "_extraction_jobs_enabled", return_value=True),
            patch.object(form_server, "_invalidate_extraction_jobs") as invalidate,
            patch.object(form_server, "clear_extraction_results") as clear,
            patch.object(form_server, "update_record_extraction_status") as update_status,
            patch.object(form_server, "_schedule_extraction_job") as schedule,
        ):
            refreshed = form_server._refresh_extraction_after_translation(
                "report-1", changed_field_code="operations.operation_details",
            )

        self.assertTrue(refreshed)
        invalidate.assert_called_once_with(["report-1"])
        clear.assert_called_once_with(form_server.DATABASE_PATH, ["report-1"])
        update_status.assert_called_once_with(
            form_server.DATABASE_PATH, "report-1", status="QUEUED", progress=0, error="", version="rules-v2",
        )
        schedule.assert_called_once_with("report-1")

    def test_all_report_type_rule_accepts_common_operation_field(self) -> None:
        config = form_server._normalize_ai_extraction_config({
            "rules": [{
                "id": "all-npt-owner", "name": "全类型NPT责任方", "report_type": "all",
                "source_section": "operations", "source_field": "operation_details",
                "instruction": "提取责任公司", "target_field": "service_line",
                "output_format": "company", "enabled": True,
            }],
        })

        self.assertEqual(config["rules"][0]["report_type"], "all")
        self.assertEqual(config["rules"][0]["report_types"], ["drilling", "completion", "workover"])
        self.assertEqual(config["catalog"]["report_types"][0]["value"], "all")
        enabled = [config["rules"][0]]
        self.assertTrue(form_server._payload_has_extraction_units(
            {"operations": [{"op_type": "NPT", "operation_details": "NPT A CARGO DE SLB"}]},
            "completion",
            enabled,
        ))

    def test_extraction_rule_supports_multiple_selected_report_types(self) -> None:
        config = form_server._normalize_ai_extraction_config({
            "rules": [{
                "id": "drilling-workover-summary",
                "name": "钻修井摘要",
                "report_types": ["drilling", "workover"],
                "source_fields": [{"section": "operations", "field": "operation_details"}],
                "instruction": "提炼作业摘要",
                "target_field": "remarks",
                "output_format": "text",
                "enabled": True,
            }],
        })

        rule = config["rules"][0]
        self.assertEqual(rule["report_types"], ["drilling", "workover"])
        self.assertTrue(form_server._ai_extraction_rule_applies(rule, "drilling"))
        self.assertTrue(form_server._ai_extraction_rule_applies(rule, "workover"))
        self.assertFalse(form_server._ai_extraction_rule_applies(rule, "completion"))

    def test_explicit_npt_responsibility_is_extracted_and_expanded_to_rig(self) -> None:
        source = "CONTINUA OPERACION; NPT A CARGO DE SINOPEC;"
        payload = {"report_fields": {"rig": "SINOPEC 248"}}

        company = form_server._explicit_responsible_party(source)

        self.assertEqual(company, "SINOPEC")
        self.assertEqual(form_server._normalize_responsible_party(company, payload), "SINOPEC 248")
        self.assertEqual(form_server._normalize_responsible_party("SINOPEC-SLB FRACTURA", payload), "SINOPEC-SLB FRACTURA")

    def test_explicit_npt_responsibility_stops_at_line_boundary(self) -> None:
        source = "NPT A CARGO DE SINOPEC\nIncident Comments:"

        self.assertEqual(form_server._explicit_responsible_party(source), "SINOPEC")

    def test_operation_extraction_status_is_row_specific(self) -> None:
        source = "NPT A CARGO DE SERVICE COMPANY"
        source_hash = hashlib.sha256(source.encode("utf-8")).hexdigest()
        current = {
            "source_hash": source_hash,
            "rule_version": "rules-v2",
            "extraction_status": "COMPLETED",
            "result_text": "SERVICE COMPANY",
            "updated_at": "2026-07-13T10:00:00",
        }

        self.assertEqual(
            form_server._current_operation_extraction(source, current, "rules-v2"),
            ("SERVICE COMPANY", "COMPLETED", "", "2026-07-13T10:00:00"),
        )
        self.assertEqual(form_server._current_operation_extraction(source, {}, "rules-v2"), ("", "PENDING", "", ""))
        self.assertEqual(form_server._current_operation_extraction(source, current, "rules-v3")[1], "STALE")
        self.assertEqual(form_server._current_operation_extraction(f"{source} CHANGED", current, "rules-v2")[1], "STALE")

    def test_npt_confirmation_rows_include_current_responsible_party(self) -> None:
        source = "NPT A CARGO DE SERVICE COMPANY"
        rows = [{
            "record_id": "report-1",
            "row_no": 2,
            "system_op_type": "NPT",
            "operation_details": source,
        }]
        result = {
            "record_id": "report-1",
            "source_section": "operations",
            "source_row_no": 2,
            "target_field": "service_line",
            "source_hash": hashlib.sha256(source.encode("utf-8")).hexdigest(),
            "rule_version": "rules-v2",
            "extraction_status": "COMPLETED",
            "result_text": "SERVICE COMPANY",
        }

        with (
            patch.object(form_server, "load_extraction_results", return_value=[result]),
            patch.object(form_server, "_load_ai_extraction_config", return_value={"version": "rules-v2"}),
        ):
            form_server._enrich_operation_extraction_rows(rows)

        self.assertEqual(rows[0]["service_line"], "SERVICE COMPANY")
        self.assertEqual(rows[0]["extraction_status"], "COMPLETED")

    def test_fault_evidence_maps_directional_tool_brand_to_service_line(self) -> None:
        source = "HALLIBURTON SPERRY REVISA FALLA / COMPORTAMIENTO ANÓMALO DE HERRAMIENTA ICRUISE MWD"

        self.assertEqual(form_server._evidence_responsible_party(source), "HALLIBURTON SPERRY")

    def test_directional_tool_mention_does_not_override_drill_pipe_failure(self) -> None:
        source = "SACA BHA CON ICRUISE Y M/LWD; OBSERVA WASHOUT EN JUNTA DE DRILL PIPE #171"

        self.assertEqual(form_server._evidence_responsible_party(source), "")

    def test_extraction_rule_filters_out_reports_with_only_p_operations(self) -> None:
        rule = {
            "report_type": "drilling", "source_section": "operations",
            "source_field": "operation_details", "target_field": "service_line",
            "condition": "仅处理 NPT", "enabled": True,
        }
        p_only = {"operations": [{"op_type": "P", "operation_details": "DRILLING"}]}
        with_npt = {"operations": [{"op_type": "NPT", "operation_details": "WAIT ON SERVICE"}]}

        self.assertFalse(form_server._payload_has_extraction_units(p_only, "drilling", [rule]))
        self.assertTrue(form_server._payload_has_extraction_units(with_npt, "drilling", [rule]))

    def test_rig_name_normalization_unifies_report_variants(self) -> None:
        self.assertEqual(form_server._normalize_rig_name("00 SINOPEC 248"), "SINOPEC 248")
        self.assertEqual(form_server._normalize_rig_name("W248"), "SINOPEC 248")

    def test_production_filter_rigs_uses_records_and_rejects_placeholders(self) -> None:
        with patch.object(form_server, "list_records", return_value=[
            {"rig": "W248"},
            {"rig": "DRPPLACEHOLDER A 168"},
        ]):
            rigs = form_server._production_filter_rigs([
                {"rig": "00 SINOPEC 168"},
                {"rig": "UNKNOWN"},
            ])
        self.assertEqual(rigs, ["SINOPEC 168", "SINOPEC 248"])

    def test_production_report_remarks_are_loaded_from_structured_storage(self) -> None:
        records = [{
            "record_id": "drilling:WELL-1:2026-07-18:1",
            "project_id": "7",
            "project_name": "Project A",
            "project_contract": "C-1",
            "rig": "W248",
            "wellbore": "WELL-1",
            "report_type": "drilling",
            "reportDate": "2026-07-18",
            "event": "DRILLING",
            "validation_status": "ok",
        }]
        with patch.object(
            form_server,
            "load_production_report_remarks",
            return_value={"7|SINOPEC 248|WELL-1": "用户维护备注"},
        ):
            details = form_server._production_report_details(records, [])
        self.assertEqual(details[0]["remarks"], "用户维护备注")
        self.assertEqual(details[0]["remark_key"], "7|SINOPEC 248|WELL-1")

    def test_drilling_validation_matches_detail_required_fields(self) -> None:
        payload = {
            "report_fields": {
                "event": "DRILLING",
                "reportDate": "2026-05-31",
                "reportNo": "19",
                "wellbore": "PCNC-039",
                "rig": "SINOPEC 248",
                "todayMd": "1000",
                "progress": "0",
                "currentOps": "Drilling",
                "summary24h": "Operations",
                "forecast24h": "Continue",
            },
            "operations": [
                {"from": "00:00", "to": "08:00", "hours": "8", "op_code": "BHA", "op_type": "P", "operation_details": "Trip"},
            ],
        }
        warnings = form_server._validation_warnings(payload, "drilling")
        self.assertIn("mudType missing", warnings)
        self.assertIn("mudDensity missing", warnings)
        self.assertIn("operation hours total 8.00", warnings)

        middle_day_warnings = form_server._validation_warnings(
            payload,
            "drilling",
            validate_operation_total=False,
        )
        self.assertNotIn("operation hours total 8.00", middle_day_warnings)

    def test_operation_duration_mismatch_is_flagged(self) -> None:
        payload = {
            "report_fields": {
                "event": "DRILLING",
                "reportDate": "2026-05-31",
                "reportNo": "19",
                "wellbore": "PCNC-039",
                "rig": "SINOPEC 248",
                "todayMd": "1000",
                "progress": "0",
                "currentOps": "Drilling",
                "summary24h": "Operations",
                "forecast24h": "Continue",
                "mudType": "WBM",
                "mudDensity": "10.2",
            },
            "operations": [
                {"from": "00:00", "to": "08:00", "hours": "7.5", "op_code": "DRILLING", "op_type": "P", "operation_details": "Drill ahead"},
                {"from": "08:00", "to": "24:00", "hours": "16.5", "op_code": "BHA", "op_type": "P", "operation_details": "Trip"},
            ],
        }
        warnings = form_server._validation_warnings(payload, "drilling")
        self.assertIn("operations row 1 time duration mismatch", warnings)

    def test_operation_same_clock_time_can_mean_next_day(self) -> None:
        payload = {
            "report_fields": {
                "event": "DRILLING",
                "reportDate": "2026-05-31",
                "reportNo": "19",
                "wellbore": "PCNC-039",
                "rig": "SINOPEC 248",
                "todayMd": "1000",
                "progress": "0",
                "currentOps": "Drilling",
                "summary24h": "Operations",
                "forecast24h": "Continue",
                "mudType": "WBM",
                "mudDensity": "10.2",
            },
            "operations": [
                {"from": "06:00", "to": "06:00", "hours": "24", "op_code": "DRILLING", "op_type": "P", "operation_details": "Drill ahead"},
            ],
        }
        warnings = form_server._validation_warnings(payload, "drilling")
        self.assertNotIn("operations row 1 time duration mismatch", warnings)

    def test_cost_inventory_are_not_validated_but_intervals_still_are(self) -> None:
        completion_payload = {
            "report_fields": {
                "event": "COMPLETION",
                "reportDate": "2026-05-31",
                "reportNo": "1",
                "wellbore": "PCNC-039",
                "rig": "SINOPEC 248",
                "currentOps": "Completion",
                "summary24h": "Operations",
                "forecast24h": "Continue",
            },
            "operations": [
                {"from": "00:00", "to": "24:00", "hours": "24", "op_code": "COMPLETION", "op_type": "P", "operation_details": "Run completion"},
            ],
            "bulks": [{"qty_start": "10", "qty_used": "3", "qty_end": "8"}],
            "perforation_intervals": [{"top_md": "1200", "base_md": "1100", "length": "10", "density": "-1"}],
        }
        warnings = form_server._validation_warnings(completion_payload, "completion")
        self.assertNotIn("bulks row 1 ending quantity inconsistent", warnings)
        self.assertIn("perforation_intervals row 1 base_md less than top_md", warnings)
        self.assertIn("perforation_intervals row 1 density negative", warnings)

    def test_login_page_supports_head_health_check(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            original_users_path = form_server.USERS_PATH
            form_server.USERS_PATH = Path(tmp) / "users.json"
            form_server.SESSIONS.clear()
            server = form_server.ThreadingHTTPServer(("127.0.0.1", 0), form_server.FormHandler)
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                response = _head(server.server_port, "/login/")
                self.assertEqual(response["status"], 200)
                self.assertEqual(response["body"], b"")
                self.assertGreater(int(response["content_length"]), 0)
            finally:
                server.shutdown()
                server.server_close()
                thread.join(timeout=2)
                form_server.USERS_PATH = original_users_path
                form_server.SESSIONS.clear()

    def test_import_drilling_pdf_endpoint_parses_and_stores_record(self) -> None:
        pdf = sample_pdf("PCNC-040")
        if not pdf.exists():
            self.skipTest(f"Sample PDF not found: {pdf}")

        with tempfile.TemporaryDirectory() as tmp:
            original_database_path = form_server.DATABASE_PATH
            original_users_path = form_server.USERS_PATH
            original_config_path = form_server.CONFIG_PATH
            original_audit_log_path = form_server.AUDIT_LOG_PATH
            form_server.DATABASE_PATH = Path(tmp) / "report_database.xlsx"
            form_server.USERS_PATH = Path(tmp) / "users.json"
            form_server.CONFIG_PATH = Path(tmp) / "system_config.json"
            form_server.AUDIT_LOG_PATH = Path(tmp) / "audit_logs.jsonl"
            form_server.SESSIONS.clear()
            server = form_server.ThreadingHTTPServer(("127.0.0.1", 0), form_server.FormHandler)
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                cookie = _login(server.server_port, "admin", "admin123")
                response = _post_pdf(server.server_port, "/api/import-pdf", pdf, cookie)
                self.assertEqual(response["status"], 200, response["body"])
                payload = json.loads(response["body"])
                self.assertEqual(payload["report_fields"]["wellbore"], "PCNC-040")
                self.assertEqual(payload["report_fields"]["reportNo"], "11")
                self.assertEqual(payload["metadata"]["record_id"], "drilling:PCNC-040:2026-06-11:11")

                stored = load_report_payload(form_server.DATABASE_PATH, payload["metadata"]["record_id"])
                self.assertEqual(stored["report_fields"]["rig"], "SINOPEC 248")
                self.assertEqual(len(stored["operations"]), 6)
            finally:
                server.shutdown()
                server.server_close()
                thread.join(timeout=2)
                form_server.DATABASE_PATH = original_database_path
                form_server.USERS_PATH = original_users_path
                form_server.CONFIG_PATH = original_config_path
                form_server.AUDIT_LOG_PATH = original_audit_log_path
                form_server.SESSIONS.clear()

    def test_translate_report_endpoint_returns_structured_result(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            original_users_path = form_server.USERS_PATH
            original_terms_path = form_server.TRANSLATION_TERMS_PATH
            original_engine = os.environ.get("DRP_TRANSLATION_ENGINE")
            form_server.USERS_PATH = Path(tmp) / "users.json"
            form_server.TRANSLATION_TERMS_PATH = Path(tmp) / "translation_terms.json"
            os.environ["DRP_TRANSLATION_ENGINE"] = "noop"
            form_server.SESSIONS.clear()
            server = form_server.ThreadingHTTPServer(("127.0.0.1", 0), form_server.FormHandler)
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                cookie = _login(server.server_port, "admin", "admin123")
                with patch.object(
                    form_server,
                    "_active_translation_config",
                    return_value=form_server.TranslationConfig(engine="noop"),
                ):
                    response = _post_json(
                        server.server_port,
                        "/api/translate-report",
                        {
                            "target_language": "zh",
                            "payload": {
                                "metadata": {"source_file": "mixed.pdf"},
                                "report_fields": {
                                    "wellbore": "EB-D-12",
                                    "currentOps": "ROP 18 m/hr while drilling 12.25 in section.",
                                },
                                "operations": [],
                            },
                        },
                        cookie,
                    )
                self.assertEqual(response["status"], 200, response["body"])
                payload = json.loads(response["body"])
                self.assertEqual(payload["metadata"]["engine"], "noop")
                self.assertEqual(payload["metadata"]["target_language"], "zh-CN")
                self.assertIn("translation_content", payload)
                self.assertEqual(
                    payload["translated_payload"]["report_fields"]["currentOps"],
                    "ROP 18 m/hr while drilling 12.25 in section.",
                )
                self.assertIsInstance(payload["translation_content"], list)
            finally:
                server.shutdown()
                server.server_close()
                thread.join(timeout=2)
                form_server.USERS_PATH = original_users_path
                form_server.TRANSLATION_TERMS_PATH = original_terms_path
                if original_engine is None:
                    os.environ.pop("DRP_TRANSLATION_ENGINE", None)
                else:
                    os.environ["DRP_TRANSLATION_ENGINE"] = original_engine
                form_server.SESSIONS.clear()

    def test_admin_translation_terms_can_be_read_and_saved(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            original_users_path = form_server.USERS_PATH
            original_terms_path = form_server.TRANSLATION_TERMS_PATH
            form_server.USERS_PATH = Path(tmp) / "users.json"
            form_server.TRANSLATION_TERMS_PATH = Path(tmp) / "translation_terms.json"
            form_server.SESSIONS.clear()
            server = form_server.ThreadingHTTPServer(("127.0.0.1", 0), form_server.FormHandler)
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                cookie = _login(server.server_port, "admin", "admin123")
                get_response = _get_json(server.server_port, "/api/admin/translation-terms", cookie)
                self.assertEqual(get_response["status"], 200, get_response["body"])
                config = json.loads(get_response["body"])
                self.assertTrue(any(term["zh"] == "机械钻速" for term in config["terms"]))

                config["terms"].append({
                    "id": "term-test-flowline",
                    "category": "equipment",
                    "zh": "流管线",
                    "en": "flowline",
                    "es": "línea de flujo",
                    "aliases": {"zh": [], "en": ["flow line"], "es": ["linea de flujo"]},
                    "protected": True,
                    "enabled": True,
                })
                with patch.object(
                    form_server,
                    "_active_ai_job_counts",
                    return_value={"translation": 0, "extraction": 0, "total": 0},
                ):
                    post_response = _post_json(server.server_port, "/api/admin/translation-terms", config, cookie)
                self.assertEqual(post_response["status"], 200, post_response["body"])
                saved = json.loads(post_response["body"])
                self.assertTrue(any(term["en"] == "flowline" for term in saved["terms"]))
                self.assertTrue(form_server.TRANSLATION_TERMS_PATH.exists())
            finally:
                server.shutdown()
                server.server_close()
                thread.join(timeout=2)
                form_server.USERS_PATH = original_users_path
                form_server.TRANSLATION_TERMS_PATH = original_terms_path
                form_server.SESSIONS.clear()


def _login(port: int, username: str, password: str) -> str:
    connection = http.client.HTTPConnection("127.0.0.1", port, timeout=20)
    body = json.dumps({"username": username, "password": password}).encode("utf-8")
    try:
        connection.request(
            "POST",
            "/api/admin/login",
            body=body,
            headers={"Content-Type": "application/json", "Content-Length": str(len(body))},
        )
        response = connection.getresponse()
        payload = response.read().decode("utf-8")
        if response.status != 200:
            raise AssertionError(payload)
        cookie = response.getheader("Set-Cookie", "").split(";", 1)[0]
        login_payload = json.loads(payload)
        if login_payload.get("user", {}).get("must_change_password"):
            changed = _post_json(
                port,
                "/api/admin/change-password",
                {"old_password": password, "new_password": f"{password}-changed"},
                cookie,
            )
            if changed["status"] != 200:
                raise AssertionError(changed["body"])
        return cookie
    finally:
        connection.close()


def _head(port: int, path: str) -> dict[str, Any]:
    connection = http.client.HTTPConnection("127.0.0.1", port, timeout=20)
    try:
        connection.request("HEAD", path)
        response = connection.getresponse()
        return {
            "status": response.status,
            "content_length": response.getheader("Content-Length", "0"),
            "body": response.read(),
        }
    finally:
        connection.close()


def _get_json(port: int, path: str, cookie: str = "") -> dict[str, Any]:
    connection = http.client.HTTPConnection("127.0.0.1", port, timeout=20)
    try:
        headers = {}
        if cookie:
            headers["Cookie"] = cookie
        connection.request("GET", path, headers=headers)
        response = connection.getresponse()
        return {"status": response.status, "body": response.read().decode("utf-8")}
    finally:
        connection.close()


def _post_json(port: int, path: str, payload: dict[str, Any], cookie: str = "") -> dict[str, Any]:
    connection = http.client.HTTPConnection("127.0.0.1", port, timeout=20)
    body = json.dumps(payload).encode("utf-8")
    try:
        headers = {"Content-Type": "application/json", "Content-Length": str(len(body))}
        if cookie:
            headers["Cookie"] = cookie
        connection.request("POST", path, body=body, headers=headers)
        response = connection.getresponse()
        return {"status": response.status, "body": response.read().decode("utf-8")}
    finally:
        connection.close()


def _post_pdf(port: int, path: str, pdf: Path, cookie: str = "") -> dict[str, Any]:
    boundary = "----drilling-report-test-boundary"
    data = pdf.read_bytes()
    head = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="report"; filename="{pdf.name}"\r\n'
        "Content-Type: application/pdf\r\n\r\n"
    ).encode("utf-8")
    tail = f"\r\n--{boundary}--\r\n".encode("utf-8")
    body = head + data + tail
    connection = http.client.HTTPConnection("127.0.0.1", port, timeout=20)
    try:
        headers = {"Content-Type": f"multipart/form-data; boundary={boundary}", "Content-Length": str(len(body))}
        if cookie:
            headers["Cookie"] = cookie
        connection.request(
            "POST",
            path,
            body=body,
            headers=headers,
        )
        response = connection.getresponse()
        return {"status": response.status, "body": response.read().decode("utf-8")}
    finally:
        connection.close()


if __name__ == "__main__":
    unittest.main()
