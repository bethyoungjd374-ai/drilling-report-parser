from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from drilling_report_parser import parse_excel_report, write_structured_workbook


class ParserTest(unittest.TestCase):
    def test_parse_sample_report_and_write_structured_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "sample.xlsx"
            output = Path(tmp) / "structured.xlsx"
            _build_sample_workbook(source)

            result = parse_excel_report(source)
            write_structured_workbook(result, output)

            fields = {row["field"]: row["value"] for row in result.fields}
            self.assertEqual(fields["Event"], "DEV DRILLING")
            self.assertEqual(fields["Report No"], "16")
            self.assertIn("BAJANDO CASING", fields["Current Ops"])
            self.assertEqual(len(result.tables["operations"]), 2)
            self.assertEqual(result.tables["operations"][0]["hours"], "1.50")
            self.assertEqual(len(result.tables["survey_data"]), 2)
            self.assertEqual(len(result.tables["bha_components"]), 2)

            wb = load_workbook(output, data_only=True)
            self.assertIn("report_fields", wb.sheetnames)
            self.assertIn("operations", wb.sheetnames)
            self.assertIn("bha_components", wb.sheetnames)


def _build_sample_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"
    ws["A1"] = "DAILY OPERATIONS REPORT"
    ws["A2"] = "Event:"
    ws["B2"] = "DEV DRILLING"
    ws["D2"] = "Date:"
    ws["E2"] = "06/11/2026"
    ws["A3"] = "Report No:"
    ws["B3"] = 16
    ws["A4"] = "Wellbore:"
    ws["B4"] = "SCHAO-611"
    ws["A5"] = "Current Ops:"
    ws["B5"] = 'BAJANDO CASING 9-5/8" EN HUECO ENTUBADO A 6800 FT'
    ws["A6"] = "24-Hr Summary:"
    ws["B6"] = "SACA Y QUIEBRA BHA #5 DIRECCIONAL EN SUPERFICIE."
    ws["A7"] = "Last Casing:"
    ws["B7"] = "13.375in @ 7,111ft"
    ws["D7"] = "Next Casing:"
    ws["E7"] = "9.625in @ 10,754ft"

    ws.append([])
    ws.append(["From", "To", "Hrs", "Op Code", "Op Sub", "Op Type", "Operation Details"])
    ws.append(["6:00", "7:30", "1.50", "BHA", "POOH", "P", "SACA BHA #5 DIRECCIONAL"])
    ws.append(["7:30", "9:00", "1.50", "CIRCULATING", "Condition Hole", "P", "REALIZA ESTACION DE CIRCULACION"])
    ws.append([])
    ws.append(["SURVEY DATA (LAST 6)"])
    ws.append(["MD", "Incl", "Azi", "TVD", "VSE", "N/-S", "DLS", "Build"])
    ws.append(["10,725.1", "23.58", "25.88", "9,418.92", "4,702.7", "4,212.3", "0.12", "-0.08"])
    ws.append(["10,687.5", "23.61", "25.80", "9,384.40", "4,687.6", "4,198.7", "0.19", "0.06"])
    ws.append([])
    ws.append(["LAST OR CURRENT BHA"])
    ws.append(["BHA No:", "05", "MD In:", "9,215.00 ft"])
    ws.append(["Component", "OD", "ID", "Jts", "Length"])
    ws.append(["Polycrystalline Diamond Bit", "12.250", "3.250", "1", "1.24"])
    ws.append(["Power Drive 900 AA", "9.625", "7.852", "1", "13.91"])

    wb.save(path)


if __name__ == "__main__":
    unittest.main()
