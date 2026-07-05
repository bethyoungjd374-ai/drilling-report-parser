from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from drilling_report_parser.excel_database import load_report_payload, save_report_payload


class ExcelDatabaseTest(unittest.TestCase):
    def test_save_drilling_report_splits_single_and_multi_row_tables(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            database = Path(tmp) / "report_database.xlsx"
            result = save_report_payload(
                database,
                {
                    "metadata": {"source_file": "daily.pdf", "parser": "pdf_report_parser_v1"},
                    "report_fields": {
                        "event": "DRILLING",
                        "reportDate": "2026-06-11",
                        "reportNo": "11",
                        "wellbore": "PCNC-040",
                        "rig": "00 SINOPEC 248",
                        "currentOps": "DRILL",
                    },
                    "operations": [
                        {"from": "00:00", "to": "12:00", "hours": "12.00", "op_code": "DRILLING", "op_type": "P"},
                        {"from": "12:00", "to": "24:00", "hours": "12.00", "op_code": "DRILLING", "op_type": "NPT"},
                    ],
                    "survey_data": [{"md": "1000", "incl": "1.20"}],
                    "bha_components": [{"component": "Bit", "od": "8.500"}],
                },
                "drilling",
            )

            workbook = load_workbook(database)
            self.assertIn("records", workbook.sheetnames)
            self.assertIn("drilling_fields", workbook.sheetnames)
            self.assertIn("drilling_operations", workbook.sheetnames)
            self.assertIn("drilling_survey", workbook.sheetnames)
            self.assertEqual(workbook["records"].max_row, 2)
            self.assertEqual(workbook["drilling_fields"].max_row, 2)
            self.assertEqual(workbook["drilling_operations"].max_row, 3)
            self.assertEqual(workbook["drilling_survey"].max_row, 2)

            loaded = load_report_payload(database, result["record_id"])
            self.assertEqual(loaded["report_fields"]["wellbore"], "PCNC-040")
            self.assertEqual(len(loaded["operations"]), 2)
            self.assertEqual(loaded["operations"][1]["op_type"], "NPT")
            self.assertEqual(loaded["survey_data"][0]["md"], "1000")

    def test_save_same_completion_record_replaces_old_detail_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            database = Path(tmp) / "report_database.xlsx"
            payload = {
                "metadata": {"record_id": "completion:SCHAS-513:2026-06-11:7"},
                "report_fields": {"reportDate": "2026-06-11", "reportNo": "7", "wellbore": "SCHAS-513", "rig": "SINOPEC 219"},
                "operations": [{"from": "00:00", "to": "24:00", "hours": "24.00"}],
                "bulks": [{"bulk": "DIESEL", "qty_used": "10"}],
            }
            save_report_payload(database, payload, "completion")
            payload["operations"] = [{"from": "06:00", "to": "18:00", "hours": "12.00"}]
            payload["bulks"] = []
            save_report_payload(database, payload, "completion")

            workbook = load_workbook(database)
            self.assertEqual(workbook["records"].max_row, 2)
            self.assertEqual(workbook["completion_operations"].max_row, 2)
            self.assertEqual(workbook["completion_bulks"].max_row, 1)
            loaded = load_report_payload(database, "completion:SCHAS-513:2026-06-11:7")
            self.assertEqual(loaded["operations"][0]["from"], "06:00")
            self.assertEqual(loaded["bulks"], [])

    def test_move_report_type_is_saved_as_drilling(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            database = Path(tmp) / "report_database.xlsx"
            result = save_report_payload(
                database,
                {
                    "metadata": {"source_file": "move.pdf", "parser": "move_pdf_parser_v1"},
                    "report_fields": {
                        "event": "MAJOR RIG MOVE",
                        "reportDate": "2026-06-10",
                        "reportNo": "5",
                        "wellbore": "TCHA-006I",
                        "rig": "00 SINOPEC 168",
                        "currentOps": "MOVE",
                    },
                    "operations": [
                        {"from": "00:00", "to": "24:00", "hours": "24.00", "op_code": "MOVE", "op_type": "P"},
                    ],
                },
                "move",
            )

            workbook = load_workbook(database)
            self.assertIn("drilling_fields", workbook.sheetnames)
            self.assertNotIn("move_fields", workbook.sheetnames)
            records = workbook["records"]
            headers = [cell.value for cell in records[1]]
            saved = {headers[index]: value for index, value in enumerate(next(records.iter_rows(min_row=2, max_row=2, values_only=True)))}
            self.assertEqual(saved["report_type"], "drilling")
            loaded = load_report_payload(database, result["record_id"])
            self.assertEqual(loaded["metadata"]["report_type"], "drilling")
            self.assertEqual(loaded["operations"][0]["op_code"], "MOVE")

    def test_save_to_existing_records_sheet_aligns_added_status_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            database = Path(tmp) / "report_database.xlsx"
            save_report_payload(
                database,
                {
                    "metadata": {"source_file": "old.pdf"},
                    "report_fields": {"reportDate": "2026-06-10", "reportNo": "1", "wellbore": "OLD-001", "rig": "SINOPEC 1"},
                },
                "drilling",
            )
            workbook = load_workbook(database)
            records = workbook["records"]
            for header in ("status", "validation_status", "validation_warnings"):
                column = [cell.value for cell in records[1]].index(header) + 1
                records.delete_cols(column)
            workbook.save(database)

            save_report_payload(
                database,
                {
                    "metadata": {"source_file": "new.pdf", "status": "parsed", "validation_status": "warning", "validation_warnings": "rig missing"},
                    "report_fields": {"reportDate": "2026-06-11", "reportNo": "2", "wellbore": "NEW-002", "rig": ""},
                },
                "drilling",
            )

            workbook = load_workbook(database)
            records = workbook["records"]
            headers = [cell.value for cell in records[1]]
            latest = {headers[index]: value for index, value in enumerate(next(records.iter_rows(min_row=3, max_row=3, values_only=True)))}
            self.assertEqual(latest["status"], "parsed")
            self.assertEqual(latest["validation_status"], "warning")
            self.assertEqual(latest["validation_warnings"], "rig missing")
            self.assertTrue(str(latest["created_at"]).startswith("2026-") or "T" in str(latest["created_at"]))


if __name__ == "__main__":
    unittest.main()
