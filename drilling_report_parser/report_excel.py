from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "123554"
NAVY_DARK = "08243A"
BLUE = "1F6FAE"
LIGHT_BLUE = "E7F1F9"
PALE = "F6F9FC"
WHITE = "FFFFFF"
BORDER = "AFC1D1"
TEXT = "152235"
MUTED = "5B6F84"
WARNING_FILL = "FFF1B8"


def build_daily_report_workbook(payload: dict[str, Any]) -> Workbook:
    fields = payload.get("report_fields", {})
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.print_options.horizontalCentered = True

    widths = [13, 15, 13, 15, 13, 15, 13, 15, 13, 15, 13, 18]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    _title(ws, 1, "DAILY OPERATIONS REPORT")
    _subtitle(ws, 2, f"{_v(fields, 'wellbore')} / {_v(fields, 'rig')} / Report No. {_v(fields, 'reportNo')} / {_v(fields, 'reportDate')}")

    row = 4
    row = _key_value_grid(ws, row, [
        ("Event", _v(fields, "event")),
        ("Date", _v(fields, "reportDate")),
        ("Report No", _v(fields, "reportNo")),
        ("Wellbore", _v(fields, "wellbore")),
        ("Rig", _v(fields, "rig")),
        ("Primary Reason", _v(fields, "primaryReason")),
        ("AFE Number", _v(fields, "afeNumber")),
        ("Ref Datum", _v(fields, "refDatum")),
        ("Today's MD (ft)", _v(fields, "todayMd")),
        ("Prev MD (ft)", _v(fields, "prevMd")),
        ("Progress (ft)", _v(fields, "progress")),
        ("Rot Hrs Today", _v(fields, "rotHrsToday")),
    ])

    row = _section(ws, row + 1, "CURRENT OPERATIONS / 24-HR SUMMARY")
    row = _text_block(ws, row, "Current Ops", _v(fields, "currentOps"), "B", "L", height=36)
    row = _two_text_blocks(ws, row, "24-Hr Summary", _v(fields, "summary24h"), "24-Hr Forecast", _v(fields, "forecast24h"))

    row = _section(ws, row + 1, "CASING / WELL CONTROL / HYDRAULICS")
    row = _key_value_grid(ws, row, [
        ("Last Casing Size", _v(fields, "lastCasingSize") or _split_at(_v(fields, "lastCasing"))[0]),
        ("Last Casing Depth", _v(fields, "lastCasingDepth") or _split_at(_v(fields, "lastCasing"))[1]),
        ("Next Casing Size", _v(fields, "nextCasingSize") or _split_at(_v(fields, "nextCasing"))[0]),
        ("Next Casing Depth", _v(fields, "nextCasingDepth") or _split_at(_v(fields, "nextCasing"))[1]),
        ("Form Test/EMW", _v(fields, "formTestEmw")),
        ("Last BOP Press Test", _v(fields, "lastBopPressTest")),
        ("Pump Rate (gpm)", _v(fields, "pumpRate")),
        ("Pump Press (psi)", _v(fields, "pumpPress")),
        ("String Wt Up/Dn", _v(fields, "stringWeightUpDown")),
        ("Torque On Btm", _v(fields, "torqueOnBottom")),
    ])

    row = _section(ws, row + 1, "SURVEY DATA (LAST 6)")
    row = _table(ws, row, ["MD", "Incl", "Azi", "TVD", "VSE", "N/-S", "DLS", "Build"], payload.get("survey_data", []), ["md", "incl", "azi", "tvd", "vse", "ns", "dls", "build"])

    row = _section(ws, row + 1, "MUD DATA")
    row = _key_value_grid(ws, row, [
        ("Engineer", _v(fields, "mudEngineer")),
        ("Sample From", _v(fields, "sampleFrom")),
        ("Mud Type", _v(fields, "mudType")),
        ("Mud Time", _v(fields, "mudTime") or _split_slash(_v(fields, "mudTimeMd"), 2)[0]),
        ("Mud MD", _v(fields, "mudMd") or _split_slash(_v(fields, "mudTimeMd"), 2)[1]),
        ("Density (ppg)", _v(fields, "mudDensity")),
        ("Mud Temp", _v(fields, "mudTemperature")),
        ("Rheology Temp", _v(fields, "rheologyTemp")),
        ("Viscosity", _v(fields, "viscosity")),
        ("PV", _v(fields, "pv") or _split_slash(_v(fields, "pvYp"), 2)[0]),
        ("YP", _v(fields, "yp") or _split_slash(_v(fields, "pvYp"), 2)[1]),
        ("Gel 10s", _v(fields, "gel10s") or _split_slash(_v(fields, "gels"), 3)[0]),
        ("Gel 10m", _v(fields, "gel10m") or _split_slash(_v(fields, "gels"), 3)[1]),
        ("Gel 30m", _v(fields, "gel30m") or _split_slash(_v(fields, "gels"), 3)[2]),
        ("API WL", _v(fields, "apiWl")),
        ("Oil (%)", _v(fields, "oilPercent") or _split_slash(_v(fields, "oilWater"), 2)[0]),
        ("Water (%)", _v(fields, "waterPercent") or _split_slash(_v(fields, "oilWater"), 2)[1]),
        ("Sand (%)", _v(fields, "sand")),
        ("ECD", _v(fields, "ecd")),
    ])
    row = _text_block(ws, row, "Mud Comments", _v(fields, "mudComments"), "B", "L", height=34)

    row = _section(ws, row + 1, "BIT RECORD / LAST OR CURRENT BHA")
    row = _key_value_grid(ws, row, [
        ("Bit No", _v(fields, "bitNo")),
        ("Bit Size", _v(fields, "bitSize")),
        ("Manufacturer", _v(fields, "bitManufacturer")),
        ("Serial No", _v(fields, "bitSerial")),
        ("BHA No", _v(fields, "bhaNo")),
        ("MD In", _v(fields, "bhaMdIn")),
        ("MD Out", _v(fields, "bhaMdOut")),
        ("Total Length (ft)", _v(fields, "bhaTotalLength")),
    ])
    row = _table(ws, row, ["Component", "OD", "ID", "Jts", "Length"], payload.get("bha_components", []), ["component", "od", "id", "joints", "length"], widths=[4, 2, 2, 2, 2])

    row = _section(ws, row + 1, "OPERATIONS")
    row = _table(
        ws,
        row,
        ["From", "To", "Hrs", "Op Code", "Op Sub", "Type", "Operation Details"],
        payload.get("operations", []),
        ["from", "to", "hours", "op_code", "op_sub", "op_type", "operation_details"],
        widths=[1, 1, 1, 1, 2, 1, 5],
        row_height=44,
        warning_keys={"op_type"},
    )

    row = _section(ws, row + 1, "DAILY COSTS / BULKS")
    row = _table(ws, row, ["Cost Description", "Vendor", "Amount"], payload.get("daily_costs", []), ["cost_description", "vendor", "amount"])
    row = _table(ws, row + 1, ["Bulk", "Qty Start", "Qty Used", "Qty End"], payload.get("bulks", []), ["bulk", "qty_start", "qty_used", "qty_end"])

    row = _section(ws, row + 1, "INCIDENTS AND REMARKS")
    row = _key_value_grid(ws, row, [
        ("Safety Incident?", _v(fields, "safetyIncident")),
        ("Environ Incident?", _v(fields, "environmentIncident")),
        ("Days since Last RI", _v(fields, "daysSinceRi")),
        ("Days since Last LTA", _v(fields, "daysSinceLta")),
    ])
    row = _text_block(ws, row, "Incident Comments", _v(fields, "incidentComments"), "B", "L", height=34)
    row = _text_block(ws, row, "Other Remarks", _v(fields, "otherRemarks"), "B", "L", height=50)

    ws.print_area = f"A1:L{row}"
    _apply_global(ws, row)
    return wb


def build_completion_report_workbook(payload: dict[str, Any]) -> Workbook:
    fields = payload.get("report_fields", {})
    wb = Workbook()
    ws = wb.active
    ws.title = "Completion Report"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.print_options.horizontalCentered = True

    widths = [12, 15, 12, 15, 12, 15, 12, 15, 12, 15, 12, 20]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    _title(ws, 1, "DAILY COMPLETION OPERATIONS REPORT")
    _subtitle(ws, 2, f"{_v(fields, 'wellbore')} / {_v(fields, 'rig')} / Report No. {_v(fields, 'reportNo')} / {_v(fields, 'reportDate')}")

    row = 4
    row = _key_value_grid(ws, row, [
        ("Event", _v(fields, "event")),
        ("Date", _v(fields, "reportDate")),
        ("Report No", _v(fields, "reportNo")),
        ("Wellbore", _v(fields, "wellbore")),
        ("Rig", _v(fields, "rig")),
        ("Primary Reason", _v(fields, "primaryReason")),
        ("Description", _v(fields, "description")),
        ("Operation Start", _v(fields, "operationStartDate")),
        ("AFP Number", _v(fields, "afeNumber")),
        ("AFP Cost", _v(fields, "afeCost")),
        ("Ref Datum", _v(fields, "refDatum")),
        ("Daily Cost", _v(fields, "dailyCost")),
        ("Cumulative Cost", _v(fields, "cumulativeCost")),
        ("Total Personnel", _v(fields, "totalPersonnel")),
    ])

    row = _section(ws, row + 1, "CURRENT OPERATIONS / 24-HR SUMMARY")
    row = _text_block(ws, row, "Current Operation", _v(fields, "currentOps"), "B", "L", height=36)
    row = _two_text_blocks(ws, row, "24-Hr Summary", _v(fields, "summary24h"), "24-Hr Forecast", _v(fields, "forecast24h"))

    row = _section(ws, row + 1, "PERSONNEL")
    row = _key_value_grid(ws, row, [
        ("Supervisor 1", _v(fields, "supervisor1")),
        ("Supervisor 2", _v(fields, "supervisor2")),
        ("Engineer", _v(fields, "engineer")),
        ("PAM Engineer", _v(fields, "pamEngineer")),
        ("Geologist", _v(fields, "geologist")),
    ])

    row = _section(ws, row + 1, "OPERATIONS")
    row = _table(
        ws,
        row,
        ["From", "To", "Hrs", "Op Code", "Op Sub", "Type", "Operation Details"],
        payload.get("operations", []),
        ["from", "to", "hours", "op_code", "op_sub", "op_type", "operation_details"],
        widths=[1, 1, 1, 2, 2, 1, 4],
        row_height=46,
        warning_keys={"op_type"},
        warning_valid_values={"op_type": {"P", "SC", "NPT"}},
    )

    row = _section(ws, row + 1, "BULKS / DAILY COSTS")
    row = _table(ws, row, ["Bulk", "Qty Start", "Qty Used", "Qty End"], payload.get("bulks", []), ["bulk", "qty_start", "qty_used", "qty_end"])
    row = _table(ws, row + 1, ["Cost Description", "Vendor", "Amount"], payload.get("daily_costs", []), ["cost_description", "vendor", "amount"])

    row = _section(ws, row + 1, "PERFORATED INTERVALS")
    row = _table(
        ws,
        row,
        ["Formation", "Top MD", "Base MD", "Length", "Density", "Charges", "Phase", "Penetration", "Diameter", "Date", "Status", "Comments"],
        payload.get("perforation_intervals", []),
        ["formation", "top_md", "base_md", "length", "density", "charges", "phase", "penetration", "diameter", "date", "status", "comments"],
        widths=[1, 1, 1, 1, 1, 2, 1, 1, 1, 1, 1, 1],
        row_height=32,
    )

    row = _section(ws, row + 1, "SAFETY COMMENTS / REMARKS")
    row = _text_block(ws, row, "Safety Comments", _v(fields, "safetyComments"), "B", "L", height=34)
    row = _text_block(ws, row, "Other Remarks", _v(fields, "otherRemarks"), "B", "L", height=42)

    ws.print_area = f"A1:L{row}"
    _apply_global(ws, row)
    return wb


def build_workover_report_workbook(payload: dict[str, Any]) -> Workbook:
    fields = payload.get("report_fields", {})
    wb = Workbook()
    ws = wb.active
    ws.title = "Workover Report"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.print_options.horizontalCentered = True

    widths = [12, 15, 12, 15, 12, 15, 12, 15, 12, 15, 12, 20]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    _title(ws, 1, "DAILY WORKOVER OPERATIONS REPORT")
    _subtitle(ws, 2, f"{_v(fields, 'wellbore')} / {_v(fields, 'rig')} / WO #{_v(fields, 'workoverNo')} / Report No. {_v(fields, 'reportNo')} / {_v(fields, 'reportDate')}")

    row = 4
    row = _key_value_grid(ws, row, [
        ("Event", _v(fields, "event")),
        ("Date", _v(fields, "reportDate")),
        ("Report No", _v(fields, "reportNo")),
        ("WO No", _v(fields, "workoverNo")),
        ("Wellbore", _v(fields, "wellbore")),
        ("Rig", _v(fields, "rig")),
        ("Primary Reason", _v(fields, "primaryReason")),
        ("Description", _v(fields, "description")),
        ("Operation Start", _v(fields, "operationStartDate")),
        ("AFP Number", _v(fields, "afeNumber")),
        ("AFP Cost", _v(fields, "afeCost")),
        ("Ref Datum", _v(fields, "refDatum")),
        ("Daily Cost", _v(fields, "dailyCost")),
        ("Cumulative Cost", _v(fields, "cumulativeCost")),
        ("Total Personnel", _v(fields, "totalPersonnel")),
    ])

    row = _section(ws, row + 1, "CURRENT OPERATIONS / 24-HR SUMMARY")
    row = _text_block(ws, row, "Current Operation", _v(fields, "currentOps"), "B", "L", height=36)
    row = _two_text_blocks(ws, row, "24-Hr Summary", _v(fields, "summary24h"), "24-Hr Forecast", _v(fields, "forecast24h"))

    row = _section(ws, row + 1, "PERSONNEL")
    row = _key_value_grid(ws, row, [
        ("Supervisor 1", _v(fields, "supervisor1")),
        ("Supervisor 2", _v(fields, "supervisor2")),
        ("Engineer", _v(fields, "engineer")),
        ("PAM Engineer", _v(fields, "pamEngineer")),
        ("Geologist", _v(fields, "geologist")),
    ])

    row = _section(ws, row + 1, "OPERATIONS")
    row = _table(
        ws,
        row,
        ["From", "To", "Hrs", "Op Code", "Op Sub", "Type", "Operation Details"],
        payload.get("operations", []),
        ["from", "to", "hours", "op_code", "op_sub", "op_type", "operation_details"],
        widths=[1, 1, 1, 2, 2, 1, 4],
        row_height=46,
        warning_keys={"op_type"},
        warning_valid_values={"op_type": {"P", "SC", "NPT"}},
    )

    row = _section(ws, row + 1, "BULKS / DAILY COSTS")
    row = _table(ws, row, ["Bulk", "Qty Start", "Qty Used", "Qty End"], payload.get("bulks", []), ["bulk", "qty_start", "qty_used", "qty_end"])
    row = _table(ws, row + 1, ["Cost Description", "Vendor", "Amount"], payload.get("daily_costs", []), ["cost_description", "vendor", "amount"])

    row = _section(ws, row + 1, "PERFORATED INTERVALS")
    row = _table(
        ws,
        row,
        ["Formation", "Top MD", "Base MD", "Length", "Density", "Charges", "Phase", "Penetration", "Diameter", "Date", "Status", "Comments"],
        payload.get("perforation_intervals", []),
        ["formation", "top_md", "base_md", "length", "density", "charges", "phase", "penetration", "diameter", "date", "status", "comments"],
        widths=[1, 1, 1, 1, 1, 2, 1, 1, 1, 1, 1, 1],
        row_height=32,
    )

    row = _section(ws, row + 1, "SAFETY COMMENTS / REMARKS")
    row = _text_block(ws, row, "Safety Comments", _v(fields, "safetyComments"), "B", "L", height=42)
    row = _text_block(ws, row, "Other Remarks", _v(fields, "otherRemarks"), "B", "L", height=36)

    ws.print_area = f"A1:L{row}"
    _apply_global(ws, row)
    return wb


def workbook_bytes(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _v(fields: dict[str, Any], key: str) -> str:
    value = fields.get(key, "")
    return "" if value is None else str(value)


def _split_at(value: str) -> tuple[str, str]:
    if "@" not in value:
        return value.strip(), ""
    left, right = value.split("@", 1)
    return left.strip(), right.strip()


def _split_slash(value: str, count: int) -> list[str]:
    parts = [part.strip() for part in value.split("/")]
    while len(parts) < count:
        parts.append("")
    return parts[:count]


def _title(ws, row: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row, 1, text)
    cell.fill = PatternFill("solid", fgColor=NAVY_DARK)
    cell.font = Font(color=WHITE, bold=True, size=18)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30


def _subtitle(ws, row: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row, 1, text)
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cell.font = Font(color=NAVY, bold=True, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22


def _section(ws, row: int, text: str) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row, 1, text)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True, size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    return row + 1


def _key_value_grid(ws, row: int, pairs: list[tuple[str, str]]) -> int:
    col_pairs = [(1, 2), (4, 5), (7, 8), (10, 11)]
    current = row
    for index, (label, value) in enumerate(pairs):
        if index and index % 4 == 0:
            current += 1
        label_col, value_col = col_pairs[index % 4]
        _label(ws, current, label_col, label)
        ws.merge_cells(start_row=current, start_column=value_col, end_row=current, end_column=value_col + 1)
        _value(ws, current, value_col, value)
    return current + 1


def _text_block(ws, row: int, label: str, text: str, start_col: str, end_col: str, height: int = 42) -> int:
    _label(ws, row, 1, label)
    ws.merge_cells(f"{start_col}{row}:{end_col}{row}")
    _value(ws, row, 2, text)
    ws.row_dimensions[row].height = height
    return row + 1


def _two_text_blocks(ws, row: int, left_label: str, left_text: str, right_label: str, right_text: str) -> int:
    _label(ws, row, 1, left_label)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    _value(ws, row, 2, left_text)
    _label(ws, row, 7, right_label)
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=12)
    _value(ws, row, 8, right_text)
    ws.row_dimensions[row].height = 58
    return row + 1


def _table(
    ws,
    row: int,
    headers: list[str],
    records: list[dict[str, Any]],
    keys: list[str],
    widths: list[int] | None = None,
    row_height: int = 24,
    warning_keys: set[str] | None = None,
    warning_valid_values: dict[str, set[str]] | None = None,
) -> int:
    widths = widths or [max(1, 12 // len(headers)) for _ in headers]
    col = 1
    spans: list[tuple[int, int]] = []
    for idx, header in enumerate(headers):
        span = widths[idx] if idx < len(widths) else 1
        start, end = col, min(12, col + span - 1)
        spans.append((start, end))
        if start != end:
            ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
        cell = ws.cell(row, start, header)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col = end + 1
        if col > 12:
            break
    _row_border(ws, row, 1, 12)
    row += 1

    if not records:
        records = [{key: "" for key in keys}]

    for record in records:
        for (start, end), key in zip(spans, keys):
            if start != end:
                ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
            cell = ws.cell(row, start, record.get(key, ""))
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.font = Font(color=TEXT, size=9)
            valid_values = (warning_valid_values or {}).get(key, {"P", "NPT"})
            is_warning = key in (warning_keys or set()) and str(record.get(key, "")).strip() not in valid_values
            cell.fill = PatternFill("solid", fgColor=WARNING_FILL if is_warning else WHITE)
        _row_border(ws, row, 1, 12)
        ws.row_dimensions[row].height = row_height
        row += 1
    return row


def _label(ws, row: int, col: int, text: str) -> None:
    cell = ws.cell(row, col, text)
    cell.fill = PatternFill("solid", fgColor=PALE)
    cell.font = Font(color=NAVY, bold=True, size=9)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = _border()


def _value(ws, row: int, col: int, text: str) -> None:
    cell = ws.cell(row, col, text)
    cell.fill = PatternFill("solid", fgColor=WHITE)
    cell.font = Font(color=TEXT, size=9)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = _border()


def _apply_global(ws, max_row: int) -> None:
    for idx in range(1, max_row + 1):
        if ws.row_dimensions[idx].height is None:
            ws.row_dimensions[idx].height = 21


def _row_border(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        ws.cell(row, col).border = _border()


def _border() -> Border:
    side = Side(style="thin", color=BORDER)
    return Border(left=side, right=side, top=side, bottom=side)
