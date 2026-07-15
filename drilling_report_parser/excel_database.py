from __future__ import annotations

import os
import threading
from datetime import datetime, timezone
from functools import wraps
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .database_common import (
    confirmation_group_status as _confirmation_group_status,
    natural_record_id as _natural_record_id,
    normalize_report_type as _normalize_report_type,
    npt_statuses as _npt_statuses,
    safe_float as _safe_float,
    slug as _slug,
)

_WORKBOOK_LOCK = threading.RLock()


def _with_workbook_lock(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        with _WORKBOOK_LOCK:
            return func(*args, **kwargs)

    return wrapper

BASE_COLUMNS = [
    "record_id",
    "report_type",
    "source_file",
    "parser",
    "reportDate",
    "reportNo",
    "wellbore",
    "rig",
    "status",
    "source_language",
    "translation_status",
    "translation_progress",
    "translation_error",
    "translation_version",
    "validation_status",
    "validation_warnings",
    "locked",
    "confirmation_status",
    "confirmed_at",
    "confirmed_by",
    "confirmation_note",
    "created_at",
    "updated_at",
]

COMMON_FIELD_COLUMNS = [
    "record_id",
    "event",
    "reportDate",
    "reportNo",
    "wellbore",
    "rig",
    "primaryReason",
    "afeNumber",
    "refDatum",
    "currentOps",
    "summary24h",
    "forecast24h",
    "otherRemarks",
]

DRILLING_FIELD_COLUMNS = COMMON_FIELD_COLUMNS + [
    "todayMd",
    "prevMd",
    "progress",
    "rotHrsToday",
    "lastCasing",
    "lastCasingSize",
    "lastCasingDepth",
    "nextCasing",
    "nextCasingSize",
    "nextCasingDepth",
    "formTestEmw",
    "lastBopPressTest",
    "pumpRate",
    "pumpPress",
    "stringWeightUpDown",
    "torqueOnBottom",
    "mudEngineer",
    "sampleFrom",
    "mudType",
    "mudTime",
    "mudMd",
    "mudDensity",
    "mudTemperature",
    "rheologyTemp",
    "viscosity",
    "pv",
    "yp",
    "gel10s",
    "gel10m",
    "gel30m",
    "apiWl",
    "oilPercent",
    "waterPercent",
    "sand",
    "ecd",
    "mudComments",
    "bitNo",
    "bitSize",
    "bitManufacturer",
    "bitSerial",
    "bitWearIodl",
    "bitWearBgor",
    "bhaNo",
    "bhaMdIn",
    "bhaMdOut",
    "bhaTotalLength",
    "safetyIncident",
    "environmentIncident",
    "daysSinceRi",
    "daysSinceLta",
    "incidentComments",
]

COMPLETION_FIELD_COLUMNS = COMMON_FIELD_COLUMNS + [
    "description",
    "operationStartDate",
    "afeCost",
    "dailyCost",
    "cumulativeCost",
    "supervisor1",
    "supervisor2",
    "engineer",
    "pamEngineer",
    "geologist",
    "totalPersonnel",
    "safetyComments",
]

WORKOVER_FIELD_COLUMNS = COMMON_FIELD_COLUMNS + [
    "workoverNo",
    "description",
    "operationStartDate",
    "afeCost",
    "dailyCost",
    "cumulativeCost",
    "supervisor1",
    "supervisor2",
    "engineer",
    "pamEngineer",
    "geologist",
    "totalPersonnel",
    "safetyComments",
]

MOVE_FIELD_COLUMNS = COMMON_FIELD_COLUMNS + [
    "todayMd",
    "prevMd",
    "progress",
    "rotHrsToday",
    "groundElev",
    "afeMdDays",
    "wellborePrefix",
]

ROW_COLUMNS = {
    "operations": ["from", "to", "hours", "op_code", "op_sub", "op_type", "operation_details", "system_op_type", "confirmed_op_type"],
    "bulks": ["bulk", "qty_start", "qty_used", "qty_end"],
    "daily_costs": ["cost_description", "vendor", "amount"],
    "survey_data": ["md", "incl", "azi", "tvd", "vse", "ns", "dls", "build"],
    "bha_components": ["component", "od", "id", "joints", "length"],
    "perforation_intervals": [
        "formation",
        "top_md",
        "base_md",
        "length",
        "density",
        "charges",
        "phase",
        "penetration",
        "diameter",
        "date",
        "status",
        "comments",
    ],
    "mud_products": ["product", "unit", "received", "used", "returned", "ending"],
}

TRANSLATION_CONTENT_COLUMNS = [
    "record_id",
    "entity_type",
    "entity_id",
    "field_code",
    "source_language",
    "target_language",
    "source_text",
    "translated_text",
    "source_hash",
    "model_config_id",
    "prompt_version",
    "translation_status",
    "error_message",
    "is_manual_modified",
    "created_at",
    "updated_at",
]

DEPRECATED_ROW_COLUMNS = {
    "operations": ["confirmation_note"],
}

REPORT_TABLES = {
    "drilling": {
        "field_sheet": "drilling_fields",
        "field_columns": DRILLING_FIELD_COLUMNS,
        "multi": {
            "survey_data": "drilling_survey",
            "bha_components": "drilling_bha",
            "operations": "drilling_operations",
            "daily_costs": "drilling_costs",
            "bulks": "drilling_bulks",
        },
    },
    "completion": {
        "field_sheet": "completion_fields",
        "field_columns": COMPLETION_FIELD_COLUMNS,
        "multi": {
            "operations": "completion_operations",
            "bulks": "completion_bulks",
            "daily_costs": "completion_costs",
            "mud_products": "completion_mud_products",
            "perforation_intervals": "completion_intervals",
        },
    },
    "workover": {
        "field_sheet": "workover_fields",
        "field_columns": WORKOVER_FIELD_COLUMNS,
        "multi": {
            "operations": "workover_operations",
            "bulks": "workover_bulks",
            "daily_costs": "workover_costs",
            "mud_products": "workover_mud_products",
            "perforation_intervals": "workover_intervals",
        },
    },
    "move": {
        "field_sheet": "move_fields",
        "field_columns": MOVE_FIELD_COLUMNS,
        "multi": {
            "operations": "move_operations",
        },
    },
}


@_with_workbook_lock
def initialize_database(database_path: str | Path) -> Path:
    path = Path(database_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = _load_or_create_workbook(path)
    _ensure_schema(workbook)
    _format_workbook(workbook)
    _save_workbook(workbook, path)
    return path


@_with_workbook_lock
def save_report_payload(
    database_path: str | Path,
    payload: dict[str, Any],
    report_type: str,
    *,
    source_file: str = "",
) -> dict[str, Any]:
    report_type = _normalize_report_type(report_type)
    path = Path(database_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = _load_or_create_workbook(path)
    _ensure_schema(workbook)

    now = _now()
    fields = payload.get("report_fields", {}) or {}
    metadata = payload.get("metadata", {}) or {}
    record_id = str(metadata.get("record_id") or payload.get("record_id") or _natural_record_id(report_type, fields) or _generated_record_id(report_type, now))
    source_file = source_file or str(metadata.get("source_file") or "")

    existing = _record_index(workbook["records"]).get(record_id, {})
    if _truthy(existing.get("locked")):
        raise PermissionError(f"Record is locked after NPT confirmation: {record_id}")
    created_at = existing.get("created_at") or now
    _replace_record_rows(workbook, report_type, record_id)

    _append_row(
        workbook["records"],
        BASE_COLUMNS,
        {
            "record_id": record_id,
            "report_type": report_type,
            "source_file": source_file,
            "parser": metadata.get("parser", ""),
            "reportDate": fields.get("reportDate", ""),
            "reportNo": fields.get("reportNo", ""),
            "wellbore": fields.get("wellbore", ""),
            "rig": fields.get("rig", ""),
            "status": metadata.get("status", "parsed"),
            "source_language": metadata.get("source_language", ""),
            "translation_status": metadata.get("translation_status", ""),
            "translation_progress": metadata.get("translation_progress", ""),
            "translation_error": metadata.get("translation_error", ""),
            "translation_version": metadata.get("translation_version", ""),
            "validation_status": metadata.get("validation_status", "ok"),
            "validation_warnings": metadata.get("validation_warnings", ""),
            "locked": metadata.get("locked", existing.get("locked", "")),
            "confirmation_status": metadata.get("confirmation_status", existing.get("confirmation_status", "")),
            "confirmed_at": metadata.get("confirmed_at", existing.get("confirmed_at", "")),
            "confirmed_by": metadata.get("confirmed_by", existing.get("confirmed_by", "")),
            "confirmation_note": metadata.get("confirmation_note", existing.get("confirmation_note", "")),
            "created_at": created_at,
            "updated_at": now,
        },
    )

    table_info = REPORT_TABLES[report_type]
    _append_row(workbook[table_info["field_sheet"]], table_info["field_columns"], {"record_id": record_id, **fields})
    for module_name, sheet_name in table_info["multi"].items():
        columns = ["record_id", "row_no", *ROW_COLUMNS[module_name]]
        for row_no, row in enumerate(payload.get(module_name, []) or [], start=1):
            _append_row(workbook[sheet_name], columns, {"record_id": record_id, "row_no": row_no, **(row or {})})

    _format_workbook(workbook)
    _save_workbook(workbook, path)
    return {"record_id": record_id, "database_path": str(path), "updated_at": now}


@_with_workbook_lock
def load_report_payload(database_path: str | Path, record_id: str) -> dict[str, Any]:
    path = Path(database_path)
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path)
    record = _record_index(workbook["records"]).get(record_id)
    if not record:
        raise KeyError(record_id)
    report_type = _normalize_report_type(str(record.get("report_type", "")))
    table_info = REPORT_TABLES[report_type]
    fields = _first_matching_row(workbook[table_info["field_sheet"]], record_id)
    payload: dict[str, Any] = {
        "metadata": {
            "record_id": record_id,
            "report_type": report_type,
            "source_file": record.get("source_file", ""),
            "parser": record.get("parser", ""),
            "source_language": record.get("source_language", ""),
            "translation_status": record.get("translation_status", ""),
            "translation_progress": record.get("translation_progress", ""),
            "translation_error": record.get("translation_error", ""),
            "translation_version": record.get("translation_version", ""),
            "locked": record.get("locked", ""),
            "confirmation_status": record.get("confirmation_status", ""),
            "confirmed_at": record.get("confirmed_at", ""),
            "confirmed_by": record.get("confirmed_by", ""),
            "confirmation_note": record.get("confirmation_note", ""),
        },
        "report_fields": fields,
    }
    for module_name, sheet_name in table_info["multi"].items():
        payload[module_name] = _matching_rows(workbook[sheet_name], record_id)
    translation_content = load_translation_content(path, record_id)
    if translation_content:
        payload["translation_content"] = translation_content
    return payload


@_with_workbook_lock
def delete_report_payload(database_path: str | Path, record_id: str) -> bool:
    path = Path(database_path)
    if not path.exists():
        return False
    workbook = load_workbook(path)
    _ensure_schema(workbook)
    record = _record_index(workbook["records"]).get(record_id)
    if not record:
        return False
    report_type = _normalize_report_type(str(record.get("report_type", "")))
    _replace_record_rows(workbook, report_type, record_id)
    _format_workbook(workbook)
    _save_workbook(workbook, path)
    return True


@_with_workbook_lock
def save_translation_content(database_path: str | Path, record_id: str, rows: list[dict[str, Any]]) -> None:
    path = Path(database_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = _load_or_create_workbook(path)
    _ensure_schema(workbook)
    worksheet = workbook["translation_content"]
    _delete_rows_by_record_id(worksheet, record_id)
    for row in rows:
        if not isinstance(row, dict):
            continue
        _append_row(worksheet, TRANSLATION_CONTENT_COLUMNS, {"record_id": record_id, **row})
    _format_workbook(workbook)
    _save_workbook(workbook, path)


@_with_workbook_lock
def load_translation_content(database_path: str | Path, record_id: str) -> list[dict[str, str]]:
    path = Path(database_path)
    if not path.exists():
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "translation_content" not in workbook.sheetnames:
        return []
    headers = _headers(workbook["translation_content"])
    rows: list[dict[str, str]] = []
    for raw in workbook["translation_content"].iter_rows(min_row=2, values_only=True):
        row = {headers[index]: _cell_value(value) for index, value in enumerate(raw) if index < len(headers)}
        if row.get("record_id") == record_id:
            rows.append(row)
    return rows


@_with_workbook_lock
def clear_translation_content(database_path: str | Path, record_id: str = "") -> None:
    path = Path(database_path)
    if not path.exists():
        return
    workbook = load_workbook(path)
    _ensure_schema(workbook)
    worksheet = workbook["translation_content"]
    if record_id:
        _delete_rows_by_record_id(worksheet, record_id)
    else:
        worksheet.delete_rows(2, max(0, worksheet.max_row - 1))
    _format_workbook(workbook)
    _save_workbook(workbook, path)


@_with_workbook_lock
def update_record_translation_status(
    database_path: str | Path,
    record_id: str,
    *,
    status: str,
    progress: int | str = "",
    error: str = "",
    version: str = "",
) -> None:
    path = Path(database_path)
    if not path.exists() or not record_id:
        return
    workbook = load_workbook(path)
    _ensure_schema(workbook)
    worksheet = workbook["records"]
    headers = _headers(worksheet)
    now = _now()
    for row_index in range(2, worksheet.max_row + 1):
        if str(worksheet.cell(row=row_index, column=headers.index("record_id") + 1).value or "") != record_id:
            continue
        worksheet.cell(row=row_index, column=headers.index("translation_status") + 1, value=str(status or ""))
        worksheet.cell(row=row_index, column=headers.index("translation_progress") + 1, value=str(progress if progress != "" else ""))
        worksheet.cell(row=row_index, column=headers.index("translation_error") + 1, value=str(error or "")[:500])
        if version:
            worksheet.cell(row=row_index, column=headers.index("translation_version") + 1, value=str(version))
        worksheet.cell(row=row_index, column=headers.index("updated_at") + 1, value=now)
        _format_workbook(workbook)
        _save_workbook(workbook, path)
        return


@_with_workbook_lock
def list_records(database_path: str | Path) -> list[dict[str, str]]:
    path = Path(database_path)
    if not path.exists():
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    records = list(_record_index(workbook["records"]).values())
    for record in records:
        report_type = str(record.get("report_type", "") or "")
        record_id = str(record.get("record_id", "") or "")
        if not report_type or report_type not in REPORT_TABLES or not record_id:
            continue
        fields = _first_matching_row(workbook[REPORT_TABLES[report_type]["field_sheet"]], record_id)
        record["afeNumber"] = fields.get("afeNumber", "")
        record["event"] = fields.get("event", "")
    records.sort(key=lambda record: (record.get("reportDate", ""), record.get("updated_at", "")), reverse=True)
    return records


@_with_workbook_lock
def list_npt_confirmation_wells(database_path: str | Path, *, rig: str = "", wellbore: str = "", status: str = "", scope_rig: str = "") -> dict[str, Any]:
    path = Path(database_path)
    if not path.exists():
        return {"items": [], "filters": {"rigs": [], "statuses": _npt_statuses()}}
    workbook = load_workbook(path)
    _ensure_schema(workbook)
    records = list(_record_index(workbook["records"]).values())
    groups: dict[tuple[str, str], dict[str, Any]] = {}
    for record in records:
        record_rig = str(record.get("rig", "") or "")
        record_well = str(record.get("wellbore", "") or "")
        if not record_well:
            continue
        if scope_rig and record_rig != scope_rig:
            continue
        if rig and record_rig != rig:
            continue
        if wellbore and wellbore.lower() not in record_well.lower():
            continue
        record_id = str(record.get("record_id", "") or "")
        report_type = str(record.get("report_type", "") or "")
        try:
            operations = _operation_rows_for_record(workbook, report_type, record_id)
        except (KeyError, ValueError):
            continue
        has_non_p = any(_system_type(row) in {"SC", "NPT"} for row in operations)
        if not has_non_p:
            continue
        key = (record_rig, record_well)
        item = groups.setdefault(key, {
            "rig": record_rig,
            "wellbore": record_well,
            "start_date": record.get("reportDate", ""),
            "end_date": record.get("reportDate", ""),
            "record_ids": [],
            "statuses": [],
            "locked_count": 0,
            "row_count": 0,
            "sc_hours": 0.0,
            "npt_hours": 0.0,
        })
        date_value = str(record.get("reportDate", "") or "")
        if date_value:
            item["start_date"] = min(str(item["start_date"] or date_value), date_value)
            item["end_date"] = max(str(item["end_date"] or date_value), date_value)
        item["record_ids"].append(record_id)
        item["statuses"].append(str(record.get("confirmation_status", "") or "pending"))
        if _truthy(record.get("locked")):
            item["locked_count"] += 1
        for row in operations:
            op_type = _system_type(row)
            hours = _safe_float(row.get("hours"))
            item["row_count"] += 1
            if op_type == "SC":
                item["sc_hours"] += hours
            elif op_type == "NPT":
                item["npt_hours"] += hours
    items = []
    for item in groups.values():
        item_status = _confirmation_group_status(item)
        if status and item_status != status:
            continue
        items.append({
            "wellbore": item["wellbore"],
            "rig": item["rig"],
            "start_date": item["start_date"],
            "end_date": item["end_date"],
            "status": item_status,
            "record_count": len(item["record_ids"]),
            "row_count": item["row_count"],
            "sc_hours": round(float(item["sc_hours"]), 2),
            "npt_hours": round(float(item["npt_hours"]), 2),
        })
    items.sort(key=lambda row: (str(row["status"] != "pending"), str(row["end_date"])), reverse=False)
    rigs = sorted({str(record.get("rig", "") or "") for record in records if record.get("rig")})
    return {"items": items, "filters": {"rigs": rigs, "statuses": _npt_statuses()}}


@_with_workbook_lock
def load_npt_confirmation_detail(database_path: str | Path, wellbore: str, *, rig: str = "", scope_rig: str = "") -> dict[str, Any]:
    path = Path(database_path)
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path)
    _ensure_schema(workbook)
    records = [
        record for record in _record_index(workbook["records"]).values()
        if str(record.get("wellbore", "") or "") == wellbore
        and (not rig or str(record.get("rig", "") or "") == rig)
        and (not scope_rig or str(record.get("rig", "") or "") == scope_rig)
    ]
    if not records:
        raise KeyError(wellbore)
    relevant_ids = set()
    for record in records:
        operations = _operation_rows_for_record(workbook, str(record.get("report_type", "") or ""), str(record.get("record_id", "") or ""))
        if any(_system_type(row) in {"SC", "NPT"} for row in operations):
            relevant_ids.add(str(record.get("record_id", "") or ""))
    records = [record for record in records if str(record.get("record_id", "") or "") in relevant_ids]
    if not records:
        raise KeyError(wellbore)
    rows: list[dict[str, Any]] = []
    for record in sorted(records, key=lambda item: str(item.get("reportDate", "") or "")):
        operations = _operation_rows_for_record(workbook, str(record.get("report_type", "") or ""), str(record.get("record_id", "") or ""))
        for row in operations:
            system_type = _system_type(row)
            confirmed_type = str(row.get("confirmed_op_type", "") or row.get("op_type", "") or system_type).strip().upper()
            rows.append({
                "record_id": record.get("record_id", ""),
                "report_type": record.get("report_type", ""),
                "reportDate": record.get("reportDate", ""),
                "row_no": row.get("row_no", ""),
                "from": row.get("from", ""),
                "to": row.get("to", ""),
                "hours": row.get("hours", ""),
                "op_code": row.get("op_code", ""),
                "op_sub": row.get("op_sub", ""),
                "operation_details": row.get("operation_details", ""),
                "system_op_type": system_type,
                "confirmed_op_type": confirmed_type,
            })
    start_date = min(str(record.get("reportDate", "") or "") for record in records if record.get("reportDate"))
    end_date = max(str(record.get("reportDate", "") or "") for record in records if record.get("reportDate"))
    meta = {
        "wellbore": wellbore,
        "rig": rig or str(records[0].get("rig", "") or ""),
        "start_date": start_date,
        "end_date": end_date,
        "status": _confirmation_group_status({
            "record_ids": [record.get("record_id", "") for record in records],
            "statuses": [str(record.get("confirmation_status", "") or "pending") for record in records],
            "locked_count": sum(1 for record in records if _truthy(record.get("locked"))),
        }),
        "record_count": len(records),
        "locked": all(_truthy(record.get("locked")) for record in records),
        "confirmation_note": next((str(record.get("confirmation_note", "") or "") for record in records if record.get("confirmation_note")), ""),
    }
    return {"meta": meta, "operations": rows}


@_with_workbook_lock
def save_npt_confirmation(
    database_path: str | Path,
    wellbore: str,
    operations: list[dict[str, Any]],
    *,
    rig: str = "",
    note: str = "",
    confirmed_by: str = "",
    submit: bool = False,
) -> dict[str, Any]:
    path = Path(database_path)
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path)
    _ensure_schema(workbook)
    records_sheet = workbook["records"]
    records = _record_index(records_sheet)
    detail = load_npt_confirmation_detail(path, wellbore, rig=rig)
    if detail["meta"].get("locked"):
        raise PermissionError(f"Well is locked after NPT confirmation: {wellbore}")
    allowed_record_ids = {str(row.get("record_id", "") or "") for row in detail["operations"]}
    updates: dict[tuple[str, int], dict[str, str]] = {}
    for row in operations:
        record_id = str(row.get("record_id", "") or "")
        if record_id not in allowed_record_ids:
            continue
        try:
            row_no = int(str(row.get("row_no", "") or "0"))
        except ValueError:
            row_no = 0
        confirmed_type = str(row.get("confirmed_op_type", "") or "").strip().upper()
        if confirmed_type not in {"P", "SC", "NPT"} or row_no <= 0:
            continue
        updates[(record_id, row_no)] = {
            "confirmed_op_type": confirmed_type,
        }
    if not updates:
        raise ValueError("No valid NPT confirmation rows.")
    touched_ids: set[str] = set()
    for record_id, row_no in updates:
        report_type = str(records.get(record_id, {}).get("report_type", "") or "")
        sheet_name = REPORT_TABLES[_normalize_report_type(report_type)]["multi"]["operations"]
        worksheet = workbook[sheet_name]
        headers = _headers(worksheet)
        for row_index in range(2, worksheet.max_row + 1):
            if str(worksheet.cell(row=row_index, column=headers.index("record_id") + 1).value or "") != record_id:
                continue
            if int(str(worksheet.cell(row=row_index, column=headers.index("row_no") + 1).value or "0")) != row_no:
                continue
            current_type = str(worksheet.cell(row=row_index, column=headers.index("op_type") + 1).value or "").strip().upper()
            system_col = headers.index("system_op_type") + 1
            if not str(worksheet.cell(row=row_index, column=system_col).value or "").strip():
                worksheet.cell(row=row_index, column=system_col, value=current_type)
            worksheet.cell(row=row_index, column=headers.index("confirmed_op_type") + 1, value=updates[(record_id, row_no)]["confirmed_op_type"])
            if submit:
                worksheet.cell(row=row_index, column=headers.index("op_type") + 1, value=updates[(record_id, row_no)]["confirmed_op_type"])
            touched_ids.add(record_id)
            break
    now = _now()
    record_headers = _headers(records_sheet)
    for row_index in range(2, records_sheet.max_row + 1):
        record_id = str(records_sheet.cell(row=row_index, column=record_headers.index("record_id") + 1).value or "")
        if record_id not in allowed_record_ids:
            continue
        records_sheet.cell(row=row_index, column=record_headers.index("confirmation_status") + 1, value="confirmed" if submit else "draft")
        records_sheet.cell(row=row_index, column=record_headers.index("confirmation_note") + 1, value=note)
        records_sheet.cell(row=row_index, column=record_headers.index("updated_at") + 1, value=now)
        if submit:
            records_sheet.cell(row=row_index, column=record_headers.index("locked") + 1, value="yes")
            records_sheet.cell(row=row_index, column=record_headers.index("confirmed_at") + 1, value=now)
            records_sheet.cell(row=row_index, column=record_headers.index("confirmed_by") + 1, value=confirmed_by)
    _format_workbook(workbook)
    _save_workbook(workbook, path)
    return {"wellbore": wellbore, "updated_records": len(touched_ids), "status": "confirmed" if submit else "draft", "locked": submit, "updated_at": now}


def _save_workbook(workbook: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f".{path.stem}.{os.getpid()}.{threading.get_ident()}.xlsx")
    try:
        workbook.save(tmp_path)
        tmp_path.replace(path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def _load_or_create_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    workbook = Workbook()
    workbook.active.title = "records"
    return workbook


def _operation_rows_for_record(workbook: Workbook, report_type: str, record_id: str) -> list[dict[str, str]]:
    table_info = REPORT_TABLES[_normalize_report_type(report_type)]
    sheet_name = table_info["multi"].get("operations")
    if not sheet_name:
        return []
    return _matching_rows(workbook[sheet_name], record_id, keep_row_no=True)


def _truthy(value: Any) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "locked", "confirmed"}


def _system_type(row: dict[str, Any]) -> str:
    return str(row.get("system_op_type", "") or row.get("op_type", "") or "").strip().upper()


def _ensure_schema(workbook: Workbook) -> None:
    _ensure_sheet(workbook, "records", BASE_COLUMNS)
    if "record_translations" in workbook.sheetnames:
        del workbook["record_translations"]
    _ensure_sheet(workbook, "translation_content", TRANSLATION_CONTENT_COLUMNS)
    for report_type, table_info in REPORT_TABLES.items():
        _ensure_sheet(workbook, table_info["field_sheet"], table_info["field_columns"])
        for module_name, sheet_name in table_info["multi"].items():
            worksheet = _ensure_sheet(workbook, sheet_name, ["record_id", "row_no", *ROW_COLUMNS[module_name]])
            _drop_columns(worksheet, DEPRECATED_ROW_COLUMNS.get(module_name, []))


def _ensure_sheet(workbook: Workbook, name: str, columns: list[str]) -> Worksheet:
    worksheet = workbook[name] if name in workbook.sheetnames else workbook.create_sheet(name)
    if worksheet.max_row == 1 and all(cell.value is None for cell in worksheet[1]):
        for index, column in enumerate(columns, start=1):
            worksheet.cell(row=1, column=index, value=column)
    else:
        existing = [cell.value for cell in worksheet[1]]
        for column in columns:
            if column not in existing:
                worksheet.cell(row=1, column=len(existing) + 1, value=column)
                existing.append(column)
    return worksheet


def _drop_columns(worksheet: Worksheet, columns: list[str]) -> None:
    if not columns:
        return
    while True:
        headers = _headers(worksheet)
        removable = [column for column in columns if column in headers]
        if not removable:
            break
        column_index = headers.index(removable[0]) + 1
        worksheet.delete_cols(column_index)


def _replace_record_rows(workbook: Workbook, report_type: str, record_id: str) -> None:
    sheets = ["records", "translation_content", REPORT_TABLES[report_type]["field_sheet"], *REPORT_TABLES[report_type]["multi"].values()]
    for sheet_name in sheets:
        worksheet = workbook[sheet_name]
        _delete_rows_by_record_id(worksheet, record_id)


def _delete_rows_by_record_id(worksheet: Worksheet, record_id: str) -> None:
    for row_number in range(worksheet.max_row, 1, -1):
        if str(worksheet.cell(row=row_number, column=1).value or "") == record_id:
            worksheet.delete_rows(row_number)


def _append_row(worksheet: Worksheet, columns: list[str], values: dict[str, Any]) -> None:
    headers = _headers(worksheet)
    for column in columns:
        if column not in headers:
            worksheet.cell(row=1, column=len(headers) + 1, value=column)
            headers.append(column)
    worksheet.append([_cell_value(values.get(column, "")) for column in headers])


def _cell_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _record_index(worksheet: Worksheet) -> dict[str, dict[str, str]]:
    headers = _headers(worksheet)
    rows: dict[str, dict[str, str]] = {}
    for raw in worksheet.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: _cell_value(value) for index, value in enumerate(raw) if index < len(headers)}
        record_id = row.get("record_id", "")
        if record_id:
            rows[record_id] = row
    return rows


def _first_matching_row(worksheet: Worksheet, record_id: str) -> dict[str, str]:
    rows = _matching_rows(worksheet, record_id)
    if not rows:
        return {}
    rows[0].pop("row_no", None)
    return rows[0]


def _matching_rows(worksheet: Worksheet, record_id: str, *, keep_row_no: bool = False) -> list[dict[str, str]]:
    headers = _headers(worksheet)
    rows: list[dict[str, str]] = []
    for raw in worksheet.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: _cell_value(value) for index, value in enumerate(raw) if index < len(headers)}
        if row.get("record_id") == record_id:
            row.pop("record_id", None)
            rows.append(row)
    rows.sort(key=lambda item: int(item.get("row_no") or 0))
    if not keep_row_no:
        for row in rows:
            row.pop("row_no", None)
    return rows


def _headers(worksheet: Worksheet) -> list[str]:
    return [_cell_value(cell.value) for cell in worksheet[1]]


def _generated_record_id(report_type: str, timestamp: str) -> str:
    return f"{report_type}:{_slug(timestamp)}"


def _now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _format_workbook(workbook: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        widths: dict[int, int] = {}
        for row in worksheet.iter_rows():
            for cell in row:
                value = _cell_value(cell.value)
                widths[cell.column] = min(max(widths.get(cell.column, 10), len(value) + 2), 48)
        for column, width in widths.items():
            worksheet.column_dimensions[worksheet.cell(row=1, column=column).column_letter].width = width
