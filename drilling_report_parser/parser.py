from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .text_structure import normalize_multiline


MAX_SCAN_COLS = 80


@dataclass(frozen=True)
class CellValue:
    sheet: str
    row: int
    col: int
    value: str

    @property
    def address(self) -> str:
        return f"{self.sheet}!{get_column_letter(self.col)}{self.row}"


@dataclass(frozen=True)
class FieldSpec:
    output: str
    labels: tuple[str, ...]


@dataclass(frozen=True)
class TableSpec:
    name: str
    sheet_name: str
    headers: tuple[tuple[str, tuple[str, ...]], ...]
    min_header_hits: int
    stop_markers: tuple[str, ...] = ()
    continuation_columns: tuple[str, ...] = ()
    first_column_pattern: str | None = None


@dataclass
class ParseResult:
    source_file: str
    metadata: dict[str, Any] = field(default_factory=dict)
    fields: list[dict[str, Any]] = field(default_factory=list)
    tables: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
    raw_cells: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = text.replace("\n", " ")
    text = re.sub(r"[^a-z0-9%/+.-]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


FIELD_SPECS: tuple[FieldSpec, ...] = (
    FieldSpec("Event", ("Event",)),
    FieldSpec("Date", ("Date", "Fecha")),
    FieldSpec("Report No", ("Report No", "Reporte No")),
    FieldSpec("Country/Area", ("ECU", "Country")),
    FieldSpec("Company", ("EP PETROECUADOR", "Company")),
    FieldSpec("Wellbore", ("Wellbore", "Pozo")),
    FieldSpec("Primary Reason", ("Prim. Reason", "Primary Reason")),
    FieldSpec("Rig", ("Rig", "Taladro")),
    FieldSpec("Current Ops", ("Current Ops", "Operacion Actual")),
    FieldSpec("24-Hr Summary", ("24-Hr Summary", "24 Hr Summary", "Resumen 24")),
    FieldSpec("24-Hr Forecast", ("24-Hr Forecast", "24 Hr Forecast", "Pronostico 24")),
    FieldSpec("Today's MD", ("Today's MD", "Today MD")),
    FieldSpec("Previous MD", ("Prev MD", "Previous MD")),
    FieldSpec("Progress", ("Progress", "Avance")),
    FieldSpec("Rot Hrs Today", ("Rot Hrs Today",)),
    FieldSpec("Avg ROP Slide", ("Avg ROP Slide",)),
    FieldSpec("Avg ROP Rot", ("Avg ROP Rot",)),
    FieldSpec("Reference Datum", ("Ref Datum",)),
    FieldSpec("Ground Elevation", ("Ground Elev",)),
    FieldSpec("AFE MD/Days", ("AFE MD/Days",)),
    FieldSpec("AFE Number", ("AFE Number",)),
    FieldSpec("DFS", ("DFS",)),
    FieldSpec("Daily Cost", ("Daily Cost",)),
    FieldSpec("Cumulative Cost", ("Cum Cost",)),
    FieldSpec("AFE Cost", ("AFE Cost",)),
    FieldSpec("Mud Gas Backgr", ("Backgr", "MUD GAS Backgr")),
    FieldSpec("Mud Gas Trip", ("Trip", "MUD GAS Trip")),
    FieldSpec("Mud Gas Conn", ("Conn", "MUD GAS Conn")),
    FieldSpec("Last Casing", ("Last Casing",)),
    FieldSpec("Next Casing", ("Next Casing",)),
    FieldSpec("Last BOP Press Test", ("Last BOP Press Test",)),
    FieldSpec("Formation Test/EMW", ("Form Test/EMW",)),
    FieldSpec("String Weight Up/Down", ("Str Wt Up/Dn",)),
    FieldSpec("String Weight Rotating", ("Str Wt Rot",)),
    FieldSpec("Torque Off Bottom", ("Torq Off Btm",)),
    FieldSpec("Torque On Bottom", ("Torq On Btm",)),
    FieldSpec("Pump Rate", ("Pump Rate",)),
    FieldSpec("Pump Pressure", ("Pump Press",)),
    FieldSpec("Supervisor 1", ("Supervisor 1",)),
    FieldSpec("Supervisor 2", ("Supervisor 2",)),
    FieldSpec("Engineer", ("Engineer",)),
    FieldSpec("Geologist", ("Geologist",)),
    FieldSpec("Total on Site", ("Total on Site",)),
    FieldSpec("Engineer PAM", ("Engineer PAM",)),
    FieldSpec("Mud Rheology Temp", ("Rheology Temp",)),
    FieldSpec("Mud Engineer", ("MUD DATA Engineer", "Engineer")),
    FieldSpec("Mud Sample From", ("Sample From",)),
    FieldSpec("Mud Type", ("Mud Type",)),
    FieldSpec("Mud Time / MD", ("Time / MD",)),
    FieldSpec("Mud Density @ Temp", ("Density @ Temp",)),
    FieldSpec("Mud Viscosity", ("Viscosity",)),
    FieldSpec("Mud PV / YP", ("PV / YP",)),
    FieldSpec("Mud Gels 10s/10m/30m", ("Gels 10s/10m/30m",)),
    FieldSpec("Mud API WL", ("API WL",)),
    FieldSpec("Mud HTHP WL", ("HTHP WL",)),
    FieldSpec("Mud Cake API / HTHP", ("Cake API / HTHP",)),
    FieldSpec("Mud Solids / Sol Corr", ("Solids / Sol Corr",)),
    FieldSpec("Mud Oil / Water", ("Oil / Water",)),
    FieldSpec("Mud Sand", ("Sand",)),
    FieldSpec("Mud Water Added", ("Water Added",)),
    FieldSpec("Mud Oil Added", ("Oil Added",)),
    FieldSpec("Mud LGS", ("LGS",)),
    FieldSpec("Mud MBT", ("MBT",)),
    FieldSpec("Mud pH", ("pH",)),
    FieldSpec("Mud Pf / Mf", ("Pf / Mf",)),
    FieldSpec("Mud Pm / Pom", ("Pm / Pom",)),
    FieldSpec("Mud Chlorides", ("Chlorides",)),
    FieldSpec("Mud Ca+ / K+", ("Ca+ / K+",)),
    FieldSpec("Mud ECD", ("ECD",)),
    FieldSpec("Mud Comments", ("Comments",)),
    FieldSpec("BHA No", ("BHA No",)),
    FieldSpec("BHA Purpose", ("Purpose",)),
    FieldSpec("BHA MD In", ("MD In",)),
    FieldSpec("BHA MD Out", ("MD Out",)),
    FieldSpec("BHA Total Length", ("Total Length",)),
    FieldSpec("BHA Wt below Jars", ("Wt below Jars",)),
    FieldSpec("Injected Volume", ("Injected Volume",)),
    FieldSpec("Returned Volume", ("Returned Volume",)),
    FieldSpec("Safety Incident", ("Safety Incident",)),
    FieldSpec("Environmental Incident", ("Environ Incident",)),
    FieldSpec("Incident Comments", ("Incident Comments",)),
    FieldSpec("Days since Last RI", ("Days since Last RI",)),
    FieldSpec("Days since Last LTA", ("Days since Last LTA",)),
    FieldSpec("Other Remarks", ("Other Remarks",)),
)


TABLE_SPECS: tuple[TableSpec, ...] = (
    TableSpec(
        name="operations",
        sheet_name="operations",
        headers=(
            ("from", ("from",)),
            ("to", ("to",)),
            ("hours", ("hrs", "hours")),
            ("op_code", ("op code",)),
            ("op_sub", ("op sub",)),
            ("op_type", ("op type",)),
            ("operation_details", ("operation details", "details")),
        ),
        min_header_hits=5,
        stop_markers=("safety incident", "environ incident", "incident comments", "other remarks"),
        continuation_columns=("operation_details",),
        first_column_pattern=r"^\d{1,2}:?\d{0,2}$",
    ),
    TableSpec(
        name="survey_data",
        sheet_name="survey_data",
        headers=(
            ("md", ("md",)),
            ("incl", ("incl", "inclination")),
            ("azi", ("azi", "azimuth")),
            ("tvd", ("tvd",)),
            ("vse", ("vse",)),
            ("ns", ("n/-s", "n s", "ns")),
            ("ew", ("e/-w", "e w", "ew")),
            ("dls", ("dls",)),
            ("build", ("build",)),
        ),
        min_header_hits=5,
        stop_markers=("personnel", "bit record", "mud data"),
    ),
    TableSpec(
        name="bit_record",
        sheet_name="bit_record",
        headers=(
            ("bit_no", ("bit no",)),
            ("size", ("size",)),
            ("manufacturer", ("manufacturer",)),
            ("model", ("model",)),
            ("serial_no", ("serial no",)),
            ("iadc_code", ("iadc code",)),
            ("nozzles", ("nozzles",)),
            ("md_in", ("md in",)),
            ("md_out", ("md out",)),
            ("bit_grade", ("b-g-o-r", "b g o r")),
        ),
        min_header_hits=5,
        stop_markers=("bit operating parameters", "mud data"),
    ),
    TableSpec(
        name="bit_operating_parameters",
        sheet_name="bit_parameters",
        headers=(
            ("bit_no", ("bit no",)),
            ("rot_hrs", ("rot hrs",)),
            ("cum_rot_hrs", ("cum rot hrs",)),
            ("progress", ("prog",)),
            ("cum_progress", ("cum prog",)),
            ("rop", ("rop",)),
            ("cum_rop", ("cum rop",)),
            ("wob_min_max", ("wob",)),
            ("rpm_min_max", ("rpm",)),
            ("tfa", ("tfa",)),
            ("p_drop_bit", ("p drop bit",)),
            ("noz_vel", ("noz vel",)),
            ("hhpsi", ("hhpsi",)),
        ),
        min_header_hits=5,
        stop_markers=("mud data", "mud products"),
    ),
    TableSpec(
        name="mud_products",
        sheet_name="mud_products",
        headers=(
            ("product", ("product",)),
            ("units", ("units",)),
            ("qty_used", ("qty used",)),
        ),
        min_header_hits=3,
        stop_markers=("last or current bha", "daily costs", "fluid losses"),
    ),
    TableSpec(
        name="bha_components",
        sheet_name="bha_components",
        headers=(
            ("component", ("component",)),
            ("od", ("od",)),
            ("id", ("id",)),
            ("joints", ("jts", "joints")),
            ("length", ("length",)),
        ),
        min_header_hits=4,
        stop_markers=("fluid losses", "bulks", "operations", "daily costs"),
        continuation_columns=("component",),
    ),
    TableSpec(
        name="daily_costs",
        sheet_name="daily_costs",
        headers=(
            ("cost_description", ("cost description",)),
            ("vendor", ("vendor",)),
            ("amount", ("amount",)),
        ),
        min_header_hits=3,
        stop_markers=("fluid losses", "bulks", "operations"),
    ),
    TableSpec(
        name="bulks",
        sheet_name="bulks",
        headers=(
            ("bulk", ("bulks", "bulk")),
            ("qty_start", ("qty start",)),
            ("qty_used", ("qty used",)),
            ("qty_end", ("qty end",)),
        ),
        min_header_hits=3,
        stop_markers=("operations", "safety incident", "other remarks"),
    ),
)


def parse_excel_report(input_path: str | Path) -> ParseResult:
    path = Path(input_path)
    workbook = load_workbook(path, data_only=True)
    grids = _read_grids(workbook)

    result = ParseResult(source_file=str(path))
    result.metadata = {
        "source_file": str(path),
        "worksheets": ", ".join(workbook.sheetnames),
        "parser": "drilling_report_parser",
    }
    result.raw_cells = [
        {"sheet": c.sheet, "cell": f"{get_column_letter(c.col)}{c.row}", "row": c.row, "col": c.col, "value": c.value}
        for grid in grids.values()
        for row in grid
        for c in row
        if c.value
    ]

    result.fields = _extract_fields(grids)
    found_field_names = {row["field"] for row in result.fields}
    for spec in FIELD_SPECS:
        if spec.output not in found_field_names:
            result.warnings.append(f"Field not found: {spec.output}")

    for spec in TABLE_SPECS:
        rows = _extract_table(grids, spec)
        result.tables[spec.name] = rows
        if not rows:
            result.warnings.append(f"Table not found or empty: {spec.name}")

    return result


def write_structured_workbook(result: ParseResult, output_path: str | Path) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    _write_rows(
        wb,
        "report_fields",
        result.fields,
        ["field", "value", "source_sheet", "source_cell"],
    )

    for spec in TABLE_SPECS:
        rows = result.tables.get(spec.name, [])
        columns = [header[0] for header in spec.headers] + ["source_sheet", "source_row"]
        _write_rows(wb, spec.sheet_name, rows, columns)

    _write_rows(
        wb,
        "raw_cells",
        result.raw_cells,
        ["sheet", "cell", "row", "col", "value"],
    )
    _write_rows(
        wb,
        "parse_warnings",
        [{"warning": warning} for warning in result.warnings],
        ["warning"],
    )
    _write_rows(
        wb,
        "metadata",
        [{"key": key, "value": value} for key, value in result.metadata.items()],
        ["key", "value"],
    )

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _read_grids(workbook) -> dict[str, list[list[CellValue]]]:
    grids: dict[str, list[list[CellValue]]] = {}
    for sheet in workbook.worksheets:
        merged_values: dict[tuple[int, int], str] = {}
        for merged in sheet.merged_cells.ranges:
            top_left = sheet.cell(merged.min_row, merged.min_col).value
            if top_left is None:
                continue
            for row in range(merged.min_row, merged.max_row + 1):
                for col in range(merged.min_col, merged.max_col + 1):
                    merged_values[(row, col)] = clean_value(top_left)

        max_col = min(sheet.max_column or 1, MAX_SCAN_COLS)
        sheet_grid: list[list[CellValue]] = []
        for row_idx in range(1, (sheet.max_row or 1) + 1):
            row_values: list[CellValue] = []
            for col_idx in range(1, max_col + 1):
                cell: Cell = sheet.cell(row_idx, col_idx)
                value = merged_values.get((row_idx, col_idx), clean_value(cell.value))
                row_values.append(CellValue(sheet.title, row_idx, col_idx, value))
            sheet_grid.append(row_values)
        grids[sheet.title] = sheet_grid
    return grids


def _extract_fields(grids: dict[str, list[list[CellValue]]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    fields: list[dict[str, Any]] = []
    label_norms = {spec.output: [normalize_text(label) for label in spec.labels] for spec in FIELD_SPECS}
    all_labels = {label for labels in label_norms.values() for label in labels}

    for spec in FIELD_SPECS:
        labels = label_norms[spec.output]
        match = _find_field_value(grids, labels, all_labels)
        if match is None:
            continue
        value, source = match
        if not value or spec.output in seen:
            continue
        seen.add(spec.output)
        fields.append(
            {
                "field": spec.output,
                "value": value,
                "source_sheet": source.sheet,
                "source_cell": f"{get_column_letter(source.col)}{source.row}",
            }
        )
    return fields


def _find_field_value(
    grids: dict[str, list[list[CellValue]]],
    labels: Iterable[str],
    all_labels: set[str],
) -> tuple[str, CellValue] | None:
    label_list = tuple(labels)
    for grid in grids.values():
        for row_idx, row in enumerate(grid):
            for col_idx, cell in enumerate(row):
                norm = normalize_text(cell.value)
                if not norm:
                    continue
                matched = next((label for label in label_list if _cell_matches_label(norm, label)), None)
                if not matched:
                    continue

                inline = _inline_value(cell.value, matched, all_labels)
                if inline:
                    return inline, cell

                nearby = _nearest_value(grid, row_idx, col_idx, all_labels)
                if nearby:
                    return nearby.value, nearby
    return None


def _cell_matches_label(cell_norm: str, label_norm: str) -> bool:
    if not label_norm:
        return False
    return (
        cell_norm == label_norm
        or cell_norm.startswith(f"{label_norm} ")
        or cell_norm.endswith(f" {label_norm}")
        or f"{label_norm} :" in cell_norm
        or f"{label_norm}:" in cell_norm
    )


def _inline_value(text: str, label_norm: str, all_labels: set[str]) -> str:
    original = clean_value(text)
    if ":" not in original:
        return ""

    pieces = re.split(r":", original, maxsplit=1)
    if len(pieces) != 2:
        return ""
    left_norm = normalize_text(pieces[0])
    if label_norm not in left_norm and left_norm not in label_norm:
        return ""
    value = pieces[1].strip()
    value_norm = normalize_text(value)
    if not value or value_norm in all_labels or _looks_like_only_labels(value_norm, all_labels):
        return ""
    return value


def _looks_like_only_labels(value_norm: str, all_labels: set[str]) -> bool:
    if not value_norm:
        return True
    return any(value_norm == label or value_norm.endswith(f" {label}") for label in all_labels)


def _nearest_value(
    grid: list[list[CellValue]],
    row_idx: int,
    col_idx: int,
    all_labels: set[str],
) -> CellValue | None:
    row = grid[row_idx]
    for right in range(col_idx + 1, min(len(row), col_idx + 8)):
        candidate = row[right]
        if _is_candidate_value(candidate.value, all_labels):
            return candidate

    for down in range(row_idx + 1, min(len(grid), row_idx + 5)):
        for col in range(col_idx, min(len(grid[down]), col_idx + 3)):
            candidate = grid[down][col]
            if _is_candidate_value(candidate.value, all_labels):
                return candidate
    return None


def _is_candidate_value(value: str, all_labels: set[str]) -> bool:
    norm = normalize_text(value)
    if not norm or norm in all_labels:
        return False
    if len(norm) <= 1 and not norm.isdigit():
        return False
    return not _looks_like_only_labels(norm, all_labels)


def _extract_table(grids: dict[str, list[list[CellValue]]], spec: TableSpec) -> list[dict[str, Any]]:
    collected: list[dict[str, Any]] = []
    for grid in grids.values():
        for row_idx, row in enumerate(grid):
            columns = _detect_header_columns(row, spec)
            if not columns:
                continue
            table_rows = _collect_rows_from_header(grid, row_idx, columns, spec)
            collected.extend(table_rows)
    return _dedupe_rows(collected)


def _detect_header_columns(row: list[CellValue], spec: TableSpec) -> list[tuple[str, int]] | None:
    hits: list[tuple[str, int]] = []
    used_cols: set[int] = set()
    row_text = normalize_text(" ".join(cell.value for cell in row if cell.value))

    for output, aliases in spec.headers:
        aliases_norm = tuple(normalize_text(alias) for alias in aliases)
        found_col: int | None = None
        for idx, cell in enumerate(row):
            cell_norm = normalize_text(cell.value)
            if not cell_norm:
                continue
            if any(alias == cell_norm or alias in cell_norm for alias in aliases_norm):
                found_col = idx
                break
        if found_col is None and any(alias in row_text for alias in aliases_norm):
            found_col = _fallback_column_for_header(output, hits)
        if found_col is not None and found_col not in used_cols:
            used_cols.add(found_col)
            hits.append((output, found_col))

    if len(hits) < spec.min_header_hits:
        return None
    return sorted(hits, key=lambda item: item[1])


def _fallback_column_for_header(output: str, hits: list[tuple[str, int]]) -> int:
    if not hits:
        return 0
    return hits[-1][1] + 1


def _collect_rows_from_header(
    grid: list[list[CellValue]],
    header_row_idx: int,
    columns: list[tuple[str, int]],
    spec: TableSpec,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    blank_streak = 0
    stop_markers = tuple(normalize_text(marker) for marker in spec.stop_markers)

    for row_idx in range(header_row_idx + 1, len(grid)):
        row = grid[row_idx]
        row_text = normalize_text(" ".join(cell.value for cell in row if cell.value))
        if any(marker and marker in row_text for marker in stop_markers):
            break
        if not row_text:
            blank_streak += 1
            if blank_streak >= 3:
                break
            continue
        blank_streak = 0

        parsed = _slice_table_row(row, columns)
        if not any(parsed.values()):
            continue

        if _is_probable_section_header(row_text, spec):
            break

        first_key = columns[0][0]
        if spec.first_column_pattern and not re.search(spec.first_column_pattern, parsed.get(first_key, "")):
            if rows and spec.continuation_columns:
                _append_continuation(rows[-1], parsed, spec.continuation_columns)
            continue

        if rows and _is_continuation_row(parsed, columns, spec):
            _append_continuation(rows[-1], parsed, spec.continuation_columns)
            continue

        parsed["source_sheet"] = row[0].sheet
        parsed["source_row"] = row[0].row
        rows.append(parsed)
    return rows


def _slice_table_row(row: list[CellValue], columns: list[tuple[str, int]]) -> dict[str, str]:
    parsed: dict[str, str] = {}
    for idx, (name, start) in enumerate(columns):
        end = columns[idx + 1][1] if idx + 1 < len(columns) else len(row)
        values = [cell.value for cell in row[start:end] if cell.value]
        combined = "\n".join(values) if name == "operation_details" else " ".join(values)
        parsed[name] = normalize_multiline(combined) if name == "operation_details" else clean_value(combined)
    return parsed


def _is_probable_section_header(row_text: str, spec: TableSpec) -> bool:
    known_sections = (
        "daily operations report",
        "survey data",
        "personnel",
        "bit record",
        "bit operating parameters",
        "mud data",
        "mud products",
        "last or current bha",
        "daily costs",
        "fluid losses",
        "bulks",
        "operations",
        "other remarks",
    )
    header_text = " ".join(normalize_text(alias) for _, aliases in spec.headers for alias in aliases)
    if row_text in known_sections and row_text not in header_text:
        return True
    return False


def _is_continuation_row(
    parsed: dict[str, str],
    columns: list[tuple[str, int]],
    spec: TableSpec,
) -> bool:
    if not spec.continuation_columns:
        return False
    first_key = columns[0][0]
    leading_value = parsed.get(first_key, "")
    non_continuation_values = [
        value
        for key, value in parsed.items()
        if key not in spec.continuation_columns and key not in {"source_sheet", "source_row"}
    ]
    return not leading_value and not any(non_continuation_values)


def _append_continuation(row: dict[str, Any], parsed: dict[str, str], continuation_columns: tuple[str, ...]) -> None:
    continuation_text = " ".join(value for value in parsed.values() if value)
    if not continuation_text:
        return
    target = continuation_columns[-1]
    row[target] = normalize_multiline(f"{row.get(target, '')}\n{continuation_text}")


def _dedupe_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[tuple[str, Any], ...]] = set()
    deduped: list[dict[str, Any]] = []
    for row in rows:
        key = tuple(sorted((k, v) for k, v in row.items() if k not in {"source_sheet", "source_row"}))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def _write_rows(wb: Workbook, sheet_name: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ws = wb.create_sheet(sheet_name[:31])
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_len = max(len(clean_value(cell.value)) for cell in column_cells)
        width = min(max(max_len + 2, 12), 60)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
